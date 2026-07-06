from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import xlwt
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from xlsxwriter.utility import xl_rowcol_to_cell


OUT = Path(__file__).resolve().parent / "FLUJO_DE_CAJA_MUSEIQ.xlsx"
OUT_XLS = OUT.with_suffix(".xls")

YEARS = list(range(1, 11))
YEAR_LABELS = [f"Año {year}" for year in YEARS]

PRICE_ADJUSTMENT = 1.20
ADDITIONAL_SERVICE_RATE = 0.08
ADDITIONAL_SERVICE_COST_RATE = 0.50
INCOME_TAX = 0.295
EQUITY_COST = 0.09
LOAN_RATE = 0.18
LOAN_RATE_STRESS = 0.24
LOAN_SHARE = 0.60
EQUITY_SHARE = 0.40
LOAN_GRACE_YEARS = 2
LOAN_TOTAL_YEARS = 8
INITIAL_INVESTMENT = 300_000.0
TANGIBLE_ASSETS = 118_000.0
INTANGIBLE_ASSETS = 132_000.0
INITIAL_CASH = INITIAL_INVESTMENT - TANGIBLE_ASSETS - INTANGIBLE_ASSETS
SALVAGE_VALUE = 35_000.0
INFLATION = [0.03, 0.03, 0.03, 0.03, 0.035, 0.04, 0.04, 0.04, 0.04, 0.04]
COLLECTION_CASH = 0.40
COLLECTION_CREDIT = 0.60
COLLECTION_CURRENT_CREDIT = 0.80
PURCHASE_CASH = 0.30
PURCHASE_CREDIT = 0.70
PAY_CURRENT_CREDIT = 0.85
SALARY_GROWTH = 0.05

# Stable zero-based rows used by sheets that consume the PLANILLA summary.
PLANILLA_ADMIN_SUMMARY_ROW = 4
PLANILLA_PRODUCTION_SUMMARY_ROW = 5
PLANILLA_SALES_SUMMARY_ROW = 6
PLANILLA_TOTAL_SUMMARY_ROW = 7
PLANILLA_DIRECT_PRODUCTION_ROW = 8
PLANILLA_INDIRECT_PRODUCTION_ROW = 9

# Stable zero-based rows in the expanded COSTOS summary.
COSTOS_IMPL_SUBTOTAL_ROW = 26
COSTOS_RECURRENT_SUBTOTAL_ROW = 39
COSTOS_EXTRA_SUBTOTAL_ROW = 46
COSTOS_DIRECT_TOTAL_ROW = 47
COSTOS_IMPLEMENTED_UNITS_ROW = 51
COSTOS_RECURRENT_UNITS_ROW = 53
COSTOS_CXP_ROW = 62
COSTOS_INDIRECT_TOTAL_ROW = 72
PLAN_INV_EQUITY_TOTAL_ROW = 26
PLAN_INV_FINANCING_TOTAL_ROW = 26

# Stable zero-based rows in the professor-style financial statements.
COST_OF_SALES_TOTAL_ROW = 8
COST_OF_SALES_FIRST_YEAR_COL = 2
INCOME_STATEMENT_SALES_ROW = 5
INCOME_STATEMENT_GROSS_PROFIT_ROW = 7
INCOME_STATEMENT_EBIT_ROW = 12
INCOME_STATEMENT_INTEREST_ROW = 13
INCOME_STATEMENT_TAX_ROW = 16
INCOME_STATEMENT_NET_INCOME_ROW = 17
INCOME_STATEMENT_RESERVE_ROW = 18
INCOME_STATEMENT_DISTRIBUTABLE_ROW = 19
CASH_BUDGET_FINAL_ROW = 31
BALANCE_CASH_ROW = 5
BALANCE_CXC_ROW = 6
BALANCE_CURRENT_ASSETS_ROW = 9
BALANCE_TOTAL_ASSETS_ROW = 17
BALANCE_CXP_ROW = 20
BALANCE_CURRENT_LIABILITIES_ROW = 23
BALANCE_LONG_DEBT_ROW = 25
BALANCE_TOTAL_LIABILITIES_ROW = 27
BALANCE_TOTAL_EQUITY_ROW = 33

# Base commercial path from Chapter 5: 35 named institutions, equivalent to
# less than 10% of the 386 museums identified in the addressable market.
NEW_BASIC = [0, 0, 0, 0, 0, 1, 1, 1, 1, 1]
NEW_STANDARD = [0, 0, 2, 2, 3, 1, 2, 3, 3, 4]
NEW_ADVANCED = [2, 2, 1, 1, 0, 2, 1, 0, 1, 0]


@dataclass(frozen=True)
class Package:
    code: str
    name: str
    base_implementation_price: float
    base_recurring_price: float
    implementation_cost: float
    recurring_cost: float
    coverage: str
    beacons: str


PACKAGES = {
    "B": Package("B", "Basico", 48_000, 14_000, 29_000, 5_000, "3 a 5 salas", "8 a 12"),
    "E": Package("E", "Estandar", 96_000, 29_000, 58_000, 9_500, "6 a 10 salas", "18 a 25"),
    "A": Package("A", "Avanzado", 185_000, 52_000, 111_000, 17_000, "12 a 18 salas", "35 a 45"),
}


IMPLEMENTATION_COST_INPUTS = [
    ("Beacons BLE basados en ESP32", "unidad", {"B": 10, "E": 22, "A": 40}, {"B": 30.0, "E": 30.0, "A": 30.0}, "Cap. 5: BLE estimados y piloto S/30 por unidad"),
    ("Baterias CR2032 para beacons", "unidad", {"B": 10, "E": 22, "A": 40}, {"B": 3.5, "E": 3.5, "A": 3.5}, "Cap. 5: bateria por beacon"),
    ("Soportes y adhesivos para beacons", "unidad", {"B": 10, "E": 22, "A": 40}, {"B": 2.0, "E": 2.0, "A": 2.0}, "Cap. 5: fijacion por beacon"),
    ("Nodo de borde Raspberry Pi 5", "unidad", {"B": 0, "E": 1, "A": 2}, {"B": 600.0, "E": 600.0, "A": 600.0}, "Cap. 5: nodo local referencial S/600"),
    ("Fuente oficial para nodo de borde", "unidad", {"B": 0, "E": 1, "A": 2}, {"B": 75.0, "E": 75.0, "A": 75.0}, "Cap. 5: fuente 27W"),
    ("Carcasa para nodo de borde", "unidad", {"B": 0, "E": 1, "A": 2}, {"B": 30.0, "E": 30.0, "A": 30.0}, "Cap. 5: carcasa protectora"),
    ("Almacenamiento microSD", "unidad", {"B": 0, "E": 1, "A": 2}, {"B": 30.0, "E": 30.0, "A": 30.0}, "Cap. 5: microSD"),
    ("Kit de cableado y conectividad", "kit", {"B": 1, "E": 2, "A": 3}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Cableado, conectores y reserva comercial"),
    ("Señaletica QR y rotulado de zonas", "unidad", {"B": 5, "E": 10, "A": 18}, {"B": 80.0, "E": 80.0, "A": 80.0}, "Cap. 5: señalizacion por sala o zona"),
    ("Diagnostico y levantamiento museografico", "hora", {"B": 20, "E": 40, "A": 80}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Diagnostico de salas, recorridos y necesidades"),
    ("Diseño de cobertura y mapa BLE", "hora", {"B": 15, "E": 30, "A": 60}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Diseño de zonas, ubicacion y reglas de activacion"),
    ("Instalacion fisica de beacons y señalizacion", "hora", {"B": 15, "E": 30, "A": 60}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Montaje en sitio"),
    ("Configuracion de app, backend y panel", "hora", {"B": 40, "E": 80, "A": 150}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Configuracion PWA/app, backend y analitica"),
    ("Estructuracion de corpus curatorial", "hora", {"B": 60, "E": 120, "A": 220}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Limpieza, etiquetado y preparacion para RAG"),
    ("Pruebas funcionales y de aceptacion", "hora", {"B": 20, "E": 40, "A": 80}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Pruebas BLE, voz, contenidos y experiencia"),
    ("Capacitacion inicial al personal", "hora", {"B": 16, "E": 32, "A": 60}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Formacion operativa y curatorial"),
    ("Documentacion y manuales operativos", "hora", {"B": 12, "E": 24, "A": 40}, {"B": 100.0, "E": 100.0, "A": 100.0}, "Manuales, protocolo y entregables"),
    ("Viajes y logistica de despliegue", "proyecto", {"B": 1, "E": 1, "A": 1}, {"B": 2_000.0, "E": 4_000.0, "A": 8_000.0}, "Traslado, viaticos y materiales de campo"),
    ("Gestion del proyecto", "proyecto", {"B": 1, "E": 1, "A": 1}, {"B": 2_220.75, "E": 4_159.40, "A": 8_179.50}, "Coordinacion tecnica e institucional"),
    ("Reserva de garantia y postimplementacion", "proyecto", {"B": 1, "E": 1, "A": 1}, {"B": 1_586.25, "E": 2_971.00, "A": 5_842.50}, "Atencion inicial y garantia"),
    ("Gestion contractual y documentaria", "proyecto", {"B": 1, "E": 1, "A": 1}, {"B": 1_269.00, "E": 2_376.80, "A": 4_674.00}, "Expediente, conformidad y cierre"),
    ("Contingencia tecnica de implementacion", "proyecto", {"B": 1, "E": 1, "A": 1}, {"B": 1_269.00, "E": 2_376.80, "A": 4_674.00}, "Reserva por ajustes de campo"),
]


RECURRING_COST_INPUTS = [
    ("Hosting y almacenamiento cloud", "año", {"B": 1, "E": 1, "A": 1}, {"B": 600.0, "E": 1_200.0, "A": 2_400.0}, "Infraestructura anual"),
    ("Consumo STT y TTS", "año", {"B": 1, "E": 1, "A": 1}, {"B": 300.0, "E": 800.0, "A": 2_000.0}, "Voz segun alcance del paquete"),
    ("Consumo LLM, embeddings y RAG", "año", {"B": 1, "E": 1, "A": 1}, {"B": 300.0, "E": 800.0, "A": 2_000.0}, "Consultas y recuperacion aumentada"),
    ("Monitoreo, respaldos y repositorios", "año", {"B": 1, "E": 1, "A": 1}, {"B": 300.0, "E": 500.0, "A": 800.0}, "Continuidad y trazabilidad"),
    ("Reposicion de baterias y accesorios", "año", {"B": 1, "E": 1, "A": 1}, {"B": 100.0, "E": 220.0, "A": 400.0}, "Reposicion preventiva"),
    ("Soporte tecnico remoto", "hora", {"B": 12, "E": 24, "A": 36}, {"B": 80.0, "E": 80.0, "A": 80.0}, "Bolsa anual de soporte"),
    ("Visitas preventivas y logistica", "año", {"B": 1, "E": 1, "A": 1}, {"B": 600.0, "E": 1_200.0, "A": 2_000.0}, "Revision y calibracion en sitio"),
    ("Actualizacion de contenidos", "hora", {"B": 8, "E": 16, "A": 24}, {"B": 80.0, "E": 80.0, "A": 80.0}, "Ajustes curatoriales menores"),
    ("Reserva de incidentes y SLA", "año", {"B": 1, "E": 1, "A": 1}, {"B": 720.0, "E": 948.0, "A": 1_560.0}, "Atencion de incidentes"),
    ("Contingencia de sobreconsumo", "año", {"B": 1, "E": 1, "A": 1}, {"B": 480.0, "E": 632.0, "A": 1_040.0}, "Voz, IA y almacenamiento"),
]


FIXED_ASSET_DETAIL = [
    ("Equipos de desarrollo y prueba", "Laptop de desarrollo", 4, 7_000.0, 4, 0.01),
    ("Equipos de desarrollo y prueba", "Laptop para pruebas y demostraciones", 2, 5_000.0, 4, 0.01),
    ("Equipos de desarrollo y prueba", "Smartphone Android de prueba", 4, 1_500.0, 4, 0.01),
    ("Equipos de desarrollo y prueba", "Tablet para pruebas de recorrido", 2, 2_000.0, 4, 0.01),
    ("Servidor y red local", "Servidor NAS para RAG y datos", 1, 18_000.0, 5, 0.01),
    ("Servidor y red local", "UPS para servidor y comunicaciones", 2, 2_000.0, 5, 0.01),
    ("Servidor y red local", "Router, switch y puntos de acceso", 4, 1_500.0, 5, 0.01),
    ("Servidor y red local", "Discos de respaldo", 4, 1_000.0, 5, 0.01),
    ("Herramientas de instalacion y medicion", "Kit de beacons para pruebas y calibracion", 20, 300.0, 5, 0.01),
    ("Herramientas de instalacion y medicion", "Multimetro y medidor de red", 4, 750.0, 5, 0.01),
    ("Herramientas de instalacion y medicion", "Kit de herramientas de instalacion", 4, 1_000.0, 5, 0.01),
    ("Herramientas de instalacion y medicion", "Impresora de etiquetas y señalizacion", 1, 2_000.0, 5, 0.01),
    ("Herramientas de instalacion y medicion", "Kit portatil de pruebas de campo", 1, 5_000.0, 5, 0.01),
    ("Mobiliario y oficina", "Escritorio de trabajo", 6, 1_200.0, 10, 0.01),
    ("Mobiliario y oficina", "Silla ergonomica de oficina", 6, 800.0, 10, 0.01),
    ("Mobiliario y oficina", "Mesa de reuniones", 1, 2_000.0, 10, 0.01),
    ("Mobiliario y oficina", "Silla para reuniones", 6, 500.0, 10, 0.01),
    ("Mobiliario y oficina", "Archivador y gabinete de seguridad", 2, 500.0, 10, 0.01),
]


def package_input_cost(cost_input: tuple, package_key: str) -> float:
    return cost_input[2][package_key] * cost_input[3][package_key]


def money(value: float) -> float:
    return round(float(value), 2)


def cached_value(value: float, kind: str) -> float:
    if kind == "integer":
        return int(round(float(value)))
    if kind == "money":
        return money(value)
    return float(value)


def cell(row: int, col: int, *, abs_row: bool = False, abs_col: bool = False) -> str:
    return xl_rowcol_to_cell(row, col, row_abs=abs_row, col_abs=abs_col)


def sheet_cell(sheet: str, row: int, col: int, *, abs_row: bool = False, abs_col: bool = False) -> str:
    return f"'{sheet}'!{cell(row, col, abs_row=abs_row, abs_col=abs_col)}"


def same_row_formula(operator: str, rows: Iterable[int], col: int) -> str:
    refs = [cell(row, col) for row in rows]
    return f"={operator}({','.join(refs)})"


def safe_div(num: float, den: float) -> float:
    return 0.0 if abs(den) < 1e-9 else num / den


def npv(rate: float, flows: list[float]) -> float:
    return sum(flow / ((1 + rate) ** i) for i, flow in enumerate(flows))


def irr(flows: list[float]) -> float:
    low, high = -0.95, 5.0
    previous_rate = low
    previous_value = npv(previous_rate, flows)
    for i in range(1, 30000):
        rate = low + (high - low) * i / 30000
        value = npv(rate, flows)
        if previous_value == 0 or previous_value * value < 0:
            a, b = previous_rate, rate
            for _ in range(100):
                mid = (a + b) / 2
                if npv(a, flows) * npv(mid, flows) <= 0:
                    b = mid
                else:
                    a = mid
            return (a + b) / 2
        previous_rate, previous_value = rate, value
    return float("nan")


def discounted_payback(flows: list[float], rate: float) -> float:
    cumulative = flows[0]
    for index in range(1, len(flows)):
        discounted = flows[index] / ((1 + rate) ** index)
        previous = cumulative
        cumulative += discounted
        if cumulative >= 0:
            return (index - 1) + abs(previous) / discounted
    return float("nan")


def inflation_factors() -> list[float]:
    factors = [1.0]
    for i in range(1, len(YEARS)):
        factors.append(factors[-1] * (1 + INFLATION[i - 1]))
    return factors


def labor_total(monthly_salary: float) -> dict[str, float]:
    annual_salary = monthly_salary * 12
    gratification = monthly_salary * 2
    essalud = annual_salary * 0.09
    cts = (annual_salary + gratification) * 0.0833
    vacations = monthly_salary * 0.50
    severance = monthly_salary * 14 / 12
    social = gratification + essalud + cts + vacations + severance
    return {
        "monthly": monthly_salary,
        "annual": annual_salary,
        "gratification": gratification,
        "essalud": essalud,
        "cts": cts,
        "vacations": vacations,
        "notice": 0.0,
        "severance": severance,
        "social": social,
        "total": annual_salary + social,
    }


def staff_plan(year_index: int) -> dict[str, list[tuple[str, float]]]:
    growth = (1 + SALARY_GROWTH) ** year_index
    admin = [
        ("Gerente general - financiero", 3_750 * growth),
        ("Asistente contable administrativo", 800 * growth),
    ]
    if year_index >= 4:
        admin.append(
            (
                "Coordinador administrativo y contratos",
                850 * ((1 + SALARY_GROWTH) ** (year_index - 4)),
            )
        )

    production = [
        ("Lider backend, RAG y plataforma", 3_000 * growth),
        ("Especialista curatorial y datos", 2_250 * growth),
    ]
    if year_index >= 1:
        production.append(
            (
                "Ingeniero mobile e IoT BLE",
                2_500 * ((1 + SALARY_GROWTH) ** (year_index - 1)),
            )
        )
    if year_index >= 2:
        production.append(
            (
                "Soporte tecnico y QA",
                850 * ((1 + SALARY_GROWTH) ** (year_index - 2)),
            )
        )
    if year_index >= 5:
        production.append(
            (
                "Implementador de campo",
                850 * ((1 + SALARY_GROWTH) ** (year_index - 5)),
            )
        )
    if year_index >= 7:
        production.append(
            (
                "Analista de datos y soporte IA",
                2_500 * ((1 + SALARY_GROWTH) ** (year_index - 7)),
            )
        )

    sales = [("Ejecutivo comercial institucional", 850 * growth)]
    if year_index >= 3:
        sales.append(
            (
                "Gestor de relaciones institucionales",
                850 * ((1 + SALARY_GROWTH) ** (year_index - 3)),
            )
        )
    if year_index >= 7:
        sales.append(
            (
                "Coordinador postventa y alianzas",
                850 * ((1 + SALARY_GROWTH) ** (year_index - 7)),
            )
        )

    return {"admin": admin, "production": production, "sales": sales}


def staff_totals(admin_factor: float = 1.0, cost_factor: float = 1.0) -> dict[str, list[float]]:
    totals = {"admin": [], "production": [], "sales": []}
    for i in range(10):
        plan = staff_plan(i)
        totals["admin"].append(sum(labor_total(salary)["total"] for _, salary in plan["admin"]) * admin_factor * cost_factor)
        totals["production"].append(sum(labor_total(salary)["total"] for _, salary in plan["production"]) * cost_factor)
        totals["sales"].append(sum(labor_total(salary)["total"] for _, salary in plan["sales"]) * cost_factor)
    return totals


def production_labor_split(cost_factor: float = 1.0) -> tuple[list[float], list[float]]:
    direct, indirect = [], []
    for i in range(10):
        production = staff_plan(i)["production"]
        indirect.append(
            sum(
                labor_total(salary)["total"]
                for role, salary in production
                if role == "Lider backend, RAG y plataforma"
            )
            * cost_factor
        )
        direct.append(
            sum(
                labor_total(salary)["total"]
                for role, salary in production
                if role != "Lider backend, RAG y plataforma"
            )
            * cost_factor
        )
    return direct, indirect


def debt_schedule(loan: float, annual_rate: float) -> dict[str, list[float]]:
    balance = loan
    amort_years = LOAN_TOTAL_YEARS - LOAN_GRACE_YEARS
    payment = loan * annual_rate / (1 - (1 + annual_rate) ** (-amort_years))
    interest, principal, installment, final_balance, initial_balance = [], [], [], [], []
    for i in range(10):
        initial_balance.append(balance)
        if i < LOAN_GRACE_YEARS:
            year_interest = balance * annual_rate
            year_principal = 0.0
        elif i < LOAN_TOTAL_YEARS:
            year_interest = balance * annual_rate
            year_principal = min(payment - year_interest, balance)
            balance -= year_principal
            if balance < 1e-6:
                balance = 0.0
        else:
            year_interest = 0.0
            year_principal = 0.0
        interest.append(year_interest)
        principal.append(year_principal)
        installment.append(year_interest + year_principal)
        final_balance.append(balance)
    return {
        "initial_balance": initial_balance,
        "interest": interest,
        "principal": principal,
        "installment": installment,
        "final_balance": final_balance,
    }


def asset_schedules() -> dict[str, list[float] | float]:
    fixed_assets = [
        ("Laptops y equipos de prueba", 48_000.0, 4),
        ("Servidor/NAS y red local", 32_000.0, 5),
        ("Herramientas de instalacion y medicion", 20_000.0, 5),
        ("Mobiliario y oficina", 18_000.0, 10),
    ]
    intangibles = [
        ("Constitucion y formalizacion", 8_000.0, 5),
        ("Desarrollo base app/API/RAG", 90_000.0, 5),
        ("Documentacion y manuales", 22_000.0, 5),
        ("Marca, preventa y dossier comercial", 12_000.0, 5),
    ]

    depreciation = [0.0] * 10
    amortization = [0.0] * 10
    fixed_net = []
    intangible_net = []
    fixed_accum = 0.0
    intangible_accum = 0.0
    for i in range(10):
        depreciation[i] = sum(
            quantity * unit_cost * (1 - residual_rate) / life
            for _, _, quantity, unit_cost, life, residual_rate in FIXED_ASSET_DETAIL
            if i < life
        )
        amortization[i] = sum(value / life for _, value, life in intangibles if i < life)
        fixed_accum += depreciation[i]
        intangible_accum += amortization[i]
        fixed_net.append(max(0.0, TANGIBLE_ASSETS - fixed_accum))
        intangible_net.append(max(0.0, INTANGIBLE_ASSETS - intangible_accum))
    return {
        "fixed_assets": fixed_assets,
        "fixed_asset_detail": FIXED_ASSET_DETAIL,
        "intangibles": intangibles,
        "depreciation": depreciation,
        "amortization": amortization,
        "fixed_net": fixed_net,
        "intangible_net": intangible_net,
        "fixed_total": TANGIBLE_ASSETS,
        "intangible_total": INTANGIBLE_ASSETS,
    }


def build_model(
    *,
    name: str = "Base",
    revenue_factor: float = 1.0,
    cost_factor: float = 1.0,
    price_adjustment: float = PRICE_ADJUSTMENT,
    interest_rate: float = LOAN_RATE,
    admin_factor: float = 1.0,
    bad_debt_rate: float = 0.0,
) -> dict[str, object]:
    factors = inflation_factors()
    staff = staff_totals(admin_factor=admin_factor, cost_factor=cost_factor)
    production_direct_labor, production_indirect_labor = production_labor_split(
        cost_factor=cost_factor
    )
    assets = asset_schedules()
    loan = INITIAL_INVESTMENT * LOAN_SHARE
    equity = INITIAL_INVESTMENT * EQUITY_SHARE
    debt = debt_schedule(loan, interest_rate)

    new = {"B": NEW_BASIC[:], "E": NEW_STANDARD[:], "A": NEW_ADVANCED[:]}
    cumulative = {"B": 0, "E": 0, "A": 0}
    active_end = {key: [] for key in new}
    effective = {key: [] for key in new}
    total_new, total_active = [], []

    sales_impl, sales_rec, sales_extra, sales_total = [], [], [], []
    direct_impl_cost, direct_rec_cost, direct_extra_cost, direct_total = [], [], [], []
    cash_sales, credit_sales, current_credit_collection = [], [], []
    previous_cxc_collection, cxc_end, cash_income = [], [], []
    purchase_cash, purchase_credit, current_supplier_payment = [], [], []
    previous_supplier_payment, cxp_end, supplier_cash_payment = [], [], []
    production_labor, admin_labor, sales_labor = [], [], []
    admin_other, sales_other, sales_commission = [], [], []
    energy, platform_tools, supplies, maintenance, indirect_cash, indirect_total = [], [], [], [], [], []
    cost_of_production, cost_of_sales = [], []
    operating_expenses, ebit, ebt, income_tax = [], [], [], []
    cash_tax_paid, cash_egress, cash_budget_flow = [], [], []
    reserve_legal, reserve_accum, retained_profit = [], [], []
    current_debt = []

    previous_cxc = 0.0
    previous_cxp = 0.0
    reserve_limit = equity * 0.20
    reserve_acc = 0.0

    for i in range(10):
        year_new = 0
        year_active = 0
        for key in new:
            effective[key].append(cumulative[key] + 0.5 * new[key][i])
            cumulative[key] += new[key][i]
            active_end[key].append(cumulative[key])
            year_new += new[key][i]
            year_active += cumulative[key]
        total_new.append(year_new)
        total_active.append(year_active)

        impl = sum(
            new[key][i] * PACKAGES[key].base_implementation_price * price_adjustment * factors[i]
            for key in new
        ) * revenue_factor
        rec = sum(
            effective[key][i] * PACKAGES[key].base_recurring_price * price_adjustment * factors[i]
            for key in new
        ) * revenue_factor
        extra = impl * ADDITIONAL_SERVICE_RATE
        total_sales_year = impl + rec + extra
        sales_impl.append(impl)
        sales_rec.append(rec)
        sales_extra.append(extra)
        sales_total.append(total_sales_year)

        contado = total_sales_year * COLLECTION_CASH
        credito = total_sales_year * COLLECTION_CREDIT
        current_credit = credito * COLLECTION_CURRENT_CREDIT * (1 - bad_debt_rate)
        prior_collection = previous_cxc * (1 - bad_debt_rate)
        ending_cxc = credito * (1 - COLLECTION_CURRENT_CREDIT)
        cash_sales.append(contado)
        credit_sales.append(credito)
        current_credit_collection.append(current_credit)
        previous_cxc_collection.append(prior_collection)
        cxc_end.append(ending_cxc)
        cash_income.append(contado + current_credit + prior_collection)

        impl_cost = sum(new[key][i] * PACKAGES[key].implementation_cost * factors[i] for key in new) * cost_factor
        rec_cost = sum(effective[key][i] * PACKAGES[key].recurring_cost * factors[i] for key in new) * cost_factor
        extra_cost = extra * ADDITIONAL_SERVICE_COST_RATE * cost_factor
        direct_cost = impl_cost + rec_cost + extra_cost
        direct_impl_cost.append(impl_cost)
        direct_rec_cost.append(rec_cost)
        direct_extra_cost.append(extra_cost)
        direct_total.append(direct_cost)

        cash_purchase = direct_cost * PURCHASE_CASH
        credit_purchase = direct_cost * PURCHASE_CREDIT
        current_payment = credit_purchase * PAY_CURRENT_CREDIT
        ending_cxp = credit_purchase * (1 - PAY_CURRENT_CREDIT)
        purchase_cash.append(cash_purchase)
        purchase_credit.append(credit_purchase)
        current_supplier_payment.append(current_payment)
        previous_supplier_payment.append(previous_cxp)
        cxp_end.append(ending_cxp)
        supplier_cash_payment.append(cash_purchase + current_payment + previous_cxp)

        admin_labor.append(staff["admin"][i])
        production_labor.append(staff["production"][i])
        sales_labor.append(staff["sales"][i])
        admin_other.append(16_000 * (1.04**i) * cost_factor * admin_factor)
        commission = total_sales_year * 0.025 * cost_factor
        sales_commission.append(commission)
        sales_other.append((16_000 * (1.04**i) * cost_factor) + commission)

        energy.append(20_000 * (1.04**i) * cost_factor)
        platform_tools.append((0 if i == 0 else 9_000 * (1.04 ** (i - 1))) * cost_factor)
        supplies.append(3_800 * (1.04**i) * cost_factor)
        maintenance.append(4_500 * (1.04**i) * cost_factor)
        indirect_cash_year = energy[i] + platform_tools[i] + supplies[i] + maintenance[i]
        indirect_cash.append(indirect_cash_year)
        indirect_total.append(indirect_cash_year + assets["depreciation"][i] * 0.60)

        prod_cost = direct_cost + production_labor[i] + indirect_total[i]
        cost_of_production.append(prod_cost)
        cost_of_sales.append(prod_cost)
        op_exp = admin_labor[i] + admin_other[i] + sales_labor[i] + sales_other[i] + assets["depreciation"][i] * 0.40 + assets["amortization"][i]
        operating_expenses.append(op_exp)
        ebit_year = total_sales_year - cost_of_sales[i] - op_exp
        ebit.append(ebit_year)
        ebt_year = ebit_year - debt["interest"][i]
        ebt.append(ebt_year)
        tax_year = max(0.0, ebt_year * INCOME_TAX)
        income_tax.append(tax_year)
        cash_tax_paid.append(0.0 if i == 0 else income_tax[i - 1])
        current_debt.append(debt["principal"][i + 1] if i + 1 < 10 else 0.0)

        net_income = ebt_year - tax_year
        legal = min(max(0.0, net_income * 0.10), max(0.0, reserve_limit - reserve_acc))
        reserve_acc += legal
        reserve_legal.append(legal)
        reserve_accum.append(reserve_acc)
        retained_profit.append(net_income - legal)

        cash_out = (
            supplier_cash_payment[i]
            + production_labor[i]
            + indirect_cash[i]
            + admin_labor[i]
            + admin_other[i]
            + sales_labor[i]
            + sales_other[i]
            + debt["interest"][i]
            + debt["principal"][i]
            + cash_tax_paid[i]
        )
        cash_egress.append(cash_out)
        cash_budget_flow.append(cash_income[i] - cash_out)

        previous_cxc = ending_cxc
        previous_cxp = ending_cxp

    cash_balance = [INITIAL_CASH]
    for i in range(10):
        cash_balance.append(cash_balance[-1] + cash_budget_flow[i])

    final_current_liabilities = cxp_end[-1] + income_tax[-1]
    flow_net = []
    for i in range(10):
        terminal = cxc_end[i] + SALVAGE_VALUE - final_current_liabilities if i == 9 else 0.0
        flow_net.append(cash_income[i] - cash_egress[i] + terminal)
    flows_for_valuation = [-INITIAL_INVESTMENT] + flow_net

    cost_capital = EQUITY_SHARE * EQUITY_COST + LOAN_SHARE * interest_rate
    avg_inflation = sum(INFLATION) / len(INFLATION)
    risk = cost_capital * avg_inflation
    cut_rate = cost_capital + avg_inflation + risk
    model_npv = npv(cut_rate, flows_for_valuation)
    model_irr = irr(flows_for_valuation)
    profitability_index = (model_npv + INITIAL_INVESTMENT) / INITIAL_INVESTMENT
    payback = discounted_payback(flows_for_valuation, cut_rate)

    net_income = [ebt[i] - income_tax[i] for i in range(10)]
    gross_profit = [sales_total[i] - cost_of_sales[i] for i in range(10)]
    current_assets = [cash_balance[i + 1] + cxc_end[i] for i in range(10)]
    current_liabilities = [cxp_end[i] + current_debt[i] + income_tax[i] for i in range(10)]
    fixed_net = assets["fixed_net"]
    intangible_net = assets["intangible_net"]
    total_assets = [current_assets[i] + fixed_net[i] + intangible_net[i] for i in range(10)]
    long_debt = [max(0.0, debt["final_balance"][i] - (debt["principal"][i + 1] if i + 1 < 10 else 0.0)) for i in range(10)]
    total_liabilities = [current_liabilities[i] + long_debt[i] for i in range(10)]
    accumulated_retained = []
    acc = 0.0
    for value in retained_profit:
        acc += value
        accumulated_retained.append(acc)
    equity_values = [equity + reserve_accum[i] + accumulated_retained[i] for i in range(10)]
    balance_gap = [total_assets[i] - total_liabilities[i] - equity_values[i] for i in range(10)]

    variable_labor = production_direct_labor
    variable_indirect = [
        energy[i] + platform_tools[i] + supplies[i]
        for i in range(10)
    ]
    fixed_production = [
        production_indirect_labor[i]
        + maintenance[i]
        + assets["depreciation"][i] * 0.60
        for i in range(10)
    ]
    fixed_admin = [
        admin_labor[i]
        + admin_other[i]
        + assets["depreciation"][i] * 0.40
        + assets["amortization"][i]
        for i in range(10)
    ]
    fixed_sales = [
        sales_labor[i] + sales_other[i]
        for i in range(10)
    ]
    fixed_financial = debt["interest"][:]
    variable_costs = [
        direct_total[i] + variable_labor[i] + variable_indirect[i]
        for i in range(10)
    ]
    fixed_costs = [
        fixed_production[i]
        + fixed_admin[i]
        + fixed_sales[i]
        + fixed_financial[i]
        for i in range(10)
    ]
    contribution_margin = [1 - safe_div(variable_costs[i], sales_total[i]) for i in range(10)]
    break_even_sales = [safe_div(fixed_costs[i], contribution_margin[i]) for i in range(10)]
    equivalent_museums = [safe_div(sales_total[i], max(1.0, total_new[i] + sum(effective[key][i] for key in new))) for i in range(10)]
    break_even_units = [safe_div(break_even_sales[i], equivalent_museums[i]) for i in range(10)]
    break_even_pct = [safe_div(break_even_sales[i], sales_total[i]) for i in range(10)]

    return {
        "name": name,
        "factors": factors,
        "new": new,
        "active_end": active_end,
        "effective": effective,
        "total_new": total_new,
        "total_active": total_active,
        "sales_impl": sales_impl,
        "sales_rec": sales_rec,
        "sales_extra": sales_extra,
        "sales_total": sales_total,
        "cash_sales": cash_sales,
        "credit_sales": credit_sales,
        "current_credit_collection": current_credit_collection,
        "previous_cxc_collection": previous_cxc_collection,
        "cxc_end": cxc_end,
        "cash_income": cash_income,
        "direct_impl_cost": direct_impl_cost,
        "direct_rec_cost": direct_rec_cost,
        "direct_extra_cost": direct_extra_cost,
        "direct_total": direct_total,
        "purchase_cash": purchase_cash,
        "purchase_credit": purchase_credit,
        "current_supplier_payment": current_supplier_payment,
        "previous_supplier_payment": previous_supplier_payment,
        "cxp_end": cxp_end,
        "supplier_cash_payment": supplier_cash_payment,
        "production_labor": production_labor,
        "production_direct_labor": production_direct_labor,
        "production_indirect_labor": production_indirect_labor,
        "admin_labor": admin_labor,
        "sales_labor": sales_labor,
        "admin_other": admin_other,
        "sales_other": sales_other,
        "sales_commission": sales_commission,
        "energy": energy,
        "platform_tools": platform_tools,
        "supplies": supplies,
        "maintenance": maintenance,
        "indirect_cash": indirect_cash,
        "indirect_total": indirect_total,
        "cost_of_production": cost_of_production,
        "cost_of_sales": cost_of_sales,
        "operating_expenses": operating_expenses,
        "gross_profit": gross_profit,
        "ebit": ebit,
        "ebt": ebt,
        "income_tax": income_tax,
        "cash_tax_paid": cash_tax_paid,
        "cash_egress": cash_egress,
        "cash_budget_flow": cash_budget_flow,
        "cash_balance": cash_balance,
        "reserve_legal": reserve_legal,
        "reserve_accum": reserve_accum,
        "retained_profit": retained_profit,
        "net_income": net_income,
        "debt": debt,
        "loan": loan,
        "equity": equity,
        "assets": assets,
        "flow_net": flow_net,
        "flows_for_valuation": flows_for_valuation,
        "cost_capital": cost_capital,
        "avg_inflation": avg_inflation,
        "risk": risk,
        "cut_rate": cut_rate,
        "npv": model_npv,
        "irr": model_irr,
        "profitability_index": profitability_index,
        "payback": payback,
        "current_assets": current_assets,
        "current_liabilities": current_liabilities,
        "fixed_net": fixed_net,
        "intangible_net": intangible_net,
        "total_assets": total_assets,
        "long_debt": long_debt,
        "total_liabilities": total_liabilities,
        "equity_values": equity_values,
        "balance_gap": balance_gap,
        "variable_costs": variable_costs,
        "variable_labor": variable_labor,
        "variable_indirect": variable_indirect,
        "fixed_costs": fixed_costs,
        "fixed_production": fixed_production,
        "fixed_admin": fixed_admin,
        "fixed_sales": fixed_sales,
        "fixed_financial": fixed_financial,
        "contribution_margin": contribution_margin,
        "break_even_sales": break_even_sales,
        "equivalent_museums": equivalent_museums,
        "break_even_units": break_even_units,
        "break_even_pct": break_even_pct,
        "price_adjustment": price_adjustment,
        "interest_rate": interest_rate,
        "bad_debt_rate": bad_debt_rate,
    }


def make_formats(workbook: xlsxwriter.Workbook) -> dict[str, object]:
    base = {
        "font_name": "Arial",
        "font_size": 10,
        "border_color": "#003333",
    }
    teal = "#008080"
    light_blue = "#A6CAF0"
    light_gray = "#D9D9D9"
    return {
        "title": workbook.add_format(
            {
                **base,
                "bold": True,
                "font_size": 12,
                "font_color": "white",
                "bg_color": teal,
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        ),
        "caption": workbook.add_format(
            {
                **base,
                "bold": True,
                "font_size": 11,
                "align": "center",
                "valign": "vcenter",
            }
        ),
        "subtitle": workbook.add_format(
            {
                **base,
                "bold": True,
                "font_color": "white",
                "bg_color": teal,
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        ),
        "header": workbook.add_format(
            {
                **base,
                "bold": True,
                "font_color": "white",
                "bg_color": teal,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
                "border": 1,
            }
        ),
        "label": workbook.add_format(
            {**base, "border": 1, "bg_color": "white", "valign": "vcenter"}
        ),
        "label_blue": workbook.add_format(
            {**base, "border": 1, "bg_color": light_blue, "valign": "vcenter"}
        ),
        "text": workbook.add_format(
            {
                **base,
                "border": 1,
                "bg_color": "white",
                "text_wrap": True,
                "valign": "top",
            }
        ),
        "money": workbook.add_format(
            {
                **base,
                "border": 1,
                "bg_color": "white",
                "num_format": "#,##0.00;[Red]-#,##0.00",
            }
        ),
        "money_blue": workbook.add_format(
            {
                **base,
                "border": 1,
                "bg_color": light_blue,
                "num_format": "#,##0.00;[Red]-#,##0.00",
            }
        ),
        "number": workbook.add_format(
            {**base, "border": 1, "bg_color": "white", "num_format": "#,##0.00"}
        ),
        "integer": workbook.add_format(
            {**base, "border": 1, "bg_color": "white", "num_format": "#,##0"}
        ),
        "percent": workbook.add_format(
            {**base, "border": 1, "bg_color": "white", "num_format": "0.00%"}
        ),
        "group_label": workbook.add_format(
            {**base, "bold": True, "border": 1, "bg_color": "white"}
        ),
        "group_money": workbook.add_format(
            {
                **base,
                "bold": True,
                "border": 1,
                "bg_color": "white",
                "num_format": "#,##0.00;[Red]-#,##0.00",
            }
        ),
        "summary_label": workbook.add_format(
            {**base, "bold": True, "border": 1, "bg_color": light_gray}
        ),
        "summary_money": workbook.add_format(
            {
                **base,
                "border": 1,
                "bg_color": light_gray,
                "num_format": "#,##0.00;[Red]-#,##0.00",
            }
        ),
        "summary_total": workbook.add_format(
            {
                **base,
                "bold": True,
                "border": 1,
                "bg_color": light_gray,
                "num_format": "#,##0.00;[Red]-#,##0.00",
            }
        ),
        "summary_note": workbook.add_format(
            {
                **base,
                "italic": True,
                "bg_color": light_gray,
                "text_wrap": True,
                "valign": "top",
            }
        ),
        "total": workbook.add_format(
            {
                **base,
                "bold": True,
                "font_color": "white",
                "bg_color": teal,
                "border": 1,
                "num_format": "#,##0.00;[Red]-#,##0.00",
            }
        ),
        "total_num": workbook.add_format(
            {
                **base,
                "bold": True,
                "font_color": "white",
                "bg_color": teal,
                "border": 1,
                "num_format": "#,##0.00",
            }
        ),
        "total_pct": workbook.add_format(
            {
                **base,
                "bold": True,
                "font_color": "white",
                "bg_color": teal,
                "border": 1,
                "num_format": "0.00%",
            }
        ),
        "note": workbook.add_format(
            {
                **base,
                "italic": True,
                "font_color": "#003333",
                "text_wrap": True,
                "valign": "top",
            }
        ),
    }


def setup(ws, title: str, formats: dict[str, object], last_col: int = 11) -> None:
    ws.set_zoom(80)
    ws.set_tab_color("#008080")
    ws.set_default_row(18)
    ws.freeze_panes(3, 1)
    ws.set_column(0, 0, 44)
    ws.set_column(1, last_col, 12.5)
    ws.set_landscape()
    ws.set_paper(9)
    ws.fit_to_pages(1, 0)
    ws.set_margins(left=0.25, right=0.25, top=0.35, bottom=0.35)
    ws.set_row(0, 22)
    ws.merge_range(0, 0, 0, last_col, title, formats["title"])


def write_year_header(ws, row: int, formats: dict[str, object], first_label: str = "Concepto", year0: bool = False) -> int:
    labels = [first_label] + (["Año 0"] if year0 else []) + YEAR_LABELS
    ws.set_row(row, 20)
    for col, label in enumerate(labels):
        ws.write(row, col, label, formats["header"])
    return row + 1


def write_series(
    ws,
    row: int,
    label: str,
    values: Iterable[float],
    formats: dict[str, object],
    *,
    kind: str = "money",
    total: bool = False,
    first_col: int = 1,
    formulas: Iterable[str] | None = None,
) -> int:
    ws.write_string(row, 0, label, formats["total"] if total and kind == "money" else formats["label"])
    fmt = formats["total"] if total and kind == "money" else formats["money"]
    if kind == "integer":
        fmt = formats["total_num"] if total else formats["integer"]
    elif kind == "number":
        fmt = formats["total_num"] if total else formats["number"]
    elif kind == "percent":
        fmt = formats["total_pct"] if total else formats["percent"]
    formula_values = list(formulas) if formulas is not None else None
    for index, value in enumerate(values):
        col = first_col + index
        cached = cached_value(value, kind)
        if formula_values is None:
            ws.write_number(row, col, cached, fmt)
        else:
            ws.write_formula(row, col, formula_values[index], fmt, cached)
    return row + 1


def write_section(ws, row: int, title: str, formats: dict[str, object], last_col: int = 10) -> int:
    ws.set_row(row, 20)
    ws.merge_range(row, 0, row, last_col, title, formats["subtitle"])
    return row + 1


def write_ingresos(workbook, model, formats):
    ws = workbook.add_worksheet("INGRESOS")
    setup(ws, "Servicio de Implementacion MuseIQ - Soporte, IA y Analitica", formats, 11)
    ws.merge_range(1, 0, 1, 10, "PROYECCION DE INGRESOS", formats["caption"])
    row = 2
    row = write_section(ws, row, "PROYECCION DE INGRESOS - IMPLEMENTACION MUSEIQ", formats)
    row = write_year_header(ws, row, formats)
    new_basic_row = row
    row = write_series(ws, row, "Museos nuevos - Basico", model["new"]["B"], formats, kind="integer")
    new_standard_row = row
    row = write_series(ws, row, "Museos nuevos - Estandar", model["new"]["E"], formats, kind="integer")
    new_advanced_row = row
    row = write_series(ws, row, "Museos nuevos - Avanzado", model["new"]["A"], formats, kind="integer")
    total_new_row = row
    row = write_series(
        ws,
        row,
        "Total museos implementados por año",
        model["total_new"],
        formats,
        kind="integer",
        total=True,
        formulas=[same_row_formula("SUM", [new_basic_row, new_standard_row, new_advanced_row], col) for col in range(1, 11)],
    )
    sales_impl_row = row
    row = write_series(ws, row, "Ventas implementacion MuseIQ", model["sales_impl"], formats, total=True)
    sales_extra_row = row
    row = write_series(
        ws,
        row,
        "Servicios adicionales de despliegue",
        model["sales_extra"],
        formats,
        formulas=[f"={cell(sales_impl_row, col)}*{ADDITIONAL_SERVICE_RATE}" for col in range(1, 11)],
    )
    row += 1
    row = write_section(ws, row, "INGRESOS - SOPORTE, IA Y ANALITICA", formats)
    row = write_year_header(ws, row, formats)
    active_row = row
    row = write_series(
        ws,
        row,
        "Museos activos al cierre",
        model["total_active"],
        formats,
        kind="integer",
        formulas=[f"=SUM($B${total_new_row + 1}:{cell(total_new_row, col)})" for col in range(1, 11)],
    )
    effective_basic_row = row
    row = write_series(
        ws,
        row,
        "Museos equivalentes recurrentes - Basico",
        model["effective"]["B"],
        formats,
        kind="number",
        formulas=[
            f"={cell(new_basic_row, col)}*0.5" if col == 1 else f"=SUM($B${new_basic_row + 1}:{cell(new_basic_row, col - 1)})+{cell(new_basic_row, col)}*0.5"
            for col in range(1, 11)
        ],
    )
    effective_standard_row = row
    row = write_series(
        ws,
        row,
        "Museos equivalentes recurrentes - Estandar",
        model["effective"]["E"],
        formats,
        kind="number",
        formulas=[
            f"={cell(new_standard_row, col)}*0.5" if col == 1 else f"=SUM($B${new_standard_row + 1}:{cell(new_standard_row, col - 1)})+{cell(new_standard_row, col)}*0.5"
            for col in range(1, 11)
        ],
    )
    effective_advanced_row = row
    row = write_series(
        ws,
        row,
        "Museos equivalentes recurrentes - Avanzado",
        model["effective"]["A"],
        formats,
        kind="number",
        formulas=[
            f"={cell(new_advanced_row, col)}*0.5" if col == 1 else f"=SUM($B${new_advanced_row + 1}:{cell(new_advanced_row, col - 1)})+{cell(new_advanced_row, col)}*0.5"
            for col in range(1, 11)
        ],
    )
    sales_rec_row = row
    row = write_series(ws, row, "Ventas soporte, IA y analitica", model["sales_rec"], formats, total=True)
    sales_total_row = row
    row = write_series(
        ws,
        row,
        "VENTAS TOTALES",
        model["sales_total"],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [sales_impl_row, sales_extra_row, sales_rec_row], col) for col in range(1, 11)],
    )
    row += 1
    row = write_section(ws, row, "DETERMINACION DE INGRESOS DE EFECTIVO", formats)
    row = write_year_header(ws, row, formats)
    cash_sales_row = row
    row = write_series(
        ws,
        row,
        "Ventas al contado 40%",
        model["cash_sales"],
        formats,
        formulas=[f"={cell(sales_total_row, col)}*{COLLECTION_CASH}" for col in range(1, 11)],
    )
    credit_sales_row = row
    row = write_series(
        ws,
        row,
        "Ventas al credito 60%",
        model["credit_sales"],
        formats,
        formulas=[f"={cell(sales_total_row, col)}*{COLLECTION_CREDIT}" for col in range(1, 11)],
    )
    current_credit_row = row
    row = write_series(
        ws,
        row,
        "Recuperacion de cuentas por cobrar del año",
        model["current_credit_collection"],
        formats,
        formulas=[f"={cell(credit_sales_row, col)}*{COLLECTION_CURRENT_CREDIT}*(1-{model['bad_debt_rate']})" for col in range(1, 11)],
    )
    previous_cxc_row = row
    row = write_series(
        ws,
        row,
        "Recuperacion de cuentas por cobrar año anterior",
        model["previous_cxc_collection"],
        formats,
    )
    cash_income_row = row
    row = write_series(
        ws,
        row,
        "TOTAL INGRESOS DE EFECTIVO",
        model["cash_income"],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [cash_sales_row, current_credit_row, previous_cxc_row], col) for col in range(1, 11)],
    )
    cxc_end_row = row
    row = write_series(
        ws,
        row,
        "Saldo de cuentas por cobrar al final del año",
        model["cxc_end"],
        formats,
        formulas=[f"={cell(credit_sales_row, col)}*(1-{COLLECTION_CURRENT_CREDIT})" for col in range(1, 11)],
    )
    row += 1
    row = write_section(ws, row, "INFLACION PROYECTADA Y PRECIOS BASE", formats)
    row = write_year_header(ws, row, formats)
    inflation_row = row
    row = write_series(ws, row, "Inflacion anual proyectada", INFLATION, formats, kind="percent")
    factor_row = row
    row = write_series(
        ws,
        row,
        "Factor acumulado aplicado a precios",
        model["factors"],
        formats,
        kind="number",
        formulas=[f"=1" if col == 1 else f"={cell(factor_row, col - 1)}*(1+{cell(inflation_row, col - 1)})" for col in range(1, 11)],
    )
    for col in range(1, 11):
        ws.write_formula(
            sales_impl_row,
            col,
            f"=({cell(new_basic_row, col)}*{PACKAGES['B'].base_implementation_price}+{cell(new_standard_row, col)}*{PACKAGES['E'].base_implementation_price}+{cell(new_advanced_row, col)}*{PACKAGES['A'].base_implementation_price})*{model['price_adjustment']}*{cell(factor_row, col)}",
            formats["total"],
            money(model["sales_impl"][col - 1]),
        )
        ws.write_formula(
            sales_rec_row,
            col,
            f"=({cell(effective_basic_row, col)}*{PACKAGES['B'].base_recurring_price}+{cell(effective_standard_row, col)}*{PACKAGES['E'].base_recurring_price}+{cell(effective_advanced_row, col)}*{PACKAGES['A'].base_recurring_price})*{model['price_adjustment']}*{cell(factor_row, col)}",
            formats["total"],
            money(model["sales_rec"][col - 1]),
        )
        ws.write_formula(
            previous_cxc_row,
            col,
            "=0" if col == 1 else f"={cell(cxc_end_row, col - 1)}",
            formats["money"],
            money(model["previous_cxc_collection"][col - 1]),
        )
    row += 2
    ws.write(row, 0, "Paquete", formats["header"])
    ws.write(row, 1, "Precio base cap. 5", formats["header"])
    ws.write(row, 2, "Factor economico", formats["header"])
    ws.write(row, 3, "Precio economico año 1", formats["header"])
    ws.write(row, 4, "Recurrente base cap. 5", formats["header"])
    ws.write(row, 5, "Recurrente economico año 1", formats["header"])
    ws.write(row, 6, "Cobertura", formats["header"])
    ws.write(row, 7, "BLE estimados", formats["header"])
    row += 1
    package_rows = {}
    for package_index, key in enumerate(["B", "E", "A"]):
        pkg = PACKAGES[key]
        package_rows[key] = row
        ws.write(row, 0, pkg.name, formats["label"])
        ws.write_number(row, 1, pkg.base_implementation_price, formats["money"])
        if package_index == 0:
            ws.write_number(row, 2, model["price_adjustment"], formats["number"])
        else:
            ws.write_formula(
                row,
                2,
                f"={cell(package_rows['B'], 2, abs_row=True, abs_col=True)}",
                formats["number"],
                model["price_adjustment"],
            )
        ws.write_formula(row, 3, f"=B{row + 1}*C{row + 1}", formats["money"], money(pkg.base_implementation_price * model["price_adjustment"]))
        ws.write_number(row, 4, pkg.base_recurring_price, formats["money"])
        ws.write_formula(row, 5, f"=E{row + 1}*C{row + 1}", formats["money"], money(pkg.base_recurring_price * model["price_adjustment"]))
        ws.write(row, 6, pkg.coverage, formats["text"])
        ws.write(row, 7, pkg.beacons, formats["text"])
        row += 1
    ws.write(row + 1, 0, "Nota: el precio economico aplica un factor comercial de 20% sobre los valores del capitulo 5 para cubrir formalizacion, garantia, gestion contractual y contingencia operativa.", formats["note"])

    row += 3
    row = write_section(ws, row, "DETALLE DE PRECIOS Y VENTAS DE IMPLEMENTACION", formats)
    row = write_year_header(ws, row, formats)
    impl_price_rows = {}
    impl_sales_rows = {}
    for key, source_row in package_rows.items():
        pkg = PACKAGES[key]
        impl_price_rows[key] = row
        values = [pkg.base_implementation_price * model["price_adjustment"] * model["factors"][i] for i in range(10)]
        row = write_series(
            ws,
            row,
            f"Precio implementacion - {pkg.name}",
            values,
            formats,
            formulas=[f"={cell(source_row, 3, abs_col=True, abs_row=True)}*{cell(factor_row, col)}" for col in range(1, 11)],
        )
    source_new_rows = {"B": new_basic_row, "E": new_standard_row, "A": new_advanced_row}
    for key in ["B", "E", "A"]:
        pkg = PACKAGES[key]
        impl_sales_rows[key] = row
        row = write_series(
            ws,
            row,
            f"Venta implementacion - {pkg.name}",
            [
                model["new"][key][i] * pkg.base_implementation_price * model["price_adjustment"] * model["factors"][i]
                for i in range(10)
            ],
            formats,
            formulas=[f"={cell(source_new_rows[key], col)}*{cell(impl_price_rows[key], col)}" for col in range(1, 11)],
        )
    impl_detail_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL VENTAS IMPLEMENTACION DETALLADAS",
        model["sales_impl"],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [impl_sales_rows[key] for key in ["B", "E", "A"]], col) for col in range(1, 11)],
    )

    row += 1
    row = write_section(ws, row, "DETALLE DE PRECIOS Y VENTAS RECURRENTES", formats)
    row = write_year_header(ws, row, formats)
    rec_price_rows = {}
    rec_sales_rows = {}
    for key, source_row in package_rows.items():
        pkg = PACKAGES[key]
        rec_price_rows[key] = row
        values = [pkg.base_recurring_price * model["price_adjustment"] * model["factors"][i] for i in range(10)]
        row = write_series(
            ws,
            row,
            f"Precio recurrente - {pkg.name}",
            values,
            formats,
            formulas=[f"={cell(source_row, 5, abs_col=True, abs_row=True)}*{cell(factor_row, col)}" for col in range(1, 11)],
        )
    source_effective_rows = {"B": effective_basic_row, "E": effective_standard_row, "A": effective_advanced_row}
    for key in ["B", "E", "A"]:
        pkg = PACKAGES[key]
        rec_sales_rows[key] = row
        row = write_series(
            ws,
            row,
            f"Venta soporte, IA y analitica - {pkg.name}",
            [
                model["effective"][key][i] * pkg.base_recurring_price * model["price_adjustment"] * model["factors"][i]
                for i in range(10)
            ],
            formats,
            formulas=[f"={cell(source_effective_rows[key], col)}*{cell(rec_price_rows[key], col)}" for col in range(1, 11)],
        )
    rec_detail_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL VENTAS RECURRENTES DETALLADAS",
        model["sales_rec"],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [rec_sales_rows[key] for key in ["B", "E", "A"]], col) for col in range(1, 11)],
    )

    row += 1
    row = write_section(ws, row, "VALORES PROMEDIO Y POLITICA DE COBRANZA", formats)
    row = write_year_header(ws, row, formats)
    row = write_series(
        ws,
        row,
        "Valor promedio por museo implementado",
        [safe_div(model["sales_impl"][i], model["total_new"][i]) for i in range(10)],
        formats,
        formulas=[f"=IFERROR({cell(sales_impl_row, col)}/{cell(total_new_row, col)},0)" for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "Valor promedio por museo recurrente equivalente",
        [
            safe_div(model["sales_rec"][i], sum(model["effective"][key][i] for key in ["B", "E", "A"]))
            for i in range(10)
        ],
        formats,
        formulas=[
            f"=IFERROR({cell(sales_rec_row, col)}/SUM({cell(effective_basic_row, col)}:{cell(effective_advanced_row, col)}),0)"
            for col in range(1, 11)
        ],
    )
    row += 1
    ws.write(row, 0, "Politica de cobranza", formats["header"])
    ws.write(row, 1, "Valor", formats["header"])
    row += 1
    collection_cash_policy_row = row
    ws.write(row, 0, "Contado", formats["label"])
    ws.write_number(row, 1, COLLECTION_CASH, formats["percent"])
    row += 1
    collection_credit_policy_row = row
    ws.write(row, 0, "Credito", formats["label"])
    ws.write_formula(row, 1, f"=1-B{collection_cash_policy_row + 1}", formats["percent"], COLLECTION_CREDIT)
    row += 1
    current_collection_policy_row = row
    ws.write(row, 0, "Recuperacion del credito en el año", formats["label"])
    ws.write_number(row, 1, COLLECTION_CURRENT_CREDIT, formats["percent"])
    row += 1
    ws.write(row, 0, "Saldo de credito al cierre", formats["label"])
    ws.write_formula(row, 1, f"=1-B{current_collection_policy_row + 1}", formats["percent"], 1 - COLLECTION_CURRENT_CREDIT)
    row += 2
    ws.write(row, 0, "Politica de servicios adicionales", formats["header"])
    ws.write(row, 1, "Valor", formats["header"])
    row += 1
    additional_service_policy_row = row
    ws.write(row, 0, "Servicios adicionales sobre implementacion", formats["label"])
    ws.write_number(row, 1, ADDITIONAL_SERVICE_RATE, formats["percent"])
    row += 2
    ws.write(row, 0, "Politica de activacion recurrente", formats["header"])
    ws.write(row, 1, "Valor", formats["header"])
    row += 1
    recurrent_activation_policy_row = row
    ws.write(
        row,
        0,
        "Fraccion del primer año reconocida como servicio recurrente",
        formats["label"],
    )
    ws.write_number(row, 1, 0.50, formats["percent"])

    for col in range(1, 11):
        for effective_row, new_row in [
            (effective_basic_row, new_basic_row),
            (effective_standard_row, new_standard_row),
            (effective_advanced_row, new_advanced_row),
        ]:
            prior_years = (
                "0"
                if col == 1
                else f"SUM($B${new_row + 1}:{cell(new_row, col - 1)})"
            )
            ws.write_formula(
                effective_row,
                col,
                f"={prior_years}+{cell(new_row, col)}"
                f"*{cell(recurrent_activation_policy_row, 1, abs_row=True, abs_col=True)}",
                formats["number"],
                model["effective"][
                    {
                        effective_basic_row: "B",
                        effective_standard_row: "E",
                        effective_advanced_row: "A",
                    }[effective_row]
                ][col - 1],
            )
        ws.write_formula(sales_impl_row, col, f"={cell(impl_detail_total_row, col)}", formats["total"], money(model["sales_impl"][col - 1]))
        ws.write_formula(sales_rec_row, col, f"={cell(rec_detail_total_row, col)}", formats["total"], money(model["sales_rec"][col - 1]))
        ws.write_formula(
            sales_extra_row,
            col,
            f"={cell(sales_impl_row, col)}"
            f"*{cell(additional_service_policy_row, 1, abs_row=True, abs_col=True)}",
            formats["money"],
            money(model["sales_extra"][col - 1]),
        )
        ws.write_formula(cash_sales_row, col, f"={cell(sales_total_row, col)}*$B${collection_cash_policy_row + 1}", formats["money"], money(model["cash_sales"][col - 1]))
        ws.write_formula(credit_sales_row, col, f"={cell(sales_total_row, col)}*$B${collection_credit_policy_row + 1}", formats["money"], money(model["credit_sales"][col - 1]))
        ws.write_formula(current_credit_row, col, f"={cell(credit_sales_row, col)}*$B${current_collection_policy_row + 1}", formats["money"], money(model["current_credit_collection"][col - 1]))
        ws.write_formula(cxc_end_row, col, f"={cell(credit_sales_row, col)}*(1-$B${current_collection_policy_row + 1})", formats["money"], money(model["cxc_end"][col - 1]))


def write_costos(workbook, model, formats):
    ws = workbook.add_worksheet("COSTOS")
    setup(ws, "Cuadro de costos de insumos directos anuales", formats, 12)
    ws.set_column(0, 0, 48)
    ws.set_column(1, 10, 13)
    ws.set_column(11, 11, 42)

    def annual_input_values(cost_input, units_by_package):
        return [
            sum(
                units_by_package[key][year_index] * package_input_cost(cost_input, key)
                for key in ["B", "E", "A"]
            )
            * model["factors"][year_index]
            for year_index in range(10)
        ]

    row = 2
    row = write_section(ws, row, "COSTOS DIRECTOS ANUALES DE IMPLEMENTACION", formats)
    row = write_year_header(ws, row, formats)
    implementation_annual_rows = []
    for cost_input in IMPLEMENTATION_COST_INPUTS:
        implementation_annual_rows.append(row)
        row = write_series(
            ws,
            row,
            cost_input[0],
            annual_input_values(cost_input, model["new"]),
            formats,
        )
    direct_impl_row = row
    row = write_series(
        ws,
        row,
        "TOTAL COSTOS DE IMPLEMENTACION",
        model["direct_impl_cost"],
        formats,
        total=True,
        formulas=[
            f"=SUM({cell(implementation_annual_rows[0], col)}:{cell(implementation_annual_rows[-1], col)})"
            for col in range(1, 11)
        ],
    )

    row = write_section(ws, row, "COSTOS DIRECTOS ANUALES DE OPERACION RECURRENTE", formats)
    row = write_year_header(ws, row, formats)
    recurring_annual_rows = []
    for cost_input in RECURRING_COST_INPUTS:
        recurring_annual_rows.append(row)
        row = write_series(
            ws,
            row,
            cost_input[0],
            annual_input_values(cost_input, model["effective"]),
            formats,
        )
    direct_rec_row = row
    row = write_series(
        ws,
        row,
        "TOTAL COSTOS RECURRENTES",
        model["direct_rec_cost"],
        formats,
        total=True,
        formulas=[
            f"=SUM({cell(recurring_annual_rows[0], col)}:{cell(recurring_annual_rows[-1], col)})"
            for col in range(1, 11)
        ],
    )

    additional_components = [
        ("Produccion y adecuacion curatorial adicional", 0.40),
        ("Traduccion, accesibilidad y nuevas narrativas", 0.20),
        ("Capacitacion adicional", 0.20),
        ("Despliegue, materiales y logistica adicional", 0.20),
    ]
    row = write_section(ws, row, "COSTOS DIRECTOS DE SERVICIOS ADICIONALES", formats)
    row = write_year_header(ws, row, formats)
    additional_annual_rows = []
    for label, share in additional_components:
        additional_annual_rows.append(row)
        row = write_series(
            ws,
            row,
            label,
            [value * share for value in model["direct_extra_cost"]],
            formats,
        )
    direct_extra_row = row
    row = write_series(
        ws,
        row,
        "TOTAL COSTOS DE SERVICIOS ADICIONALES",
        model["direct_extra_cost"],
        formats,
        total=True,
        formulas=[
            f"=SUM({cell(additional_annual_rows[0], col)}:{cell(additional_annual_rows[-1], col)})"
            for col in range(1, 11)
        ],
    )
    direct_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL COSTO DIRECTO ANUAL",
        model["direct_total"],
        formats,
        total=True,
        formulas=[
            same_row_formula("SUM", [direct_impl_row, direct_rec_row, direct_extra_row], col)
            for col in range(1, 11)
        ],
    )

    row += 1
    row = write_section(ws, row, "CUADRO DE PRODUCCION - UNIDADES DE SERVICIO", formats)
    row = write_year_header(ws, row, formats)
    implemented_units_row = row
    row = write_series(
        ws,
        row,
        "Museos implementados",
        model["total_new"],
        formats,
        kind="integer",
        formulas=[f"={sheet_cell('INGRESOS', 7, col)}" for col in range(1, 11)],
    )
    active_units_row = row
    row = write_series(
        ws,
        row,
        "Museos activos al cierre",
        model["total_active"],
        formats,
        kind="integer",
        formulas=[f"={sheet_cell('INGRESOS', 13, col)}" for col in range(1, 11)],
    )
    recurrent_units = [
        sum(model["effective"][key][year_index] for key in ["B", "E", "A"])
        for year_index in range(10)
    ]
    recurrent_units_row = row
    row = write_series(
        ws,
        row,
        "Museos equivalentes recurrentes",
        recurrent_units,
        formats,
        kind="number",
        formulas=[
            f"=SUM({sheet_cell('INGRESOS', 14, col)}:{cell(16, col)})"
            for col in range(1, 11)
        ],
    )

    row += 1
    row = write_section(ws, row, "PRESUPUESTO DE COMPRA Y PAGO A PROVEEDORES", formats)
    row = write_year_header(ws, row, formats)
    purchase_cash_row = row
    row = write_series(ws, row, "Compra de contado 30%", model["purchase_cash"], formats)
    purchase_credit_row = row
    row = write_series(ws, row, "Compra a proveedores al credito 70%", model["purchase_credit"], formats)
    current_payment_row = row
    row = write_series(ws, row, "Pago del credito en el año 85%", model["current_supplier_payment"], formats)
    previous_supplier_row = row
    row = write_series(ws, row, "Pago a proveedores del año anterior", model["previous_supplier_payment"], formats)
    supplier_payment_row = row
    row = write_series(
        ws,
        row,
        "TOTAL PAGO A PROVEEDORES",
        model["supplier_cash_payment"],
        formats,
        total=True,
        formulas=[
            same_row_formula("SUM", [purchase_cash_row, current_payment_row, previous_supplier_row], col)
            for col in range(1, 11)
        ],
    )
    cxp_row = row
    row = write_series(ws, row, "Saldo de cuentas por pagar", model["cxp_end"], formats)

    electricity = [value * 0.52 for value in model["energy"]]
    communications = [value * 0.48 for value in model["energy"]]
    row += 1
    row = write_section(ws, row, "COSTOS INDIRECTOS ANUALES DE PRODUCCION", formats)
    row = write_year_header(ws, row, formats)
    electricity_row = row
    row = write_series(ws, row, "Energia electrica y ambiente de pruebas", electricity, formats)
    communications_row = row
    row = write_series(ws, row, "Internet y comunicaciones de produccion", communications, formats)
    tools_row = row
    row = write_series(ws, row, "Herramientas cloud, repositorios y monitoreo", model["platform_tools"], formats)
    supplies_row = row
    row = write_series(ws, row, "Papeleria, insumos de oficina y pruebas", model["supplies"], formats)
    maintenance_row = row
    row = write_series(ws, row, "Mantenimiento menor de equipos", model["maintenance"], formats)
    production_dep_row = row
    row = write_series(
        ws,
        row,
        "Depreciacion asignada a produccion",
        [value * 0.60 for value in model["assets"]["depreciation"]],
        formats,
    )
    indirect_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL COSTO INDIRECTO DE PRODUCCION",
        model["indirect_total"],
        formats,
        total=True,
        formulas=[
            f"=SUM({cell(electricity_row, col)}:{cell(production_dep_row, col)})"
            for col in range(1, 11)
        ],
    )

    row += 2
    row = write_section(ws, row, "CEDULA DE COSTOS UNITARIOS DE IMPLEMENTACION POR PAQUETE", formats, 11)
    cost_headers = [
        "Concepto",
        "Unidad",
        "Cant. Basico",
        "Tarifa Basico",
        "Costo Basico",
        "Cant. Estandar",
        "Tarifa Estandar",
        "Costo Estandar",
        "Cant. Avanzado",
        "Tarifa Avanzado",
        "Costo Avanzado",
        "Base / fuente",
    ]
    for col, header in enumerate(cost_headers):
        ws.write(row, col, header, formats["header"])
    row += 1
    implementation_unit_rows = []
    for cost_input in IMPLEMENTATION_COST_INPUTS:
        implementation_unit_rows.append(row)
        ws.write(row, 0, cost_input[0], formats["label"])
        ws.write(row, 1, cost_input[1], formats["text"])
        for key, qty_col, rate_col, total_col in [
            ("B", 2, 3, 4),
            ("E", 5, 6, 7),
            ("A", 8, 9, 10),
        ]:
            quantity = cost_input[2][key]
            rate = cost_input[3][key]
            ws.write_number(row, qty_col, quantity, formats["number"])
            ws.write_number(row, rate_col, rate, formats["money"])
            ws.write_formula(
                row,
                total_col,
                f"={cell(row, qty_col)}*{cell(row, rate_col)}",
                formats["money"],
                money(quantity * rate),
            )
        ws.write(row, 11, cost_input[4], formats["note"])
        row += 1
    implementation_unit_total_row = row
    ws.write(row, 0, "TOTAL COSTO UNITARIO DE IMPLEMENTACION", formats["total"])
    for col, package_key in [(4, "B"), (7, "E"), (10, "A")]:
        ws.write_formula(
            row,
            col,
            f"=SUM({cell(implementation_unit_rows[0], col)}:{cell(implementation_unit_rows[-1], col)})",
            formats["total"],
            PACKAGES[package_key].implementation_cost,
        )

    row += 2
    row = write_section(ws, row, "CEDULA DE COSTOS UNITARIOS RECURRENTES POR PAQUETE", formats, 11)
    for col, header in enumerate(cost_headers):
        ws.write(row, col, header, formats["header"])
    row += 1
    recurring_unit_rows = []
    for cost_input in RECURRING_COST_INPUTS:
        recurring_unit_rows.append(row)
        ws.write(row, 0, cost_input[0], formats["label"])
        ws.write(row, 1, cost_input[1], formats["text"])
        for key, qty_col, rate_col, total_col in [
            ("B", 2, 3, 4),
            ("E", 5, 6, 7),
            ("A", 8, 9, 10),
        ]:
            quantity = cost_input[2][key]
            rate = cost_input[3][key]
            ws.write_number(row, qty_col, quantity, formats["number"])
            ws.write_number(row, rate_col, rate, formats["money"])
            ws.write_formula(
                row,
                total_col,
                f"={cell(row, qty_col)}*{cell(row, rate_col)}",
                formats["money"],
                money(quantity * rate),
            )
        ws.write(row, 11, cost_input[4], formats["note"])
        row += 1
    recurring_unit_total_row = row
    ws.write(row, 0, "TOTAL COSTO UNITARIO RECURRENTE", formats["total"])
    for col, package_key in [(4, "B"), (7, "E"), (10, "A")]:
        ws.write_formula(
            row,
            col,
            f"=SUM({cell(recurring_unit_rows[0], col)}:{cell(recurring_unit_rows[-1], col)})",
            formats["total"],
            PACKAGES[package_key].recurring_cost,
        )

    row += 2
    row = write_section(ws, row, "COMPOSICION DEL COSTO DE SERVICIOS ADICIONALES", formats, 3)
    ws.write(row, 0, "Concepto", formats["header"])
    ws.write(row, 1, "Participacion", formats["header"])
    ws.write(row, 2, "Base", formats["header"])
    row += 1
    additional_component_input_rows = []
    for component_index, (label, share) in enumerate(additional_components):
        additional_component_input_rows.append(row)
        ws.write(row, 0, label, formats["label"])
        ws.write_number(row, 1, share, formats["percent"])
        if component_index == 0:
            ws.write_number(
                row,
                2,
                ADDITIONAL_SERVICE_COST_RATE,
                formats["percent"],
            )
        else:
            ws.write_formula(
                row,
                2,
                f"={cell(additional_component_input_rows[0], 2, abs_row=True, abs_col=True)}",
                formats["percent"],
                ADDITIONAL_SERVICE_COST_RATE,
            )
        row += 1

    row += 2
    row = write_section(ws, row, "POLITICA DE COMPRA DE INSUMOS Y SERVICIOS DIRECTOS", formats, 3)
    ws.write(row, 0, "Concepto", formats["header"])
    ws.write(row, 1, "Valor", formats["header"])
    row += 1
    purchase_cash_policy_row = row
    ws.write(row, 0, "Compra de contado", formats["label"])
    ws.write_number(row, 1, PURCHASE_CASH, formats["percent"])
    row += 1
    purchase_credit_policy_row = row
    ws.write(row, 0, "Compra al credito", formats["label"])
    ws.write_formula(row, 1, f"=1-B{purchase_cash_policy_row + 1}", formats["percent"], PURCHASE_CREDIT)
    row += 1
    supplier_payment_policy_row = row
    ws.write(row, 0, "Pago del credito en el año", formats["label"])
    ws.write_number(row, 1, PAY_CURRENT_CREDIT, formats["percent"])

    row += 2
    row = write_section(ws, row, "DETALLE DE ENERGIA Y COMUNICACIONES DE PRODUCCION", formats, 9)
    equipment_headers = [
        "Equipo / recurso",
        "KW o base",
        "Horas/dia",
        "Dias/año",
        "Tarifa",
        "Gasto anual",
        "% produccion",
        "Asignado produccion",
        "Tipo",
        "Base",
    ]
    for col, header in enumerate(equipment_headers):
        ws.write(row, col, header, formats["header"])
    row += 1
    energy_equipment = [
        ("Laptops de desarrollo y pruebas", 1.20, 8, 270, 0.55, "Energia"),
        ("Servidor/NAS para RAG y datos", 0.80, 24, 365, 0.55, "Energia"),
        ("Smartphones y tablets de prueba", 0.25, 4, 270, 0.55, "Energia"),
        ("Red WiFi, router y UPS", 0.18, 24, 365, 0.55, "Energia"),
        ("Iluminacion y ambiente de pruebas", 4_104.26, 1, 1, 1.00, "Energia"),
        ("Internet y comunicaciones anual", 9_600.00, 1, 1, 1.00, "Comunicaciones"),
    ]
    energy_detail_rows = []
    communications_detail_rows = []
    for name, base, hours, days, tariff, cost_type in energy_equipment:
        target_rows = energy_detail_rows if cost_type == "Energia" else communications_detail_rows
        target_rows.append(row)
        annual_cost = base * hours * days * tariff
        ws.write(row, 0, name, formats["label"])
        ws.write_number(row, 1, base, formats["number"])
        ws.write_number(row, 2, hours, formats["number"])
        ws.write_number(row, 3, days, formats["integer"])
        ws.write_number(row, 4, tariff, formats["number"])
        ws.write_formula(row, 5, f"=B{row + 1}*C{row + 1}*D{row + 1}*E{row + 1}", formats["money"], money(annual_cost))
        ws.write_number(row, 6, 1.0, formats["percent"])
        ws.write_formula(row, 7, f"=F{row + 1}*G{row + 1}", formats["money"], money(annual_cost))
        ws.write(row, 8, cost_type, formats["text"])
        ws.write(row, 9, "Año 1", formats["text"])
        row += 1
    energy_equipment_total_row = row
    ws.write(row, 0, "TOTAL ENERGIA ELECTRICA", formats["total"])
    ws.write_formula(
        row,
        7,
        f"=SUM({cell(energy_detail_rows[0], 7)}:{cell(energy_detail_rows[-1], 7)})",
        formats["total"],
        10_400.0,
    )
    row += 1
    communications_equipment_total_row = row
    ws.write(row, 0, "TOTAL INTERNET Y COMUNICACIONES", formats["total"])
    ws.write_formula(
        row,
        7,
        f"=SUM({cell(communications_detail_rows[0], 7)}:{cell(communications_detail_rows[-1], 7)})",
        formats["total"],
        9_600.0,
    )

    row += 2
    row = write_section(
        ws,
        row,
        "PARAMETROS DE OTROS COSTOS INDIRECTOS",
        formats,
        4,
    )
    for col, header in enumerate(
        ["Concepto", "Base anual", "Año de inicio", "Crecimiento", "Fuente"]
    ):
        ws.write(row, col, header, formats["header"])
    row += 1
    indirect_parameter_rows = {}
    indirect_parameters = [
        ("Herramientas cloud, repositorios y monitoreo", 9_000.0, 2),
        ("Papeleria, insumos de oficina y pruebas", 3_800.0, 1),
        ("Mantenimiento menor de equipos", 4_500.0, 1),
    ]
    for parameter_index, (label, base, start_year) in enumerate(
        indirect_parameters
    ):
        indirect_parameter_rows[label] = row
        ws.write(row, 0, label, formats["label"])
        ws.write_number(row, 1, base, formats["money"])
        ws.write_number(row, 2, start_year, formats["integer"])
        if parameter_index == 0:
            ws.write_formula(
                row,
                3,
                f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 59, 1)}",
                formats["percent"],
                0.04,
            )
        else:
            ws.write_formula(
                row,
                3,
                f"={cell(indirect_parameter_rows[indirect_parameters[0][0]], 3, abs_row=True, abs_col=True)}",
                formats["percent"],
                0.04,
            )
        ws.write(row, 4, "Politica anual comun de gastos", formats["text"])
        row += 1

    operating_growth_ref = cell(
        indirect_parameter_rows[indirect_parameters[0][0]],
        3,
        abs_row=True,
        abs_col=True,
    )

    new_refs = {"B": 4, "E": 5, "A": 6}
    effective_refs = {"B": 14, "E": 15, "A": 16}
    package_total_cols = {"B": 4, "E": 7, "A": 10}
    for year_col in range(1, 11):
        factor_ref = sheet_cell("INGRESOS", 32, year_col)
        for annual_row, unit_row in zip(implementation_annual_rows, implementation_unit_rows):
            package_terms = [
                f"{sheet_cell('INGRESOS', new_refs[key], year_col)}*{cell(unit_row, package_total_cols[key], abs_row=True)}"
                for key in ["B", "E", "A"]
            ]
            ws.write_formula(
                annual_row,
                year_col,
                f"=({'+'.join(package_terms)})*{factor_ref}",
                formats["money"],
                money(annual_input_values(IMPLEMENTATION_COST_INPUTS[implementation_annual_rows.index(annual_row)], model["new"])[year_col - 1]),
            )
        for annual_row, unit_row in zip(recurring_annual_rows, recurring_unit_rows):
            package_terms = [
                f"{sheet_cell('INGRESOS', effective_refs[key], year_col)}*{cell(unit_row, package_total_cols[key], abs_row=True)}"
                for key in ["B", "E", "A"]
            ]
            ws.write_formula(
                annual_row,
                year_col,
                f"=({'+'.join(package_terms)})*{factor_ref}",
                formats["money"],
                money(annual_input_values(RECURRING_COST_INPUTS[recurring_annual_rows.index(annual_row)], model["effective"])[year_col - 1]),
            )
        for component_row, input_row, (_, share) in zip(
            additional_annual_rows,
            additional_component_input_rows,
            additional_components,
        ):
            ws.write_formula(
                component_row,
                year_col,
                f"={sheet_cell('INGRESOS', 9, year_col)}"
                f"*{cell(input_row, 2, abs_row=True, abs_col=True)}"
                f"*{cell(input_row, 1, abs_row=True, abs_col=True)}",
                formats["money"],
                money(model["direct_extra_cost"][year_col - 1] * share),
            )

        ws.write_formula(
            purchase_cash_row,
            year_col,
            f"={cell(direct_total_row, year_col)}*$B${purchase_cash_policy_row + 1}",
            formats["money"],
            money(model["purchase_cash"][year_col - 1]),
        )
        ws.write_formula(
            purchase_credit_row,
            year_col,
            f"={cell(direct_total_row, year_col)}*$B${purchase_credit_policy_row + 1}",
            formats["money"],
            money(model["purchase_credit"][year_col - 1]),
        )
        ws.write_formula(
            current_payment_row,
            year_col,
            f"={cell(purchase_credit_row, year_col)}*$B${supplier_payment_policy_row + 1}",
            formats["money"],
            money(model["current_supplier_payment"][year_col - 1]),
        )
        ws.write_formula(
            previous_supplier_row,
            year_col,
            "=0" if year_col == 1 else f"={cell(cxp_row, year_col - 1)}",
            formats["money"],
            money(model["previous_supplier_payment"][year_col - 1]),
        )
        ws.write_formula(
            cxp_row,
            year_col,
            f"={cell(purchase_credit_row, year_col)}*(1-$B${supplier_payment_policy_row + 1})",
            formats["money"],
            money(model["cxp_end"][year_col - 1]),
        )
        ws.write_formula(
            electricity_row,
            year_col,
            f"=$H${energy_equipment_total_row + 1}*(1+{operating_growth_ref})^{year_col - 1}",
            formats["money"],
            money(electricity[year_col - 1]),
        )
        ws.write_formula(
            communications_row,
            year_col,
            f"=$H${communications_equipment_total_row + 1}*(1+{operating_growth_ref})^{year_col - 1}",
            formats["money"],
            money(communications[year_col - 1]),
        )
        for target_row, label, values in [
            (
                tools_row,
                "Herramientas cloud, repositorios y monitoreo",
                model["platform_tools"],
            ),
            (
                supplies_row,
                "Papeleria, insumos de oficina y pruebas",
                model["supplies"],
            ),
            (
                maintenance_row,
                "Mantenimiento menor de equipos",
                model["maintenance"],
            ),
        ]:
            parameter_row = indirect_parameter_rows[label]
            ws.write_formula(
                target_row,
                year_col,
                f"=IF({year_col}<{cell(parameter_row, 2, abs_row=True, abs_col=True)},0,"
                f"{cell(parameter_row, 1, abs_row=True, abs_col=True)}"
                f"*(1+{cell(parameter_row, 3, abs_row=True, abs_col=True)})"
                f"^({year_col}-{cell(parameter_row, 2, abs_row=True, abs_col=True)}))",
                formats["money"],
                money(values[year_col - 1]),
            )
        ws.write_formula(
            production_dep_row,
            year_col,
            f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 103, year_col)}",
            formats["money"],
            money(model["assets"]["depreciation"][year_col - 1] * 0.60),
        )

    assert direct_impl_row == COSTOS_IMPL_SUBTOTAL_ROW
    assert direct_rec_row == COSTOS_RECURRENT_SUBTOTAL_ROW
    assert direct_extra_row == COSTOS_EXTRA_SUBTOTAL_ROW
    assert direct_total_row == COSTOS_DIRECT_TOTAL_ROW
    assert implemented_units_row == COSTOS_IMPLEMENTED_UNITS_ROW
    assert recurrent_units_row == COSTOS_RECURRENT_UNITS_ROW
    assert cxp_row == COSTOS_CXP_ROW
    assert indirect_total_row == COSTOS_INDIRECT_TOTAL_ROW


def write_unit_cost(workbook, model, formats):
    ws = workbook.add_worksheet("COSTO DE PRODUCCION UNITARIO")
    setup(ws, "Costo Total de Operacion - MuseIQ", formats, 11)
    row = 3
    row = write_year_header(ws, row, formats)
    direct_row = row
    row = write_series(ws, row, "Insumos y servicios directos", model["direct_total"], formats)
    labor_row = row
    row = write_series(ws, row, "Mano de obra directa", model["production_labor"], formats)
    indirect_row = row
    row = write_series(ws, row, "Costos indirectos de servicio", model["indirect_total"], formats)
    total_cost_row = row
    row = write_series(
        ws,
        row,
        "TOTAL COSTO DE PRODUCCION",
        model["cost_of_production"],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [direct_row, labor_row, indirect_row], col) for col in range(1, 11)],
    )
    unit = [safe_div(model["cost_of_production"][i], max(1, model["total_new"][i] + sum(model["effective"][key][i] for key in ["B", "E", "A"]))) for i in range(10)]
    unit_cost_row = row
    row = write_series(
        ws,
        row,
        "Costo unitario por museo equivalente",
        unit,
        formats,
        formulas=[
            f"={cell(total_cost_row, col)}/MAX(1,{sheet_cell('COSTOS', COSTOS_IMPLEMENTED_UNITS_ROW, col)}+{sheet_cell('COSTOS', COSTOS_RECURRENT_UNITS_ROW, col)})"
            for col in range(1, 11)
        ],
    )
    row += 1
    row = write_section(ws, row, "CUADRO DE PRODUCCION - UNIDADES EQUIVALENTES", formats)
    row = write_year_header(ws, row, formats)
    implemented_units_row = row
    row = write_series(
        ws,
        row,
        "Museos implementados en el año",
        model["total_new"],
        formats,
        kind="integer",
        formulas=[f"={sheet_cell('COSTOS', COSTOS_IMPLEMENTED_UNITS_ROW, col)}" for col in range(1, 11)],
    )
    recurrent_units_row = row
    recurrent_units = [sum(model["effective"][key][i] for key in ["B", "E", "A"]) for i in range(10)]
    row = write_series(
        ws,
        row,
        "Museos recurrentes equivalentes",
        recurrent_units,
        formats,
        kind="number",
        formulas=[f"={sheet_cell('COSTOS', COSTOS_RECURRENT_UNITS_ROW, col)}" for col in range(1, 11)],
    )
    total_equivalent_units_row = row
    row = write_series(
        ws,
        row,
        "TOTAL MUSEOS EQUIVALENTES",
        [model["total_new"][i] + recurrent_units[i] for i in range(10)],
        formats,
        kind="number",
        total=True,
        formulas=[same_row_formula("SUM", [implemented_units_row, recurrent_units_row], col) for col in range(1, 11)],
    )
    active_units_row = row
    row = write_series(
        ws,
        row,
        "Museos activos al cierre",
        model["total_active"],
        formats,
        kind="integer",
        formulas=[f"={sheet_cell('INGRESOS', 13, col)}" for col in range(1, 11)],
    )

    row += 1
    row = write_section(ws, row, "COMPOSICION DEL COSTO DE PRODUCCION", formats)
    row = write_year_header(ws, row, formats)
    direct_detail_row = row
    row = write_series(
        ws,
        row,
        "Insumos y servicios directos",
        model["direct_total"],
        formats,
        formulas=[f"={cell(direct_row, col)}" for col in range(1, 11)],
    )
    labor_detail_row = row
    row = write_series(
        ws,
        row,
        "Mano de obra directa",
        model["production_labor"],
        formats,
        formulas=[f"={cell(labor_row, col)}" for col in range(1, 11)],
    )
    indirect_detail_row = row
    row = write_series(
        ws,
        row,
        "Costos indirectos de servicio",
        model["indirect_total"],
        formats,
        formulas=[f"={cell(indirect_row, col)}" for col in range(1, 11)],
    )
    detailed_total_cost_row = row
    row = write_series(
        ws,
        row,
        "TOTAL COSTO DE PRODUCCION DETALLADO",
        model["cost_of_production"],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [direct_detail_row, labor_detail_row, indirect_detail_row], col) for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "Costo unitario total por museo equivalente",
        unit,
        formats,
        formulas=[f"=IFERROR({cell(detailed_total_cost_row, col)}/{cell(total_equivalent_units_row, col)},0)" for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "% insumos directos sobre costo",
        [safe_div(model["direct_total"][i], model["cost_of_production"][i]) for i in range(10)],
        formats,
        kind="percent",
        formulas=[f"=IFERROR({cell(direct_detail_row, col)}/{cell(detailed_total_cost_row, col)},0)" for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "% mano de obra directa sobre costo",
        [safe_div(model["production_labor"][i], model["cost_of_production"][i]) for i in range(10)],
        formats,
        kind="percent",
        formulas=[f"=IFERROR({cell(labor_detail_row, col)}/{cell(detailed_total_cost_row, col)},0)" for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "% costos indirectos sobre costo",
        [safe_div(model["indirect_total"][i], model["cost_of_production"][i]) for i in range(10)],
        formats,
        kind="percent",
        formulas=[f"=IFERROR({cell(indirect_detail_row, col)}/{cell(detailed_total_cost_row, col)},0)" for col in range(1, 11)],
    )

    row += 1
    row = write_section(ws, row, "COSTOS UNITARIOS DIRECTOS POR PAQUETE", formats)
    row = write_year_header(ws, row, formats)
    package_cost_columns = {"B": 4, "E": 7, "A": 10}
    for key in ["B", "E", "A"]:
        pkg = PACKAGES[key]
        row = write_series(
            ws,
            row,
            f"Costo directo implementacion - {pkg.name}",
            [pkg.implementation_cost * model["factors"][i] for i in range(10)],
            formats,
            formulas=[
                f"={sheet_cell('COSTOS', 99, package_cost_columns[key], abs_row=True)}"
                f"*{sheet_cell('INGRESOS', 32, col)}"
                for col in range(1, 11)
            ],
        )
    for key in ["B", "E", "A"]:
        pkg = PACKAGES[key]
        row = write_series(
            ws,
            row,
            f"Costo directo recurrente - {pkg.name}",
            [pkg.recurring_cost * model["factors"][i] for i in range(10)],
            formats,
            formulas=[
                f"={sheet_cell('COSTOS', 113, package_cost_columns[key], abs_row=True)}"
                f"*{sheet_cell('INGRESOS', 32, col)}"
                for col in range(1, 11)
            ],
        )

    row += 1
    row = write_section(ws, row, "COSTOS UNITARIOS PROMEDIO POR NATURALEZA", formats)
    row = write_year_header(ws, row, formats)
    row = write_series(
        ws,
        row,
        "Costo directo por museo implementado",
        [safe_div(model["direct_impl_cost"][i], model["total_new"][i]) for i in range(10)],
        formats,
        formulas=[f"=IFERROR({sheet_cell('COSTOS', COSTOS_IMPL_SUBTOTAL_ROW, col)}/{cell(implemented_units_row, col)},0)" for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "Costo directo por museo recurrente equivalente",
        [safe_div(model["direct_rec_cost"][i], recurrent_units[i]) for i in range(10)],
        formats,
        formulas=[f"=IFERROR({sheet_cell('COSTOS', COSTOS_RECURRENT_SUBTOTAL_ROW, col)}/{cell(recurrent_units_row, col)},0)" for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "Mano de obra directa por museo equivalente",
        [safe_div(model["production_labor"][i], model["total_new"][i] + recurrent_units[i]) for i in range(10)],
        formats,
        formulas=[f"=IFERROR({cell(labor_row, col)}/{cell(total_equivalent_units_row, col)},0)" for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "Costo indirecto por museo equivalente",
        [safe_div(model["indirect_total"][i], model["total_new"][i] + recurrent_units[i]) for i in range(10)],
        formats,
        formulas=[f"=IFERROR({cell(indirect_row, col)}/{cell(total_equivalent_units_row, col)},0)" for col in range(1, 11)],
    )
    for col in range(1, 11):
        ws.write_formula(direct_row, col, f"={sheet_cell('COSTOS', COSTOS_DIRECT_TOTAL_ROW, col)}", formats["money"], money(model["direct_total"][col - 1]))
        ws.write_formula(
            labor_row,
            col,
            f"={sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, col)}",
            formats["money"],
            money(model["production_labor"][col - 1]),
        )
        ws.write_formula(indirect_row, col, f"={sheet_cell('COSTOS', COSTOS_INDIRECT_TOTAL_ROW, col)}", formats["money"], money(model["indirect_total"][col - 1]))
        ws.write_formula(total_cost_row, col, f"={cell(detailed_total_cost_row, col)}", formats["total"], money(model["cost_of_production"][col - 1]))
        ws.write_formula(unit_cost_row, col, f"=IFERROR({cell(total_cost_row, col)}/{cell(total_equivalent_units_row, col)},0)", formats["money"], money(unit[col - 1]))
    ws.write(row + 1, 0, "El costo unitario se expresa por museo equivalente: museos implementados en el año mas museos recurrentes ponderados. Las filas inferiores separan insumos, mano de obra e indirectos para explicar la variacion anual.", formats["note"])


def write_planilla(workbook, model, formats, *, admin_factor: float = 1.0, cost_factor: float = 1.0, sheet_name: str = "PLANILLA"):
    ws = workbook.add_worksheet(sheet_name)
    setup(ws, "Prestacion de servicios por año - Planilla MuseIQ", formats, 11)

    production_direct = []
    production_indirect = []
    for i in range(10):
        direct = 0.0
        indirect = 0.0
        for role, monthly in staff_plan(i)["production"]:
            labor = labor_total(monthly * cost_factor)["total"]
            if role == "Lider backend, RAG y plataforma":
                indirect += labor
            else:
                direct += labor
        production_direct.append(direct)
        production_indirect.append(indirect)

    row = 2
    row = write_section(ws, row, "RESUMEN ANUAL DE PLANILLA", formats)
    row = write_year_header(ws, row, formats)
    admin_summary_row = row
    row = write_series(ws, row, "Planilla de administracion", model["admin_labor"], formats)
    production_summary_row = row
    row = write_series(ws, row, "Planilla de produccion MuseIQ", model["production_labor"], formats)
    sales_summary_row = row
    row = write_series(ws, row, "Planilla de ventas", model["sales_labor"], formats)
    total_summary_row = row
    row = write_series(
        ws,
        row,
        "TOTAL PLANILLA",
        [
            model["admin_labor"][i] + model["production_labor"][i] + model["sales_labor"][i]
            for i in range(10)
        ],
        formats,
        total=True,
        formulas=[
            same_row_formula("SUM", [admin_summary_row, production_summary_row, sales_summary_row], col)
            for col in range(1, 11)
        ],
    )
    direct_summary_row = row
    row = write_series(ws, row, "Mano de obra directa de produccion", production_direct, formats)
    indirect_summary_row = row
    row = write_series(ws, row, "Mano de obra indirecta de produccion", production_indirect, formats)

    row += 2
    row = write_section(ws, row, "PARAMETROS DE CALCULO LABORAL", formats)
    ws.write(row, 0, "Concepto", formats["header"])
    ws.write(row, 1, "Tasa / base", formats["header"])
    ws.write(row, 2, "Aplicacion", formats["header"])
    labor_parameters = [
        ("Incremento salarial anual", SALARY_GROWTH, "Sobre sueldo mensual del año anterior"),
        ("Gratificaciones", 2.00, "Dos remuneraciones mensuales por año"),
        ("EsSalud", 0.09, "Sobre sueldo anual"),
        ("CTS", 0.0833, "Sobre sueldo anual mas gratificaciones"),
        ("Vacaciones", 0.50, "Media remuneracion mensual provisionada"),
        ("Cesantia", 14 / 12, "Proporcion de una remuneracion anual"),
    ]
    row += 1
    for label, value, application in labor_parameters:
        ws.write(row, 0, label, formats["label"])
        parameter_format = formats["percent"] if value < 1 else formats["number"]
        if sheet_name == "Riesgo Administrativo":
            ws.write_formula(
                row,
                1,
                f"={sheet_cell('PLANILLA', row, 1)}",
                parameter_format,
                value,
            )
        else:
            ws.write_number(row, 1, value, parameter_format)
        ws.write(row, 2, application, formats["text"])
        row += 1

    row += 2
    annual_refs = []
    previous_role_rows = {
        "admin": {},
        "production": {},
        "sales": {},
    }
    for i, year in enumerate(YEARS):
        row = write_section(ws, row, f"PRESTACION DE SERVICIOS POR AÑO {year}", formats, 10)
        headers = ["CARGO", "SUELDO MENSUAL", "SUELDO ANUAL", "GRATIFICACION JULIO DICIEMBRE", "ESSALUD (9%)", "CTS (8.33%)", "VACACIONES", "PREAVISO", "CESANTIA", "TOTAL C. SOCIAL", "TOTAL"]
        for col, header in enumerate(headers):
            ws.write(row, col, header, formats["header"])
        row += 1
        plan = staff_plan(i)
        section_total_rows = {}
        production_direct_rows = []
        production_indirect_rows = []
        for section, label, factor in [("admin", "Administracion", admin_factor * cost_factor), ("production", "Produccion MuseIQ", cost_factor), ("sales", "Ventas", cost_factor)]:
            ws.write(row, 0, label, formats["subtitle"])
            row += 1
            section_total = 0.0
            section_start = row
            for role, monthly in plan[section]:
                labor = labor_total(monthly * factor)
                ws.write(row, 0, role, formats["label_blue"])
                if sheet_name == "Riesgo Administrativo":
                    salary_formula = f"={sheet_cell('PLANILLA', row, 1)}"
                    if section == "admin":
                        salary_formula += "*(1+$B$275)"
                    ws.write_formula(
                        row,
                        1,
                        salary_formula,
                        formats["money_blue"],
                        money(labor["monthly"]),
                    )
                elif role in previous_role_rows[section]:
                    ws.write_formula(
                        row,
                        1,
                        f"={cell(previous_role_rows[section][role], 1)}*(1+$B$15)",
                        formats["money_blue"],
                        money(labor["monthly"]),
                    )
                else:
                    ws.write_number(
                        row,
                        1,
                        labor["monthly"],
                        formats["money_blue"],
                    )
                previous_role_rows[section][role] = row
                excel_row = row + 1
                labor_formulas = {
                    2: f"=B{excel_row}*12",
                    3: f"=B{excel_row}*$B$16",
                    4: f"=C{excel_row}*$B$17",
                    5: f"=(C{excel_row}+D{excel_row})*$B$18",
                    6: f"=B{excel_row}*$B$19",
                    7: "=0",
                    8: f"=B{excel_row}*$B$20",
                    9: f"=SUM(D{excel_row}:I{excel_row})",
                    10: f"=C{excel_row}+J{excel_row}",
                }
                for col, key in enumerate(["annual", "gratification", "essalud", "cts", "vacations", "notice", "severance", "social", "total"], start=2):
                    ws.write_formula(
                        row,
                        col,
                        labor_formulas[col],
                        formats["money_blue"],
                        money(labor[key]),
                    )
                section_total += labor["total"]
                if section == "production":
                    if role == "Lider backend, RAG y plataforma":
                        production_indirect_rows.append(row)
                    else:
                        production_direct_rows.append(row)
                row += 1
            ws.write(row, 0, f"TOTAL PLANILLA {label.upper()}", formats["total"])
            ws.write_formula(row, 10, f"=SUM(K{section_start + 1}:K{row})", formats["total"], money(section_total))
            section_total_rows[section] = row
            row += 1

        row += 1
        ws.merge_range(
            row,
            0,
            row,
            10,
            "RESUMEN PLANILLA DE PRODUCCION",
            formats["summary_label"],
        )
        row += 1
        indirect_detail_row = row
        ws.merge_range(
            row,
            0,
            row,
            9,
            "Mano de obra indirecta",
            formats["summary_label"],
        )
        ws.write_formula(
            row,
            10,
            "=" + "+".join(cell(source_row, 10) for source_row in production_indirect_rows),
            formats["summary_money"],
            money(production_indirect[i]),
        )
        row += 1
        direct_detail_row = row
        ws.merge_range(
            row,
            0,
            row,
            9,
            "Mano de obra directa",
            formats["summary_label"],
        )
        ws.write_formula(
            row,
            10,
            "=" + "+".join(cell(source_row, 10) for source_row in production_direct_rows),
            formats["summary_money"],
            money(production_direct[i]),
        )
        row += 1
        production_detail_total_row = row
        ws.merge_range(
            row,
            0,
            row,
            9,
            "TOTAL PRODUCCION",
            formats["summary_label"],
        )
        ws.write_formula(
            row,
            10,
            f"=SUM(K{indirect_detail_row + 1}:K{direct_detail_row + 1})",
            formats["summary_total"],
            money(model["production_labor"][i]),
        )
        annual_refs.append(
            {
                "admin": section_total_rows["admin"],
                "production": production_detail_total_row,
                "sales": section_total_rows["sales"],
                "direct": direct_detail_row,
                "indirect": indirect_detail_row,
            }
        )
        row += 1
        ws.set_row(row, 30)
        ws.merge_range(
            row,
            0,
            row,
            10,
            "Politica: incremento salarial anual de 5% y ampliacion gradual del equipo segun la cartera de museos.",
            formats["summary_note"],
        )
        row += 3

    for col, refs in enumerate(annual_refs, start=1):
        ws.write_formula(
            admin_summary_row,
            col,
            f"={cell(refs['admin'], 10)}",
            formats["money"],
            money(model["admin_labor"][col - 1]),
        )
        ws.write_formula(
            production_summary_row,
            col,
            f"={cell(refs['production'], 10)}",
            formats["money"],
            money(model["production_labor"][col - 1]),
        )
        ws.write_formula(
            sales_summary_row,
            col,
            f"={cell(refs['sales'], 10)}",
            formats["money"],
            money(model["sales_labor"][col - 1]),
        )
        ws.write_formula(
            direct_summary_row,
            col,
            f"={cell(refs['direct'], 10)}",
            formats["money"],
            money(production_direct[col - 1]),
        )
        ws.write_formula(
            indirect_summary_row,
            col,
            f"={cell(refs['indirect'], 10)}",
            formats["money"],
            money(production_indirect[col - 1]),
        )

    assert admin_summary_row == PLANILLA_ADMIN_SUMMARY_ROW
    assert production_summary_row == PLANILLA_PRODUCTION_SUMMARY_ROW
    assert sales_summary_row == PLANILLA_SALES_SUMMARY_ROW
    assert total_summary_row == PLANILLA_TOTAL_SUMMARY_ROW
    assert direct_summary_row == PLANILLA_DIRECT_PRODUCTION_ROW
    assert indirect_summary_row == PLANILLA_INDIRECT_PRODUCTION_ROW


def write_expenses_finance(workbook, model, formats):
    ws = workbook.add_worksheet("GASTOS_ADM_VTAS_FINANZAS")
    setup(ws, "Gastos de Venta, Administracion y Finanzas", formats, 12)
    row = 2
    row = write_section(ws, row, "GASTOS DE VENTA", formats)
    row = write_year_header(ws, row, formats)
    sales_labor_row = row
    row = write_series(
        ws,
        row,
        "Sueldo y carga social depto. venta",
        model["sales_labor"],
        formats,
        formulas=[
            f"={sheet_cell('PLANILLA', PLANILLA_SALES_SUMMARY_ROW, col)}"
            for col in range(1, 11)
        ],
    )
    sales_commission_row = row
    row = write_series(
        ws,
        row,
        "Comisiones a vendedores 2.5% ventas",
        model["sales_commission"],
        formats,
        formulas=[
            f"={sheet_cell('INGRESOS', 18, col)}*$B$59"
            for col in range(1, 11)
        ],
    )
    prospect_row = row
    row = write_series(
        ws,
        row,
        "Prospeccion, viajes y representacion",
        [16_000 * (1.04**i) for i in range(10)],
        formats,
        formulas=[
            "=SUM(B29:B31)"
            if col == 1
            else f"={cell(prospect_row, col - 1)}*(1+$B$60)"
            for col in range(1, 11)
        ],
    )
    sales_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL GASTOS DE VENTA",
        [model["sales_labor"][i] + model["sales_other"][i] for i in range(10)],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [sales_labor_row, sales_commission_row, prospect_row], col) for col in range(1, 11)],
    )
    row += 1
    row = write_section(ws, row, "GASTOS DE ADMINISTRACION", formats)
    row = write_year_header(ws, row, formats)
    admin_labor_row = row
    row = write_series(
        ws,
        row,
        "Sueldo y carga social administracion",
        model["admin_labor"],
        formats,
        formulas=[
            f"={sheet_cell('PLANILLA', PLANILLA_ADMIN_SUMMARY_ROW, col)}"
            for col in range(1, 11)
        ],
    )
    admin_services_row = row
    row = write_series(ws, row, "Servicios contables, legales, oficina y comunicaciones", model["admin_other"], formats)
    admin_dep_row = row
    row = write_series(
        ws,
        row,
        "Depreciacion y amortizacion asignada",
        [model["assets"]["depreciation"][i] * 0.40 + model["assets"]["amortization"][i] for i in range(10)],
        formats,
        formulas=[
            f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 104, col)}"
            f"+{sheet_cell('DEPRECIACIONES AMORTIZACION', 105, col)}"
            for col in range(1, 11)
        ],
    )
    admin_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL GASTOS ADMINISTRATIVOS",
        [model["admin_labor"][i] + model["admin_other"][i] + model["assets"]["depreciation"][i] * 0.40 + model["assets"]["amortization"][i] for i in range(10)],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [admin_labor_row, admin_services_row, admin_dep_row], col) for col in range(1, 11)],
    )
    row += 1
    row = write_section(ws, row, "GASTOS FINANCIEROS - AMORTIZACION DEL PRESTAMO", formats)
    row = write_year_header(ws, row, formats)
    initial_debt_row = row
    row = write_series(ws, row, "Saldo inicial de deuda", model["debt"]["initial_balance"], formats)
    interest_row = row
    row = write_series(
        ws,
        row,
        "Intereses",
        model["debt"]["interest"],
        formats,
        formulas=[
            f"={cell(initial_debt_row, col)}*$B$49"
            for col in range(1, 11)
        ],
    )
    principal_row = row
    payment_formula = "PMT($B$49,$B$51,-$B$48)"
    row = write_series(
        ws,
        row,
        "Abono a capital",
        model["debt"]["principal"],
        formats,
        formulas=[
            f"=IF(COLUMN({cell(0, col)})-COLUMN($B$1)+1<=$B$50,0,"
            f"IF(COLUMN({cell(0, col)})-COLUMN($B$1)+1<=$B$50+$B$51,"
            f"MIN({payment_formula}-{cell(interest_row, col)},{cell(initial_debt_row, col)}),0))"
            for col in range(1, 11)
        ],
    )
    installment_row = row
    row = write_series(
        ws,
        row,
        "Cuota anual",
        model["debt"]["installment"],
        formats,
        total=True,
        formulas=[same_row_formula("SUM", [interest_row, principal_row], col) for col in range(1, 11)],
    )
    final_debt_row = row
    row = write_series(
        ws,
        row,
        "Saldo final de deuda",
        model["debt"]["final_balance"],
        formats,
        formulas=[f"=MAX(0,{cell(initial_debt_row, col)}-{cell(principal_row, col)})" for col in range(1, 11)],
    )
    for col in range(1, 11):
        ws.write_formula(
            initial_debt_row,
            col,
            "=$B$48" if col == 1 else f"={cell(final_debt_row, col - 1)}",
            formats["money"],
            money(model["debt"]["initial_balance"][col - 1]),
        )
    ws.write(row + 1, 0, "Condicion adoptada: 2 años de gracia de capital y amortizacion en 6 cuotas anuales; tasa base 18% anual.", formats["note"])

    row += 3
    row = write_section(ws, row, "DETALLE DE GASTOS DE VENTA", formats)
    row = write_year_header(ws, row, formats)
    sales_detail_rows = []
    sales_detail = [
        ("Prospeccion y movilidad comercial", 6_000.0),
        ("Marketing digital y materiales comerciales", 5_000.0),
        ("Viajes y representacion institucional", 5_000.0),
    ]
    for label, base in sales_detail:
        detail_row = row
        sales_detail_rows.append(detail_row)
        row = write_series(
            ws,
            row,
            label,
            [base * (1.04**i) for i in range(10)],
            formats,
            formulas=[
                f"={base}"
                if col == 1
                else f"={cell(detail_row, col - 1)}*(1+$B$60)"
                for col in range(1, 11)
            ],
        )
        ws.write_number(detail_row, 1, base, formats["money"])
    sales_other_detail_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL OTROS GASTOS DE VENTA",
        [16_000 * (1.04**i) for i in range(10)],
        formats,
        total=True,
        formulas=[
            same_row_formula("SUM", sales_detail_rows, col)
            for col in range(1, 11)
        ],
    )

    row += 1
    row = write_section(ws, row, "DETALLE DE GASTOS DE ADMINISTRACION", formats)
    row = write_year_header(ws, row, formats)
    admin_detail_rows = []
    admin_detail = [
        ("Servicios contables y tributarios", 4_800.0),
        ("Asesoria legal y gestion contractual", 2_400.0),
        ("Comunicaciones e internet administrativo", 3_600.0),
        ("Papeleria e insumos administrativos", 2_800.0),
        ("Seguros, comisiones bancarias y tramites", 2_400.0),
    ]
    for label, base in admin_detail:
        detail_row = row
        admin_detail_rows.append(detail_row)
        row = write_series(
            ws,
            row,
            label,
            [base * (1.04**i) for i in range(10)],
            formats,
            formulas=[
                f"={base}"
                if col == 1
                else f"={cell(detail_row, col - 1)}*(1+$B$60)"
                for col in range(1, 11)
            ],
        )
        ws.write_number(detail_row, 1, base, formats["money"])
    admin_services_detail_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL SERVICIOS Y GASTOS ADMINISTRATIVOS",
        model["admin_other"],
        formats,
        total=True,
        formulas=[
            same_row_formula("SUM", admin_detail_rows, col)
            for col in range(1, 11)
        ],
    )
    admin_dep_detail_row = row
    row = write_series(
        ws,
        row,
        "Depreciacion asignada a administracion 40%",
        [model["assets"]["depreciation"][i] * 0.40 for i in range(10)],
        formats,
        formulas=[
            f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 104, col)}"
            for col in range(1, 11)
        ],
    )
    admin_amort_detail_row = row
    row = write_series(
        ws,
        row,
        "Amortizacion de intangibles y diferidos",
        model["assets"]["amortization"],
        formats,
        formulas=[
            f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 105, col)}"
            for col in range(1, 11)
        ],
    )
    admin_non_cash_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL DEPRECIACION Y AMORTIZACION",
        [
            model["assets"]["depreciation"][i] * 0.40 + model["assets"]["amortization"][i]
            for i in range(10)
        ],
        formats,
        total=True,
        formulas=[
            same_row_formula("SUM", [admin_dep_detail_row, admin_amort_detail_row], col)
            for col in range(1, 11)
        ],
    )

    row += 2
    ws.write(row, 0, "PARAMETROS DEL FINANCIAMIENTO", formats["subtitle"])
    ws.write(row, 1, "Valor", formats["header"])
    row += 1
    financing_parameters = [
        ("Monto financiado", model["loan"], "money"),
        ("Tasa efectiva anual", model["interest_rate"], "percent"),
        ("Periodo de gracia de capital", LOAN_GRACE_YEARS, "integer"),
        ("Años de amortizacion", LOAN_TOTAL_YEARS - LOAN_GRACE_YEARS, "integer"),
    ]
    for parameter_index, (label, value, kind) in enumerate(financing_parameters):
        ws.write(row, 0, label, formats["label"])
        if parameter_index == 0:
            ws.write_formula(
                row,
                1,
                f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}",
                formats[kind],
                value,
            )
        else:
            ws.write_number(row, 1, value, formats[kind])
        row += 1
    ws.write(row, 0, "Cuota nivelada durante amortizacion", formats["label"])
    ws.write_formula(
        row,
        1,
        f"=PMT(B{row - 2},B{row},-B{row - 3})",
        formats["money"],
        money(
            model["loan"]
            * model["interest_rate"]
            / (1 - (1 + model["interest_rate"]) ** (-(LOAN_TOTAL_YEARS - LOAN_GRACE_YEARS)))
        ),
    )
    row += 2
    ws.write(row, 0, "Total intereses del financiamiento", formats["label"])
    ws.write_formula(row, 1, f"=SUM(B{interest_row + 1}:K{interest_row + 1})", formats["money"], money(sum(model["debt"]["interest"])))
    row += 1
    ws.write(row, 0, "Total amortizacion de capital", formats["label"])
    ws.write_formula(row, 1, f"=SUM(B{principal_row + 1}:K{principal_row + 1})", formats["money"], money(sum(model["debt"]["principal"])))
    row += 1
    ws.write(row, 0, "Total servicio de deuda", formats["total"])
    ws.write_formula(row, 1, f"=SUM(B{installment_row + 1}:K{installment_row + 1})", formats["total"], money(sum(model["debt"]["installment"])))

    operating_params_title_row = 57
    ws.write(
        operating_params_title_row,
        0,
        "PARAMETROS OPERATIVOS",
        formats["subtitle"],
    )
    ws.write(operating_params_title_row + 1, 0, "Comision sobre ventas", formats["label"])
    ws.write_number(
        operating_params_title_row + 1,
        1,
        0.025,
        formats["percent"],
    )
    ws.write(operating_params_title_row + 2, 0, "Crecimiento anual de gastos", formats["label"])
    ws.write_number(
        operating_params_title_row + 2,
        1,
        0.04,
        formats["percent"],
    )

    for col in range(1, 11):
        ws.write_formula(
            prospect_row,
            col,
            f"={cell(sales_other_detail_total_row, col)}",
            formats["money"],
            money(16_000 * (1.04 ** (col - 1))),
        )
        ws.write_formula(
            admin_services_row,
            col,
            f"={cell(admin_services_detail_total_row, col)}",
            formats["money"],
            money(model["admin_other"][col - 1]),
        )
        ws.write_formula(
            admin_dep_row,
            col,
            f"={cell(admin_non_cash_total_row, col)}",
            formats["money"],
            money(model["assets"]["depreciation"][col - 1] * 0.40 + model["assets"]["amortization"][col - 1]),
        )


def write_depreciation(workbook, model, formats):
    ws = workbook.add_worksheet("DEPRECIACIONES AMORTIZACION")
    setup(ws, "Depreciaciones y Amortizacion", formats, 11)
    ws.set_column(0, 0, 48)
    ws.set_column(1, 10, 14)
    row = 2
    row = write_section(ws, row, "ACTIVOS TANGIBLES", formats)
    ws.write(row, 0, "Activo", formats["header"])
    ws.write(row, 1, "Valor", formats["header"])
    ws.write(row, 2, "Vida util", formats["header"])
    row += 1
    fixed_start_row = row
    for name, value, life in model["assets"]["fixed_assets"]:
        ws.write(row, 0, name, formats["label"])
        ws.write_number(row, 1, value, formats["money"])
        ws.write_number(row, 2, life, formats["integer"])
        row += 1
    fixed_end_row = row - 1
    row += 1
    row = write_year_header(ws, row, formats)
    depreciation_row = row
    row = write_series(
        ws,
        row,
        "Depreciacion anual",
        model["assets"]["depreciation"],
        formats,
        formulas=[
            "="
            + "+".join(
                f"IF({year}<={cell(asset_row, 2, abs_row=True, abs_col=True)},{cell(asset_row, 1, abs_row=True, abs_col=True)}/{cell(asset_row, 2, abs_row=True, abs_col=True)},0)"
                for asset_row in range(fixed_start_row, fixed_end_row + 1)
            )
            for year in YEARS
        ],
    )
    row = write_series(
        ws,
        row,
        "Activo fijo neto",
        model["fixed_net"],
        formats,
        total=True,
        formulas=[
            f"=MAX(0,SUM({cell(fixed_start_row, 1, abs_row=True, abs_col=True)}:{cell(fixed_end_row, 1, abs_row=True, abs_col=True)})-SUM($B${depreciation_row + 1}:{cell(depreciation_row, col)}))"
            for col in range(1, 11)
        ],
    )
    row += 2
    row = write_section(ws, row, "ACTIVOS INTANGIBLES Y DIFERIDOS", formats)
    ws.write(row, 0, "Activo", formats["header"])
    ws.write(row, 1, "Valor", formats["header"])
    ws.write(row, 2, "Vida util", formats["header"])
    row += 1
    intangible_start_row = row
    for name, value, life in model["assets"]["intangibles"]:
        ws.write(row, 0, name, formats["label"])
        ws.write_number(row, 1, value, formats["money"])
        ws.write_number(row, 2, life, formats["integer"])
        row += 1
    intangible_end_row = row - 1
    row += 1
    row = write_year_header(ws, row, formats)
    amortization_row = row
    row = write_series(
        ws,
        row,
        "Amortizacion anual",
        model["assets"]["amortization"],
        formats,
        formulas=[
            "="
            + "+".join(
                f"IF({year}<={cell(asset_row, 2, abs_row=True, abs_col=True)},{cell(asset_row, 1, abs_row=True, abs_col=True)}/{cell(asset_row, 2, abs_row=True, abs_col=True)},0)"
                for asset_row in range(intangible_start_row, intangible_end_row + 1)
            )
            for year in YEARS
        ],
    )
    row = write_series(
        ws,
        row,
        "Activo intangible neto",
        model["intangible_net"],
        formats,
        total=True,
        formulas=[
            f"=MAX(0,SUM({cell(intangible_start_row, 1, abs_row=True, abs_col=True)}:{cell(intangible_end_row, 1, abs_row=True, abs_col=True)})-SUM($B${amortization_row + 1}:{cell(amortization_row, col)}))"
            for col in range(1, 11)
        ],
    )

    row += 2
    row = write_section(ws, row, "CEDULAS DETALLADAS DE DEPRECIACION POR ACTIVO TANGIBLE", formats)
    fixed_headers = [
        "Activo",
        "Cantidad",
        "Costo unitario",
        "Costo total",
        "% residual",
        "Valor residual",
        "Base depreciable",
        "Vida util",
        "Depreciacion anual",
        "Depreciacion acumulada",
        "Valor en libros año 10",
    ]
    fixed_detail = model["assets"]["fixed_asset_detail"]
    category_order = [
        "Equipos de desarrollo y prueba",
        "Servidor y red local",
        "Herramientas de instalacion y medicion",
        "Mobiliario y oficina",
    ]
    fixed_detail_rows = []
    fixed_detail_by_row = {}
    category_total_rows = {}
    category_first_rows = {}
    for category in category_order:
        row = write_section(ws, row, category.upper(), formats)
        for col, header in enumerate(fixed_headers):
            ws.write(row, col, header, formats["header"])
        row += 1
        category_item_rows = []
        category_items = [item for item in fixed_detail if item[0] == category]
        for _, name, quantity, unit_cost, life, residual_rate in category_items:
            asset_row = row
            fixed_detail_rows.append(asset_row)
            fixed_detail_by_row[asset_row] = (name, quantity, unit_cost, life, residual_rate)
            category_item_rows.append(asset_row)
            total_cost = quantity * unit_cost
            residual = total_cost * residual_rate
            annual_dep = (total_cost - residual) / life
            accumulated = annual_dep * min(10, life)
            ws.write(row, 0, name, formats["label"])
            ws.write_number(row, 1, quantity, formats["integer"])
            ws.write_number(row, 2, unit_cost, formats["money"])
            ws.write_formula(row, 3, f"=B{row + 1}*C{row + 1}", formats["money"], money(total_cost))
            ws.write_number(row, 4, residual_rate, formats["percent"])
            ws.write_formula(row, 5, f"=D{row + 1}*E{row + 1}", formats["money"], money(residual))
            ws.write_formula(row, 6, f"=D{row + 1}-F{row + 1}", formats["money"], money(total_cost - residual))
            ws.write_number(row, 7, life, formats["integer"])
            ws.write_formula(row, 8, f"=G{row + 1}/H{row + 1}", formats["money"], money(annual_dep))
            ws.write_formula(row, 9, f"=I{row + 1}*MIN(10,H{row + 1})", formats["money"], money(accumulated))
            ws.write_formula(row, 10, f"=MAX(F{row + 1},D{row + 1}-J{row + 1})", formats["money"], money(max(residual, total_cost - accumulated)))
            row += 1
        category_first_rows[category] = category_item_rows[0]
        category_total_row = row
        category_total_rows[category] = category_total_row
        category_cost = sum(item[2] * item[3] for item in category_items)
        category_residual = sum(item[2] * item[3] * item[5] for item in category_items)
        category_annual = sum(item[2] * item[3] * (1 - item[5]) / item[4] for item in category_items)
        ws.write(row, 0, f"TOTAL {category.upper()}", formats["total"])
        for col, cached in [
            (3, category_cost),
            (5, category_residual),
            (6, category_cost - category_residual),
            (8, category_annual),
            (9, category_cost - category_residual),
            (10, category_residual),
        ]:
            ws.write_formula(
                row,
                col,
                f"=SUM({cell(category_item_rows[0], col)}:{cell(category_item_rows[-1], col)})",
                formats["total"],
                money(cached),
            )
        row += 2

    fixed_detail_total_row = row
    ws.write(row, 0, "TOTAL ACTIVOS TANGIBLES", formats["total"])
    total_residual = sum(item[2] * item[3] * item[5] for item in fixed_detail)
    total_depreciable = TANGIBLE_ASSETS - total_residual
    for col in [3, 5, 6, 8, 9, 10]:
        ws.write_formula(
            row,
            col,
            "="
            + "+".join(
                cell(category_total_rows[category], col)
                for category in category_order
            ),
            formats["total"],
            money(
                {
                    3: TANGIBLE_ASSETS,
                    5: total_residual,
                    6: total_depreciable,
                    8: model["assets"]["depreciation"][0],
                    9: total_depreciable,
                    10: total_residual,
                }[col]
            ),
        )

    row += 2
    row = write_section(ws, row, "PROGRAMA ANUAL DE DEPRECIACION", formats)
    row = write_year_header(ws, row, formats)
    fixed_schedule_rows = []
    for source_row in fixed_detail_rows:
        schedule_row = row
        fixed_schedule_rows.append(schedule_row)
        name, quantity, unit_cost, life, residual_rate = fixed_detail_by_row[source_row]
        ws.write(row, 0, f"Depreciacion - {name}", formats["label"])
        annual_dep = quantity * unit_cost * (1 - residual_rate) / life
        for col, year in enumerate(YEARS, start=1):
            ws.write_formula(
                row,
                col,
                f"=IF({year}<={cell(source_row, 7, abs_row=True, abs_col=True)},{cell(source_row, 8, abs_row=True, abs_col=True)},0)",
                formats["money"],
                money(annual_dep if year <= life else 0.0),
            )
        row += 1
    fixed_schedule_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL DEPRECIACION ANUAL",
        model["assets"]["depreciation"],
        formats,
        total=True,
        formulas=[
            f"=SUM({cell(fixed_schedule_rows[0], col)}:{cell(fixed_schedule_rows[-1], col)})"
            for col in range(1, 11)
        ],
    )

    row += 1
    row = write_section(ws, row, "CEDULA DE AMORTIZACION POR ACTIVO INTANGIBLE", formats)
    intangible_headers = [
        "Activo",
        "Cantidad",
        "Costo unitario",
        "Costo total",
        "% residual",
        "Valor residual",
        "Base amortizable",
        "Vida util",
        "Amortizacion anual",
        "Amortizacion acumulada",
        "Valor en libros año 10",
    ]
    for col, header in enumerate(intangible_headers):
        ws.write(row, col, header, formats["header"])
    row += 1
    intangible_detail = [
        ("Constitucion y formalizacion", 1, 8_000.0, 5, 0.0),
        ("Desarrollo base app/API/RAG", 1, 90_000.0, 5, 0.0),
        ("Documentacion y manuales", 1, 22_000.0, 5, 0.0),
        ("Marca, preventa y dossier comercial", 1, 12_000.0, 5, 0.0),
    ]
    intangible_detail_rows = []
    for name, quantity, unit_cost, life, residual_rate in intangible_detail:
        asset_row = row
        intangible_detail_rows.append(asset_row)
        total_cost = quantity * unit_cost
        residual = total_cost * residual_rate
        annual_amort = (total_cost - residual) / life
        accumulated = annual_amort * min(10, life)
        ws.write(row, 0, name, formats["label"])
        ws.write_number(row, 1, quantity, formats["integer"])
        ws.write_number(row, 2, unit_cost, formats["money"])
        ws.write_formula(row, 3, f"=B{row + 1}*C{row + 1}", formats["money"], money(total_cost))
        ws.write_number(row, 4, residual_rate, formats["percent"])
        ws.write_formula(row, 5, f"=D{row + 1}*E{row + 1}", formats["money"], money(residual))
        ws.write_formula(row, 6, f"=D{row + 1}-F{row + 1}", formats["money"], money(total_cost - residual))
        ws.write_number(row, 7, life, formats["integer"])
        ws.write_formula(row, 8, f"=G{row + 1}/H{row + 1}", formats["money"], money(annual_amort))
        ws.write_formula(row, 9, f"=I{row + 1}*MIN(10,H{row + 1})", formats["money"], money(accumulated))
        ws.write_formula(row, 10, f"=MAX(F{row + 1},D{row + 1}-J{row + 1})", formats["money"], money(max(residual, total_cost - accumulated)))
        row += 1
    intangible_detail_total_row = row
    ws.write(row, 0, "TOTAL ACTIVOS INTANGIBLES", formats["total"])
    for col in [3, 5, 6, 8, 9, 10]:
        ws.write_formula(
            row,
            col,
            f"=SUM({cell(intangible_detail_rows[0], col)}:{cell(intangible_detail_rows[-1], col)})",
            formats["total"],
            money(
                {
                    3: INTANGIBLE_ASSETS,
                    5: 0.0,
                    6: INTANGIBLE_ASSETS,
                    8: model["assets"]["amortization"][0],
                    9: INTANGIBLE_ASSETS,
                    10: 0.0,
                }[col]
            ),
        )

    row += 2
    row = write_section(ws, row, "PROGRAMA ANUAL DE AMORTIZACION", formats)
    row = write_year_header(ws, row, formats)
    intangible_schedule_rows = []
    for index, source_row in enumerate(intangible_detail_rows):
        schedule_row = row
        intangible_schedule_rows.append(schedule_row)
        ws.write(row, 0, f"Amortizacion - {intangible_detail[index][0]}", formats["label"])
        life = intangible_detail[index][3]
        annual_amort = intangible_detail[index][1] * intangible_detail[index][2] / life
        for col, year in enumerate(YEARS, start=1):
            ws.write_formula(
                row,
                col,
                f"=IF({year}<={cell(source_row, 7, abs_row=True, abs_col=True)},{cell(source_row, 8, abs_row=True, abs_col=True)},0)",
                formats["money"],
                money(annual_amort if year <= life else 0.0),
            )
        row += 1
    intangible_schedule_total_row = row
    row = write_series(
        ws,
        row,
        "TOTAL AMORTIZACION ANUAL",
        model["assets"]["amortization"],
        formats,
        total=True,
        formulas=[
            f"=SUM({cell(intangible_schedule_rows[0], col)}:{cell(intangible_schedule_rows[-1], col)})"
            for col in range(1, 11)
        ],
    )

    row += 1
    row = write_section(ws, row, "ASIGNACION CONTABLE Y VALOR TERMINAL", formats)
    row = write_year_header(ws, row, formats)
    allocation_policy_title_row = row + 6
    production_allocation_policy_row = row + 7
    admin_allocation_policy_row = row + 8
    row = write_series(
        ws,
        row,
        "Depreciacion asignada a produccion 60%",
        [value * 0.60 for value in model["assets"]["depreciation"]],
        formats,
        formulas=[
            f"={cell(fixed_schedule_total_row, col)}"
            f"*{cell(production_allocation_policy_row, 1, abs_row=True, abs_col=True)}"
            for col in range(1, 11)
        ],
    )
    row = write_series(
        ws,
        row,
        "Depreciacion asignada a administracion 40%",
        [value * 0.40 for value in model["assets"]["depreciation"]],
        formats,
        formulas=[
            f"={cell(fixed_schedule_total_row, col)}"
            f"*{cell(admin_allocation_policy_row, 1, abs_row=True, abs_col=True)}"
            for col in range(1, 11)
        ],
    )
    row = write_series(
        ws,
        row,
        "Amortizacion asignada a administracion 100%",
        model["assets"]["amortization"],
        formats,
        formulas=[f"={cell(intangible_schedule_total_row, col)}" for col in range(1, 11)],
    )
    row += 1
    ws.write(row, 0, "Valor de salvamento comercial considerado en el flujo final", formats["label"])
    ws.write_number(row, 1, SALVAGE_VALUE, formats["money"])
    ws.write(row, 2, "Estimacion comercial separada del valor contable neto.", formats["note"])
    ws.write(
        allocation_policy_title_row,
        0,
        "POLITICA DE ASIGNACION DE DEPRECIACION",
        formats["subtitle"],
    )
    ws.write(
        production_allocation_policy_row,
        0,
        "Asignacion a produccion",
        formats["label"],
    )
    ws.write_number(
        production_allocation_policy_row,
        1,
        0.60,
        formats["percent"],
    )
    ws.write(
        admin_allocation_policy_row,
        0,
        "Asignacion a administracion",
        formats["label"],
    )
    ws.write_formula(
        admin_allocation_policy_row,
        1,
        f"=1-{cell(production_allocation_policy_row, 1)}",
        formats["percent"],
        0.40,
    )

    category_to_total_row = {
        "Laptops y equipos de prueba": category_total_rows["Equipos de desarrollo y prueba"],
        "Servidor/NAS y red local": category_total_rows["Servidor y red local"],
        "Herramientas de instalacion y medicion": category_total_rows["Herramientas de instalacion y medicion"],
        "Mobiliario y oficina": category_total_rows["Mobiliario y oficina"],
    }
    category_to_first_row = {
        "Laptops y equipos de prueba": category_first_rows["Equipos de desarrollo y prueba"],
        "Servidor/NAS y red local": category_first_rows["Servidor y red local"],
        "Herramientas de instalacion y medicion": category_first_rows["Herramientas de instalacion y medicion"],
        "Mobiliario y oficina": category_first_rows["Mobiliario y oficina"],
    }
    for offset, (summary_name, summary_value, summary_life) in enumerate(model["assets"]["fixed_assets"]):
        summary_row = fixed_start_row + offset
        category_total_row = category_to_total_row[summary_name]
        ws.write_formula(summary_row, 1, f"={cell(category_total_row, 3)}", formats["money"], money(summary_value))
        ws.write_formula(
            summary_row,
            2,
            f"={cell(category_to_first_row[summary_name], 7)}",
            formats["integer"],
            summary_life,
        )
    for col in range(1, 11):
        ws.write_formula(
            depreciation_row,
            col,
            f"={cell(fixed_schedule_total_row, col)}",
            formats["money"],
            money(model["assets"]["depreciation"][col - 1]),
        )

    for offset, source_row in enumerate(intangible_detail_rows):
        summary_row = intangible_start_row + offset
        ws.write_formula(summary_row, 1, f"={cell(source_row, 3)}", formats["money"], money(model["assets"]["intangibles"][offset][1]))
        ws.write_formula(summary_row, 2, f"={cell(source_row, 7)}", formats["integer"], model["assets"]["intangibles"][offset][2])
    for col in range(1, 11):
        ws.write_formula(
            amortization_row,
            col,
            f"={cell(intangible_schedule_total_row, col)}",
            formats["money"],
            money(model["assets"]["amortization"][col - 1]),
        )


def write_investment_plan(workbook, model, formats):
    ws = workbook.add_worksheet("PLAN DE INVERSION")
    setup(ws, "Plan de Inversion - MuseIQ", formats, 6)
    ws.set_column(0, 1, 24)
    ws.set_column(2, 2, 16)
    ws.set_column(3, 6, 16)
    ws.freeze_panes(4, 2)

    fixed_equity_rate = 0.20
    working_equity_rate = 1.00
    intangible_equity_rate = (
        model["equity"] - TANGIBLE_ASSETS * fixed_equity_rate - INITIAL_CASH
    ) / INTANGIBLE_ASSETS
    policy_title_row = 39
    fixed_equity_policy_row = 40
    working_equity_policy_row = 41
    intangible_equity_policy_row = 42
    advance_policy_row = 43
    minimum_cash_policy_row = 44
    total_equity_policy_row = 45
    fixed_equity_ref = cell(
        fixed_equity_policy_row, 6, abs_row=True, abs_col=True
    )
    working_equity_ref = cell(
        working_equity_policy_row, 6, abs_row=True, abs_col=True
    )
    intangible_equity_ref = cell(
        intangible_equity_policy_row, 6, abs_row=True, abs_col=True
    )

    ws.merge_range(2, 0, 2, 1, "ACTIVOS", formats["header"])
    ws.write(2, 2, "MONTO", formats["header"])
    ws.merge_range(2, 3, 2, 4, "FONDOS PROPIOS", formats["header"])
    ws.merge_range(2, 5, 2, 6, "FINANCIAMIENTO", formats["header"])

    def write_main_line(
        row,
        label,
        value,
        formula,
        equity_rate,
        equity_rate_ref,
        *,
        total=False,
    ):
        label_fmt = formats["group_label"] if total else formats["label"]
        value_fmt = formats["group_money"] if total else formats["money"]
        ws.merge_range(row, 0, row, 1, label, label_fmt)
        if formula:
            ws.write_formula(row, 2, formula, value_fmt, money(value))
        else:
            ws.write_number(row, 2, value, value_fmt)
        ws.merge_range(row, 3, row, 4, "", value_fmt)
        ws.write_formula(
            row,
            3,
            f"=C{row + 1}*{equity_rate_ref}",
            value_fmt,
            money(value * equity_rate),
        )
        ws.merge_range(row, 5, row, 6, "", value_fmt)
        ws.write_formula(row, 5, f"=C{row + 1}-D{row + 1}", value_fmt, money(value * (1 - equity_rate)))

    fixed_values = model["assets"]["fixed_assets"]
    write_main_line(
        4,
        "I. ACTIVOS FIJOS",
        TANGIBLE_ASSETS,
        "=SUM(C6:C9)",
        fixed_equity_rate,
        fixed_equity_ref,
        total=True,
    )
    for index, (name, value, _) in enumerate(fixed_values):
        write_main_line(
            5 + index,
            name,
            value,
            f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 4 + index, 1)}",
            fixed_equity_rate,
            fixed_equity_ref,
        )

    working_days = 81
    production_cash = (
        model["direct_total"][0] + model["production_labor"][0] + model["indirect_cash"][0]
    ) / 365 * working_days
    admin_sales_payroll = (
        model["admin_labor"][0] + model["sales_labor"][0]
    ) / 365 * working_days
    sales_cash_expense = model["sales_other"][0] / 365 * working_days
    admin_cash_expense = model["admin_other"][0] / 365 * working_days
    cash_advance = -model["sales_total"][0] * 0.05
    supplier_credit = -model["purchase_credit"][0] / 365 * working_days
    working_adjustment = INITIAL_CASH - sum(
        [
            production_cash,
            admin_sales_payroll,
            sales_cash_expense,
            admin_cash_expense,
            cash_advance,
            supplier_credit,
        ]
    )
    working_items = [
        (
            "Costos de produccion",
            production_cash,
            f"=({sheet_cell('COSTOS', COSTOS_DIRECT_TOTAL_ROW, 1)}"
            f"+{sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, 1)}"
            f"+{sheet_cell('COSTOS', COSTOS_INDIRECT_TOTAL_ROW, 1)}"
            f"-{sheet_cell('DEPRECIACIONES AMORTIZACION', 103, 1)})/365*$B$38",
        ),
        (
            "Sueldos y carga social de administracion y ventas",
            admin_sales_payroll,
            f"=({sheet_cell('PLANILLA', PLANILLA_ADMIN_SUMMARY_ROW, 1)}+{sheet_cell('PLANILLA', PLANILLA_SALES_SUMMARY_ROW, 1)})/365*$B$38",
        ),
        (
            "Gastos de venta desembolsables",
            sales_cash_expense,
            f"=({sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 7, 1)}-{sheet_cell('PLANILLA', PLANILLA_SALES_SUMMARY_ROW, 1)})/365*$B$38",
        ),
        (
            "Gastos de administracion desembolsables",
            admin_cash_expense,
            f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 12, 1)}/365*$B$38",
        ),
        (
            "Menos anticipo contractual de ventas 5%",
            cash_advance,
            f"=-{sheet_cell('INGRESOS', 18, 1)}"
            f"*{cell(advance_policy_row, 6, abs_row=True, abs_col=True)}",
        ),
        (
            "Menos compras financiadas por proveedores",
            supplier_credit,
            f"=-{sheet_cell('COSTOS', 58, 1)}/365*$B$38",
        ),
        (
            "Ajuste de reserva minima de caja",
            working_adjustment,
            f"={cell(minimum_cash_policy_row, 6, abs_row=True, abs_col=True)}"
            f"-SUM(C12:C17)",
        ),
    ]
    write_main_line(
        10,
        "II. CAPITAL DE TRABAJO",
        INITIAL_CASH,
        "=SUM(C12:C18)",
        working_equity_rate,
        working_equity_ref,
        total=True,
    )
    for offset, (label, value, formula) in enumerate(working_items):
        write_main_line(
            11 + offset,
            label,
            value,
            formula,
            working_equity_rate,
            working_equity_ref,
        )

    intangible_values = model["assets"]["intangibles"]
    write_main_line(
        19,
        "III. ACTIVOS NOMINALES",
        INTANGIBLE_ASSETS,
        "=SUM(C21:C24)",
        intangible_equity_rate,
        intangible_equity_ref,
        total=True,
    )
    for index, (name, value, _) in enumerate(intangible_values):
        write_main_line(
            20 + index,
            name,
            value,
            f"=$G${31 + index}",
            intangible_equity_rate,
            intangible_equity_ref,
        )

    ws.merge_range(25, 0, 26, 1, "TOTAL", formats["total"])
    ws.write_blank(25, 2, None, formats["total"])
    ws.merge_range(25, 3, 25, 4, "", formats["total_pct"])
    ws.write_formula(25, 3, "=D27/C27", formats["total_pct"], EQUITY_SHARE)
    ws.merge_range(25, 5, 25, 6, "", formats["total_pct"])
    ws.write_formula(25, 5, "=F27/C27", formats["total_pct"], LOAN_SHARE)

    total_investment_row = 26
    ws.write_formula(total_investment_row, 2, "=SUM(C5,C11,C20)", formats["total"], money(INITIAL_INVESTMENT))
    ws.merge_range(total_investment_row, 3, total_investment_row, 4, "", formats["total"])
    ws.write_formula(total_investment_row, 3, "=SUM(D5,D11,D20)", formats["total"], money(model["equity"]))
    ws.merge_range(total_investment_row, 5, total_investment_row, 6, "", formats["total"])
    ws.write_formula(total_investment_row, 5, "=SUM(F5,F11,F20)", formats["total"], money(model["loan"]))

    ws.merge_range(30, 0, 30, 1, "CALCULO DIAS CAPITAL DE TRABAJO", formats["subtitle"])
    day_rows = [
        ("Tiempo de diagnostico y produccion inicial", 5),
        ("Tiempo de contratacion de insumos", 5),
        ("Tiempo promedio de cobro", 45),
        ("Tiempo de entrega y conformidad", 20),
        ("Dias de seguridad", 6),
    ]
    for offset, (label, days) in enumerate(day_rows):
        helper_row = 31 + offset
        ws.write(helper_row, 0, label, formats["label"])
        ws.write_number(helper_row, 1, days, formats["integer"])
    ws.write(37, 0, "TOTAL DIAS", formats["total"])
    ws.write_formula(37, 1, "=SUM(B32:B36)", formats["total_num"], working_days)

    ws.merge_range(29, 4, 29, 6, "ACTIVOS NOMINALES", formats["subtitle"])
    for index, (name, value, _) in enumerate(intangible_values):
        helper_row = 30 + index
        ws.merge_range(helper_row, 4, helper_row, 5, name, formats["label"])
        ws.write_formula(
            helper_row,
            6,
            f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 16 + index, 1)}",
            formats["money"],
            money(value),
        )
    ws.merge_range(35, 4, 35, 5, "TOTAL ACTIVOS NOMINALES", formats["total"])
    ws.write_formula(35, 6, "=SUM(G31:G34)", formats["total"], INTANGIBLE_ASSETS)

    ws.merge_range(
        39,
        0,
        41,
        2,
        "El capital de trabajo considera el ciclo de diagnostico, contratacion, entrega, conformidad y cobro institucional. El ajuste final mantiene la caja inicial aprobada en S/50,000.",
        formats["note"],
    )
    ws.merge_range(
        policy_title_row,
        4,
        policy_title_row,
        6,
        "POLITICA DE FINANCIAMIENTO Y CAPITAL DE TRABAJO",
        formats["subtitle"],
    )
    policy_rows = [
        (
            fixed_equity_policy_row,
            "Fondos propios en activos fijos",
            fixed_equity_rate,
            "percent",
        ),
        (
            working_equity_policy_row,
            "Fondos propios en capital de trabajo",
            working_equity_rate,
            "percent",
        ),
        (
            advance_policy_row,
            "Anticipo contractual sobre ventas",
            0.05,
            "percent",
        ),
        (
            minimum_cash_policy_row,
            "Reserva minima de caja",
            INITIAL_CASH,
            "money",
        ),
        (
            total_equity_policy_row,
            "Participacion total objetivo de fondos propios",
            EQUITY_SHARE,
            "percent",
        ),
    ]
    for policy_row, label, value, kind in policy_rows:
        ws.merge_range(policy_row, 4, policy_row, 5, label, formats["label"])
        ws.write_number(policy_row, 6, value, formats[kind])
    ws.merge_range(
        intangible_equity_policy_row,
        4,
        intangible_equity_policy_row,
        5,
        "Fondos propios en activos nominales",
        formats["label"],
    )
    ws.write_formula(
        intangible_equity_policy_row,
        6,
        f"=(C27*{cell(total_equity_policy_row, 6, abs_row=True, abs_col=True)}"
        f"-C5*{fixed_equity_ref}-C11*{working_equity_ref})/C20",
        formats["percent"],
        intangible_equity_rate,
    )

    assert total_investment_row == PLAN_INV_EQUITY_TOTAL_ROW
    assert total_investment_row == PLAN_INV_FINANCING_TOTAL_ROW


def write_opening_balance(workbook, model, formats):
    ws = workbook.add_worksheet("BALANCE DE APERTURA")
    setup(ws, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats, 4)
    ws.set_column(0, 0, 34)
    ws.set_column(1, 1, 18)
    ws.set_column(2, 2, 5)
    ws.set_column(3, 3, 34)
    ws.set_column(4, 4, 18)
    ws.freeze_panes(4, 0)
    ws.merge_range(1, 0, 1, 4, "BALANCE DE APERTURA", formats["caption"])
    ws.merge_range(2, 0, 2, 4, "", formats["caption"])

    ws.merge_range(3, 0, 3, 1, "ACTIVOS", formats["header"])
    ws.merge_range(3, 3, 3, 4, "PASIVOS Y PATRIMONIO", formats["header"])

    ws.write(4, 0, "Circulante:", formats["group_label"])
    ws.write(4, 3, "Pasivo:", formats["group_label"])
    ws.write(5, 0, "Caja y bancos", formats["label"])
    ws.write_formula(5, 1, f"={sheet_cell('PLAN DE INVERSION', 10, 2)}", formats["money"], INITIAL_CASH)
    ws.write(6, 0, "Inventario", formats["label"])
    ws.write_formula(
        6,
        1,
        f"={sheet_cell('PLAN DE INVERSION', 10, 2)}-{sheet_cell('PLAN DE INVERSION', 10, 2)}",
        formats["money"],
        0,
    )
    ws.write(8, 0, "Total circulante", formats["total"])
    ws.write_formula(
        8,
        1,
        f"={sheet_cell('PLAN DE INVERSION', 10, 2)}",
        formats["total"],
        INITIAL_CASH,
    )

    ws.write(5, 3, "Deuda de corto plazo", formats["label"])
    ws.write_formula(
        5,
        4,
        f"=MIN({sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)},{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 20, 1)})",
        formats["money"],
        money(model["debt"]["principal"][0]),
    )
    ws.write(6, 3, "Deuda de largo plazo", formats["label"])
    ws.write_formula(
        6,
        4,
        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}-E6",
        formats["money"],
        money(model["loan"] - model["debt"]["principal"][0]),
    )
    ws.write(8, 3, "Total pasivo", formats["total"])
    ws.write_formula(
        8,
        4,
        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}",
        formats["total"],
        money(model["loan"]),
    )

    ws.write(10, 0, "Fijo:", formats["group_label"])
    ws.write(10, 3, "Patrimonio:", formats["group_label"])
    for index, (name, value, _) in enumerate(model["assets"]["fixed_assets"]):
        excel_row = 11 + index
        ws.write(excel_row, 0, name, formats["label"])
        ws.write_formula(excel_row, 1, f"={sheet_cell('PLAN DE INVERSION', 5 + index, 2)}", formats["money"], money(value))
    ws.write(16, 0, "Total fijo", formats["total"])
    ws.write_formula(
        16,
        1,
        f"={sheet_cell('PLAN DE INVERSION', 4, 2)}",
        formats["total"],
        TANGIBLE_ASSETS,
    )

    ws.write(12, 3, "Capital social aportado", formats["label"])
    ws.write_formula(12, 4, f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}", formats["money"], money(model["equity"]))

    ws.write(18, 0, "Diferido:", formats["group_label"])
    for index, (name, value, _) in enumerate(model["assets"]["intangibles"]):
        excel_row = 19 + index
        ws.write(excel_row, 0, name, formats["label"])
        ws.write_formula(excel_row, 1, f"={sheet_cell('PLAN DE INVERSION', 20 + index, 2)}", formats["money"], money(value))
    ws.write(23, 0, "Total diferido", formats["total"])
    ws.write_formula(
        23,
        1,
        f"={sheet_cell('PLAN DE INVERSION', 19, 2)}",
        formats["total"],
        INTANGIBLE_ASSETS,
    )
    ws.write(23, 3, "Total patrimonio", formats["total"])
    ws.write_formula(
        23,
        4,
        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}",
        formats["total"],
        money(model["equity"]),
    )

    ws.write(25, 0, "TOTAL DE ACTIVOS", formats["total"])
    ws.write_formula(
        25,
        1,
        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 2)}",
        formats["total"],
        INITIAL_INVESTMENT,
    )
    ws.write(25, 3, "TOTAL PASIVO Y PATRIMONIO", formats["total"])
    ws.write_formula(
        25,
        4,
        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}+{sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}",
        formats["total"],
        INITIAL_INVESTMENT,
    )
    ws.write(26, 3, "Diferencia de cuadre", formats["label"])
    ws.write_formula(26, 4, "=B26-E26", formats["money"], 0)

    ws.write(32, 0, "Firma representante legal", formats["note"])
    ws.merge_range(32, 3, 32, 4, "Firma responsable contable", formats["note"])


def write_cost_of_sales(workbook, model, formats):
    ws = workbook.add_worksheet("COSTO DE VENTA")
    setup(ws, "Costo de lo Vendido", formats, 11)
    ws.set_column(0, 0, 4)
    ws.set_column(1, 1, 34)
    ws.set_column(2, 11, 13)
    row = 3
    ws.write(row, 0, "", formats["header"])
    ws.write(row, 1, "Concepto", formats["header"])
    for col, label in enumerate(YEAR_LABELS, start=2):
        ws.write(row, col, label, formats["header"])
    row += 1
    zero = [0] * 10
    initial_inventory_row = row
    ws.write(row, 0, "", formats["label"])
    ws.write(row, 1, "Inventario inicial", formats["label"])
    for col in range(2, 12):
        formula = "=0" if col == 2 else f"={cell(7, col - 1)}"
        ws.write_formula(row, col, formula, formats["money"], 0)
    row += 1
    production_cost_row = row
    ws.write(row, 0, "+", formats["group_label"])
    ws.write(row, 1, "Costo de produccion", formats["label"])
    for col in range(2, 12):
        ws.write_formula(
            row,
            col,
            f"={sheet_cell('COSTO DE PRODUCCION UNITARIO', 7, col - 1)}",
            formats["money"],
            money(model["cost_of_production"][col - 2]),
        )
    row += 1
    available_product_row = row
    ws.write_string(row, 0, "=", formats["group_label"])
    ws.write(row, 1, "Producto disponible para la venta", formats["group_label"])
    for col in range(2, 12):
        ws.write_formula(
            row,
            col,
            same_row_formula("SUM", [initial_inventory_row, production_cost_row], col),
            formats["group_money"],
            money(model["cost_of_production"][col - 2]),
        )
    row += 1
    final_inventory_row = row
    ws.write(row, 0, "-", formats["group_label"])
    ws.write(row, 1, "Inventario final", formats["label"])
    for col in range(2, 12):
        ws.write_formula(row, col, "=0", formats["money"], 0)
    row += 1
    cost_sales_row = row
    ws.write_string(row, 0, "=", formats["total"])
    ws.write(row, 1, "Costo de venta", formats["total"])
    for col in range(2, 12):
        ws.write_formula(
            row,
            col,
            f"={cell(available_product_row, col)}-{cell(final_inventory_row, col)}",
            formats["total"],
            money(model["cost_of_sales"][col - 2]),
        )
    row += 1

    row += 1
    row = write_section(ws, row, "COMPOSICION DEL COSTO DE PRODUCCION TRANSFERIDO", formats, 11)
    ws.write(row, 0, "", formats["header"])
    ws.write(row, 1, "Concepto", formats["header"])
    for col, label in enumerate(YEAR_LABELS, start=2):
        ws.write(row, col, label, formats["header"])
    row += 1
    direct_row = row
    detail_rows = [
        ("Insumos y servicios directos", model["direct_total"], 4),
        ("Mano de obra directa", model["production_labor"], 5),
        ("Costos indirectos de servicio", model["indirect_total"], 6),
    ]
    detail_source_rows = []
    for label, values, source_row in detail_rows:
        detail_source_rows.append(row)
        ws.write(row, 0, "", formats["label"])
        ws.write(row, 1, label, formats["label"])
        for col in range(2, 12):
            ws.write_formula(
                row,
                col,
                f"={sheet_cell('COSTO DE PRODUCCION UNITARIO', source_row, col - 1)}",
                formats["money"],
                money(values[col - 2]),
            )
        row += 1
    transferred_total_row = row
    ws.write_string(row, 0, "=", formats["total"])
    ws.write(row, 1, "TOTAL COSTO TRANSFERIDO", formats["total"])
    for col in range(2, 12):
        ws.write_formula(
            row,
            col,
            same_row_formula("SUM", detail_source_rows, col),
            formats["total"],
            money(model["cost_of_production"][col - 2]),
        )
    row += 1
    ws.write(row, 0, "", formats["label"])
    ws.write(row, 1, "Diferencia contra costo de venta", formats["label"])
    for col in range(2, 12):
        ws.write_formula(
            row,
            col,
            f"={cell(cost_sales_row, col)}-{cell(transferred_total_row, col)}",
            formats["money"],
            0,
        )
    row += 1

    row += 1
    ws.merge_range(row, 0, row, 1, "Politica de inventarios", formats["subtitle"])
    ws.merge_range(
        row + 1,
        0,
        row + 2,
        11,
        "MuseIQ presta un servicio por proyecto y no mantiene productos terminados. Por ello, el inventario final es cero y el costo de produccion se reconoce como costo de venta en el mismo año.",
        formats["note"],
    )

    assert cost_sales_row == COST_OF_SALES_TOTAL_ROW


def write_income_statement(workbook, model, formats):
    ws = workbook.add_worksheet("ESTADO DE RESULTADOS")
    setup(ws, "Estado de Resultados", formats, 11)
    ws.merge_range(1, 0, 1, 10, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats["caption"])
    ws.merge_range(2, 0, 3, 0, "Descripcion", formats["header"])
    for col, label in enumerate(YEAR_LABELS, start=1):
        ws.merge_range(2, col, 3, col, label, formats["header"])
    row = 5
    sales_row = row
    row = write_series(ws, row, "Ventas totales", model["sales_total"], formats)
    cost_sales_row = row
    row = write_series(ws, row, "Costo de venta", model["cost_of_sales"], formats)
    gross_profit_row = row
    row = write_series(
        ws,
        row,
        "Utilidad bruta en ventas",
        model["gross_profit"],
        formats,
        total=True,
        formulas=[f"={cell(sales_row, col)}-{cell(cost_sales_row, col)}" for col in range(1, 11)],
    )
    operating_expenses_row = row
    row = write_series(ws, row, "Total gastos de operacion", model["operating_expenses"], formats)
    sales_expense_row = row
    row = write_series(ws, row, "Gastos de venta", [model["sales_labor"][i] + model["sales_other"][i] for i in range(10)], formats)
    admin_expense_row = row
    row = write_series(ws, row, "Gastos de administracion", [model["admin_labor"][i] + model["admin_other"][i] + model["assets"]["depreciation"][i] * 0.40 + model["assets"]["amortization"][i] for i in range(10)], formats)
    row += 1
    ebit_row = row
    row = write_series(
        ws,
        row,
        "Utilidad de operacion",
        model["ebit"],
        formats,
        total=True,
        formulas=[f"={cell(gross_profit_row, col)}-{cell(operating_expenses_row, col)}" for col in range(1, 11)],
    )
    interest_row = row
    row = write_series(ws, row, "Gastos financieros", model["debt"]["interest"], formats)
    other_income_row = row
    row = write_series(
        ws,
        row,
        "Otros ingresos",
        [0] * 10,
        formats,
        formulas=["=0" for _ in range(10)],
    )
    ebt_row = row
    row = write_series(
        ws,
        row,
        "Utilidad antes del I.R.",
        model["ebt"],
        formats,
        formulas=[f"={cell(ebit_row, col)}-{cell(interest_row, col)}+{cell(other_income_row, col)}" for col in range(1, 11)],
    )
    tax_row = row
    row = write_series(
        ws,
        row,
        "Impuesto a la renta 29.5%",
        model["income_tax"],
        formats,
        formulas=[
            f"=MAX(0,{cell(ebt_row, col)}*$B$23)"
            for col in range(1, 11)
        ],
    )
    net_income_row = row
    row = write_series(
        ws,
        row,
        "Utilidad neta",
        model["net_income"],
        formats,
        total=True,
        formulas=[f"={cell(ebt_row, col)}-{cell(tax_row, col)}" for col in range(1, 11)],
    )
    reserve_row = row
    row = write_series(
        ws,
        row,
        "Reserva legal (10%, hasta 20% del capital)",
        model["reserve_legal"],
        formats,
        formulas=[
            f"=MIN(MAX(0,{cell(net_income_row, col)}*$B$24),"
            f"MAX(0,{sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}*$B$25"
            f"-SUM($B${reserve_row + 1}:{cell(reserve_row, col - 1)})))"
            if col > 1
            else f"=MIN(MAX(0,{cell(net_income_row, col)}*$B$24),"
            f"{sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}*$B$25)"
            for col in range(1, 11)
        ],
    )
    row = write_series(
        ws,
        row,
        "Utilidad a distribuir",
        model["retained_profit"],
        formats,
        total=True,
        formulas=[f"={cell(net_income_row, col)}-{cell(reserve_row, col)}" for col in range(1, 11)],
    )
    ws.write(21, 0, "PARAMETROS TRIBUTARIOS Y LEGALES", formats["subtitle"])
    ws.write(22, 0, "Tasa del impuesto a la renta", formats["label"])
    ws.write_number(22, 1, INCOME_TAX, formats["percent"])
    ws.write(23, 0, "Reserva legal sobre utilidad neta", formats["label"])
    ws.write_number(23, 1, 0.10, formats["percent"])
    ws.write(24, 0, "Limite de reserva sobre capital", formats["label"])
    ws.write_number(24, 1, 0.20, formats["percent"])

    ws.write(32, 0, "Firma representante legal", formats["note"])
    ws.write(32, 3, "Firma responsable contable", formats["note"])
    for col in range(1, 11):
        ws.write_formula(sales_row, col, f"={sheet_cell('INGRESOS', 18, col)}", formats["money"], money(model["sales_total"][col - 1]))
        ws.write_formula(
            cost_sales_row,
            col,
            f"={sheet_cell('COSTO DE VENTA', COST_OF_SALES_TOTAL_ROW, col + 1)}",
            formats["money"],
            money(model["cost_of_sales"][col - 1]),
        )
        ws.write_formula(sales_expense_row, col, f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 7, col)}", formats["money"], money(model["sales_labor"][col - 1] + model["sales_other"][col - 1]))
        ws.write_formula(admin_expense_row, col, f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 14, col)}", formats["money"], money(model["admin_labor"][col - 1] + model["admin_other"][col - 1] + model["assets"]["depreciation"][col - 1] * 0.40 + model["assets"]["amortization"][col - 1]))
        ws.write_formula(interest_row, col, f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 19, col)}", formats["money"], money(model["debt"]["interest"][col - 1]))
        ws.write_formula(
            operating_expenses_row,
            col,
            f"={cell(sales_expense_row, col)}+{cell(admin_expense_row, col)}",
            formats["money"],
            money(model["operating_expenses"][col - 1]),
        )

    assert sales_row == INCOME_STATEMENT_SALES_ROW
    assert gross_profit_row == INCOME_STATEMENT_GROSS_PROFIT_ROW
    assert ebit_row == INCOME_STATEMENT_EBIT_ROW
    assert interest_row == INCOME_STATEMENT_INTEREST_ROW
    assert tax_row == INCOME_STATEMENT_TAX_ROW
    assert net_income_row == INCOME_STATEMENT_NET_INCOME_ROW
    assert reserve_row == INCOME_STATEMENT_RESERVE_ROW
    assert row - 1 == INCOME_STATEMENT_DISTRIBUTABLE_ROW


def write_cash_budget(
    workbook,
    model,
    formats,
    *,
    sheet_name: str = "PRESUPUESTO CAJA",
    base_model: dict[str, object] | None = None,
    scenario_value: float | None = None,
):
    ws = workbook.add_worksheet(sheet_name)
    setup(ws, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats, 11)
    caption = (
        "PRESUPUESTO DE CAJA"
        if sheet_name == "PRESUPUESTO CAJA"
        else "PRESUPUESTO DE CAJA - PRECIO REDUCIDO"
    )
    ws.merge_range(1, 0, 1, 11, caption, formats["caption"])
    ws.set_column(0, 0, 42)
    ws.set_column(1, 11, 13)
    row = 2
    headers = ["Concepto", "Año 0"] + YEAR_LABELS
    for col, header in enumerate(headers):
        ws.write(row, col, header, formats["header"])
    row += 1
    ws.write(row, 0, "INGRESOS DE EFECTIVO", formats["subtitle"])
    row += 1

    linked_to_base = sheet_name == "PRESUPUESTO CAJA"
    linked_to_reduced_price = (
        sheet_name == "Presupuesto Caja Bajando Precio"
        and base_model is not None
        and scenario_value is not None
    )

    def write_formula_or_value(target_row, col, formula, value, fmt):
        if linked_to_base:
            ws.write_formula(target_row, col, formula, fmt, money(value))
        else:
            ws.write_formula(target_row, col, f"={money(value)}", fmt, money(value))

    rows = [
        ("Fondos propios", [model["equity"]] + [0] * 10),
        ("Financiamiento", [model["loan"]] + [0] * 10),
        ("Ingresos por venta de contado", [0] + model["cash_sales"]),
        ("Otros ingresos", [0] * 11),
        ("Recuperacion cuentas por cobrar del año", [0] + model["current_credit_collection"]),
        ("Cuentas por cobrar año anterior", [0] + model["previous_cxc_collection"]),
        ("Total ingresos", [INITIAL_INVESTMENT] + model["cash_income"]),
    ]
    income_rows = {}
    for label, values in rows:
        income_rows[label] = row
        ws.write(row, 0, label, formats["total"] if label == "Total ingresos" else formats["label"])
        for col, value in enumerate(values, start=1):
            ws.write_formula(
                row,
                col,
                f"={money(value)}",
                formats["total"] if label == "Total ingresos" else formats["money"],
                money(value),
            )
        row += 1

    write_formula_or_value(
        income_rows["Fondos propios"],
        1,
        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}",
        model["equity"],
        formats["money"],
    )
    write_formula_or_value(
        income_rows["Financiamiento"],
        1,
        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}",
        model["loan"],
        formats["money"],
    )
    for col in range(2, 12):
        year_col = col - 1
        write_formula_or_value(
            income_rows["Ingresos por venta de contado"],
            col,
            f"={sheet_cell('INGRESOS', 22, year_col)}",
            model["cash_sales"][year_col - 1],
            formats["money"],
        )
        write_formula_or_value(
            income_rows["Recuperacion cuentas por cobrar del año"],
            col,
            f"={sheet_cell('INGRESOS', 24, year_col)}",
            model["current_credit_collection"][year_col - 1],
            formats["money"],
        )
        write_formula_or_value(
            income_rows["Cuentas por cobrar año anterior"],
            col,
            f"={sheet_cell('INGRESOS', 25, year_col)}",
            model["previous_cxc_collection"][year_col - 1],
            formats["money"],
        )
    for col in range(1, 12):
        ws.write_formula(
            income_rows["Total ingresos"],
            col,
            f"=SUM({cell(income_rows['Fondos propios'], col)}:{cell(income_rows['Cuentas por cobrar año anterior'], col)})",
            formats["total"],
            money(([INITIAL_INVESTMENT] + model["cash_income"])[col - 1]),
        )
    row += 1
    ws.write(row, 0, "EGRESOS DE EFECTIVO", formats["subtitle"])
    row += 1
    egress_rows = [
        ("Compra de activos fijos", [TANGIBLE_ASSETS] + [0] * 10),
        ("Activos nominales", [INTANGIBLE_ASSETS] + [0] * 10),
        ("Planilla de produccion", [0] + model["production_labor"]),
        ("Costos indirectos de fabricacion", [0] + model["indirect_cash"]),
        ("Compra de contado de insumos y servicios", [0] + model["purchase_cash"]),
        ("Pago a proveedores del año", [0] + model["current_supplier_payment"]),
        ("Pago a proveedores del año anterior", [0] + model["previous_supplier_payment"]),
        ("Gastos de administracion", [0] + [model["admin_labor"][i] + model["admin_other"][i] for i in range(10)]),
        ("Gastos de venta", [0] + [model["sales_labor"][i] + model["sales_other"][i] for i in range(10)]),
        ("Impuesto a la renta pagado", [0] + model["cash_tax_paid"]),
        ("Pago de intereses", [0] + model["debt"]["interest"]),
        ("Abono a capital", [0] + model["debt"]["principal"]),
    ]
    total_egress = [TANGIBLE_ASSETS + INTANGIBLE_ASSETS] + model["cash_egress"]
    egress_start_row = row
    for label, values in egress_rows:
        ws.write(row, 0, label, formats["label"])
        for col, value in enumerate(values, start=1):
            ws.write_formula(row, col, f"={money(value)}", formats["money"], money(value))
        row += 1
    egress_end_row = row - 1

    write_formula_or_value(
        egress_start_row,
        1,
        f"={sheet_cell('PLAN DE INVERSION', 4, 2)}",
        TANGIBLE_ASSETS,
        formats["money"],
    )
    write_formula_or_value(
        egress_start_row + 1,
        1,
        f"={sheet_cell('PLAN DE INVERSION', 19, 2)}",
        INTANGIBLE_ASSETS,
        formats["money"],
    )
    for col in range(2, 12):
        source_col = col - 1
        year_index = col - 2
        formulas_and_values = [
            (
                egress_start_row + 2,
                f"={sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, source_col)}",
                model["production_labor"][year_index],
            ),
            (
                egress_start_row + 3,
                f"=SUM({sheet_cell('COSTOS', 66, source_col)}:{cell(70, source_col)})",
                model["indirect_cash"][year_index],
            ),
            (
                egress_start_row + 4,
                f"={sheet_cell('COSTOS', 57, source_col)}",
                model["purchase_cash"][year_index],
            ),
            (
                egress_start_row + 5,
                f"={sheet_cell('COSTOS', 59, source_col)}",
                model["current_supplier_payment"][year_index],
            ),
            (
                egress_start_row + 6,
                f"={sheet_cell('COSTOS', 60, source_col)}",
                model["previous_supplier_payment"][year_index],
            ),
            (
                egress_start_row + 7,
                f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 11, source_col)}+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 12, source_col)}",
                model["admin_labor"][year_index] + model["admin_other"][year_index],
            ),
            (
                egress_start_row + 8,
                f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 7, source_col)}",
                model["sales_labor"][year_index] + model["sales_other"][year_index],
            ),
            (
                egress_start_row + 9,
                "=0"
                if year_index == 0
                else f"={sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_TAX_ROW, source_col - 1)}",
                model["cash_tax_paid"][year_index],
            ),
            (
                egress_start_row + 10,
                f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 19, source_col)}",
                model["debt"]["interest"][year_index],
            ),
            (
                egress_start_row + 11,
                f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 20, source_col)}",
                model["debt"]["principal"][year_index],
            ),
        ]
        for target_row, formula, value in formulas_and_values:
            write_formula_or_value(target_row, col, formula, value, formats["money"])

    total_egress_row = row
    ws.write(row, 0, "Total egresos", formats["total"])
    for col, value in enumerate(total_egress, start=1):
        ws.write_formula(row, col, f"=SUM({cell(egress_start_row, col)}:{cell(egress_end_row, col)})", formats["total"], money(value))
    row += 2
    flow = [INITIAL_CASH] + model["cash_budget_flow"]
    flow_row = row
    ws.write(row, 0, "Flujo de efectivo", formats["total"])
    for col, value in enumerate(flow, start=1):
        ws.write_formula(row, col, f"={cell(income_rows['Total ingresos'], col)}-{cell(total_egress_row, col)}", formats["total"], money(value))
    row += 2
    initial = [0] + model["cash_balance"][:-1]
    initial_row = row
    ws.write(row, 0, "Saldo inicial", formats["label"])
    for col, value in enumerate(initial, start=1):
        ws.write_number(row, col, value, formats["money"])
    row += 2
    final_row = row
    ws.write(row, 0, "Saldo final", formats["total"])
    for col, value in enumerate(model["cash_balance"], start=1):
        ws.write_formula(row, col, f"={cell(initial_row, col)}+{cell(flow_row, col)}", formats["total"], money(value))
    for col, value in enumerate(initial, start=1):
        ws.write_formula(initial_row, col, "=0" if col == 1 else f"={cell(final_row, col - 1)}", formats["money"], money(value))

    if linked_to_reduced_price:
        assumption_row = 35
        ws.write(assumption_row, 0, "Reduccion de precio", formats["label"])
        ws.write_number(
            assumption_row,
            1,
            scenario_value,
            formats["percent"],
        )
        assumption_ref = f"$B${assumption_row + 1}"
        ws.write_formula(
            income_rows["Fondos propios"],
            1,
            f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}",
            formats["money"],
            model["equity"],
        )
        ws.write_formula(
            income_rows["Financiamiento"],
            1,
            f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}",
            formats["money"],
            model["loan"],
        )
        ws.write_formula(
            egress_start_row,
            1,
            f"={sheet_cell('PLAN DE INVERSION', 4, 2)}",
            formats["money"],
            TANGIBLE_ASSETS,
        )
        ws.write_formula(
            egress_start_row + 1,
            1,
            f"={sheet_cell('PLAN DE INVERSION', 19, 2)}",
            formats["money"],
            INTANGIBLE_ASSETS,
        )
        reduced_income_labels = [
            "Ingresos por venta de contado",
            "Recuperacion cuentas por cobrar del año",
            "Cuentas por cobrar año anterior",
        ]
        for col in range(2, 12):
            for label in reduced_income_labels:
                ws.write_formula(
                    income_rows[label],
                    col,
                    f"={sheet_cell('PRESUPUESTO CAJA', income_rows[label], col)}*(1+{assumption_ref})",
                    formats["money"],
                    money(
                        {
                            "Ingresos por venta de contado": model["cash_sales"],
                            "Recuperacion cuentas por cobrar del año": model["current_credit_collection"],
                            "Cuentas por cobrar año anterior": model["previous_cxc_collection"],
                        }[label][col - 2]
                    ),
                )

        def reduced_direct_cost(year_col: int) -> str:
            return (
                f"{sheet_cell('COSTOS', COSTOS_IMPL_SUBTOTAL_ROW, year_col)}"
                f"+{sheet_cell('COSTOS', COSTOS_RECURRENT_SUBTOTAL_ROW, year_col)}"
                f"+{sheet_cell('COSTOS', COSTOS_EXTRA_SUBTOTAL_ROW, year_col)}"
                f"*(1+{assumption_ref})"
            )

        def reduced_ebt(year_col: int) -> str:
            sales = (
                f"{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_SALES_ROW, year_col)}"
                f"*(1+{assumption_ref})"
            )
            production_cost = (
                f"{reduced_direct_cost(year_col)}"
                f"+{sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, year_col)}"
                f"+{sheet_cell('COSTOS', COSTOS_INDIRECT_TOTAL_ROW, year_col)}"
            )
            operating_expenses = (
                f"{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 14, year_col)}"
                f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 4, year_col)}"
                f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 6, year_col)}"
                f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 5, year_col)}"
                f"*(1+{assumption_ref})"
            )
            return (
                f"({sales}-({production_cost})-({operating_expenses})"
                f"-{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 19, year_col)})"
            )

        cash_share = sheet_cell("COSTOS", 125, 1, abs_row=True, abs_col=True)
        credit_share = sheet_cell("COSTOS", 126, 1, abs_row=True, abs_col=True)
        paid_current = sheet_cell("COSTOS", 127, 1, abs_row=True, abs_col=True)
        for col in range(2, 12):
            year_col = col - 1
            current_direct = reduced_direct_cost(year_col)
            previous_direct = (
                "0"
                if year_col == 1
                else reduced_direct_cost(year_col - 1)
            )
            formulas = [
                f"={sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, year_col)}",
                f"=SUM({sheet_cell('COSTOS', 66, year_col)}:{cell(70, year_col)})",
                f"={current_direct}*{cash_share}",
                f"={current_direct}*{credit_share}*{paid_current}",
                f"={previous_direct}*{credit_share}*(1-{paid_current})",
                f"={sheet_cell('PRESUPUESTO CAJA', egress_start_row + 7, col)}",
                (
                    f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 4, year_col)}"
                    f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 6, year_col)}"
                    f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 5, year_col)}"
                    f"*(1+{assumption_ref})"
                ),
                (
                    "=0"
                    if year_col == 1
                    else (
                        f"=MAX(0,{reduced_ebt(year_col - 1)})"
                        f"*{sheet_cell('ESTADO DE RESULTADOS', 22, 1, abs_row=True, abs_col=True)}"
                    )
                ),
                f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 19, year_col)}",
                f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 20, year_col)}",
            ]
            cached_rows = [
                model["production_labor"],
                model["indirect_cash"],
                model["purchase_cash"],
                model["current_supplier_payment"],
                model["previous_supplier_payment"],
                [
                    model["admin_labor"][i] + model["admin_other"][i]
                    for i in range(10)
                ],
                [
                    model["sales_labor"][i] + model["sales_other"][i]
                    for i in range(10)
                ],
                model["cash_tax_paid"],
                model["debt"]["interest"],
                model["debt"]["principal"],
            ]
            for offset, formula in enumerate(formulas, start=2):
                ws.write_formula(
                    egress_start_row + offset,
                    col,
                    formula,
                    formats["money"],
                    money(cached_rows[offset - 2][year_col - 1]),
                )
    row += 2
    ws.merge_range(
        row,
        0,
        row + 1,
        11,
        "Nota: el presupuesto se alimenta de ingresos, costos, planilla, gastos, impuestos y deuda. Un saldo negativo identifica una necesidad real de capital de trabajo.",
        formats["note"],
    )
    ws.write(row + 5, 0, "Firma representante legal", formats["note"])
    ws.write(row + 5, 8, "Firma responsable contable", formats["note"])

    assert final_row == CASH_BUDGET_FINAL_ROW


def write_balance(workbook, model, formats):
    ws = workbook.add_worksheet("BALANCE GENERAL PROYECTADO")
    setup(ws, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats, 11)
    ws.merge_range(1, 0, 1, 11, "BALANCE GENERAL PROYECTADO", formats["caption"])
    ws.set_column(0, 0, 36)
    ws.set_column(1, 11, 13)
    headers = ["Descripcion", "Año 0"] + YEAR_LABELS
    for col, header in enumerate(headers):
        ws.write(2, col, header, formats["header"])

    ws.merge_range(3, 0, 3, 11, "ACTIVOS", formats["subtitle"])
    ws.write(4, 0, "Circulante:", formats["group_label"])
    cash_row = 5
    cxc_row = 6
    inventory_row = 7
    current_assets_row = BALANCE_CURRENT_ASSETS_ROW
    ws.write(cash_row, 0, "Caja y bancos", formats["label"])
    ws.write(cxc_row, 0, "Cuentas por cobrar", formats["label"])
    ws.write(inventory_row, 0, "Inventario", formats["label"])
    ws.write(current_assets_row, 0, "Total activo circulante", formats["total"])

    ws.write(10, 0, "Fijo:", formats["group_label"])
    fixed_net_row = 11
    fixed_total_row = 13
    ws.write(fixed_net_row, 0, "Activo fijo neto", formats["label"])
    ws.write(fixed_total_row, 0, "Total activo fijo neto", formats["total"])

    ws.write(14, 0, "Nominales:", formats["group_label"])
    intangible_net_row = 15
    intangible_total_row = 16
    total_assets_row = BALANCE_TOTAL_ASSETS_ROW
    ws.write(intangible_net_row, 0, "Activo intangible neto", formats["label"])
    ws.write(intangible_total_row, 0, "Total activo nominal", formats["total"])
    ws.write(total_assets_row, 0, "TOTAL DE ACTIVOS", formats["total"])

    ws.merge_range(18, 0, 18, 11, "PASIVOS", formats["subtitle"])
    ws.write(19, 0, "Circulante:", formats["group_label"])
    cxp_row = BALANCE_CXP_ROW
    short_debt_row = 21
    tax_row = 22
    current_liabilities_row = BALANCE_CURRENT_LIABILITIES_ROW
    ws.write(cxp_row, 0, "Cuentas por pagar", formats["label"])
    ws.write(short_debt_row, 0, "Prestamo de corto plazo", formats["label"])
    ws.write(tax_row, 0, "Impuesto a la renta por pagar", formats["label"])
    ws.write(current_liabilities_row, 0, "Total pasivo circulante", formats["total"])

    ws.write(24, 0, "A largo plazo:", formats["group_label"])
    long_debt_row = BALANCE_LONG_DEBT_ROW
    long_liabilities_total_row = 26
    total_liabilities_row = BALANCE_TOTAL_LIABILITIES_ROW
    ws.write(long_debt_row, 0, "Prestamo de largo plazo", formats["label"])
    ws.write(long_liabilities_total_row, 0, "Total pasivo largo plazo", formats["total"])
    ws.write(total_liabilities_row, 0, "TOTAL PASIVOS", formats["total"])

    ws.merge_range(28, 0, 28, 11, "PATRIMONIO", formats["subtitle"])
    capital_row = 29
    current_profit_row = 30
    reserve_accum_row = 31
    prior_profit_row = 32
    equity_total_row = BALANCE_TOTAL_EQUITY_ROW
    total_liabilities_equity_row = 35
    balance_gap_row = 36
    ws.write(capital_row, 0, "Capital social", formats["label"])
    ws.write(current_profit_row, 0, "Utilidad a distribuir del año", formats["label"])
    ws.write(reserve_accum_row, 0, "Reserva legal acumulada", formats["label"])
    ws.write(prior_profit_row, 0, "Utilidad acumulada de años anteriores", formats["label"])
    ws.write(equity_total_row, 0, "TOTAL PATRIMONIO", formats["total"])
    ws.write(total_liabilities_equity_row, 0, "TOTAL PASIVO Y PATRIMONIO", formats["total"])
    ws.write(balance_gap_row, 0, "Diferencia de cuadre", formats["total"])

    year0_values = {
        cash_row: INITIAL_CASH,
        cxc_row: 0,
        inventory_row: 0,
        current_assets_row: INITIAL_CASH,
        fixed_net_row: TANGIBLE_ASSETS,
        fixed_total_row: TANGIBLE_ASSETS,
        intangible_net_row: INTANGIBLE_ASSETS,
        intangible_total_row: INTANGIBLE_ASSETS,
        total_assets_row: INITIAL_INVESTMENT,
        cxp_row: 0,
        short_debt_row: model["debt"]["principal"][0],
        tax_row: 0,
        current_liabilities_row: model["debt"]["principal"][0],
        long_debt_row: model["loan"] - model["debt"]["principal"][0],
        long_liabilities_total_row: model["loan"] - model["debt"]["principal"][0],
        total_liabilities_row: model["loan"],
        capital_row: model["equity"],
        current_profit_row: 0,
        reserve_accum_row: 0,
        prior_profit_row: 0,
        equity_total_row: model["equity"],
        total_liabilities_equity_row: INITIAL_INVESTMENT,
        balance_gap_row: 0,
    }
    for target_row, value in year0_values.items():
        formula = {
            cash_row: f"={sheet_cell('BALANCE DE APERTURA', 5, 1)}",
            current_assets_row: f"={sheet_cell('BALANCE DE APERTURA', 8, 1)}",
            fixed_net_row: f"={sheet_cell('BALANCE DE APERTURA', 16, 1)}",
            fixed_total_row: f"={cell(fixed_net_row, 1)}",
            intangible_net_row: f"={sheet_cell('BALANCE DE APERTURA', 23, 1)}",
            intangible_total_row: f"={cell(intangible_net_row, 1)}",
            total_assets_row: f"={sheet_cell('BALANCE DE APERTURA', 25, 1)}",
            short_debt_row: f"={sheet_cell('BALANCE DE APERTURA', 5, 4)}",
            long_debt_row: f"={sheet_cell('BALANCE DE APERTURA', 6, 4)}",
            long_liabilities_total_row: f"={cell(long_debt_row, 1)}",
            total_liabilities_row: f"={sheet_cell('BALANCE DE APERTURA', 8, 4)}",
            capital_row: f"={sheet_cell('BALANCE DE APERTURA', 12, 4)}",
            equity_total_row: f"={sheet_cell('BALANCE DE APERTURA', 23, 4)}",
            total_liabilities_equity_row: f"={sheet_cell('BALANCE DE APERTURA', 25, 4)}",
            balance_gap_row: f"={cell(total_assets_row, 1)}-{cell(total_liabilities_equity_row, 1)}",
        }.get(target_row, "=0")
        fmt = formats["total"] if target_row in {
            current_assets_row,
            fixed_total_row,
            intangible_total_row,
            total_assets_row,
            current_liabilities_row,
            long_liabilities_total_row,
            total_liabilities_row,
            equity_total_row,
            total_liabilities_equity_row,
            balance_gap_row,
        } else formats["money"]
        ws.write_formula(target_row, 1, formula, fmt, money(value))

    accumulated_prior = 0.0
    for year_index, col in enumerate(range(2, 12)):
        source_col = year_index + 1
        next_principal = (
            sheet_cell("GASTOS_ADM_VTAS_FINANZAS", 20, source_col + 1)
            if year_index < 9
            else "0"
        )
        prior_profit = accumulated_prior
        accumulated_prior += model["retained_profit"][year_index]
        row_formulas = {
            cash_row: f"={sheet_cell('PRESUPUESTO CAJA', CASH_BUDGET_FINAL_ROW, col)}",
            cxc_row: f"={sheet_cell('INGRESOS', 27, source_col)}",
            inventory_row: "=0",
            current_assets_row: f"=SUM({cell(cash_row, col)}:{cell(inventory_row, col)})",
            fixed_net_row: f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 11, source_col)}",
            fixed_total_row: f"={cell(fixed_net_row, col)}",
            intangible_net_row: f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 23, source_col)}",
            intangible_total_row: f"={cell(intangible_net_row, col)}",
            total_assets_row: f"=SUM({cell(current_assets_row, col)},{cell(fixed_total_row, col)},{cell(intangible_total_row, col)})",
            cxp_row: f"={sheet_cell('COSTOS', COSTOS_CXP_ROW, source_col)}",
            short_debt_row: f"={next_principal}",
            tax_row: f"={sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_TAX_ROW, source_col)}",
            current_liabilities_row: f"=SUM({cell(cxp_row, col)}:{cell(tax_row, col)})",
            long_debt_row: f"=MAX(0,{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 22, source_col)}-{next_principal})",
            long_liabilities_total_row: f"={cell(long_debt_row, col)}",
            total_liabilities_row: f"=SUM({cell(current_liabilities_row, col)},{cell(long_liabilities_total_row, col)})",
            capital_row: f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}",
            current_profit_row: f"={sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_DISTRIBUTABLE_ROW, source_col)}",
            reserve_accum_row: f"=SUM('ESTADO DE RESULTADOS'!$B${INCOME_STATEMENT_RESERVE_ROW + 1}:{cell(INCOME_STATEMENT_RESERVE_ROW, source_col)})",
            prior_profit_row: (
                "=0"
                if year_index == 0
                else f"=SUM('ESTADO DE RESULTADOS'!$B${INCOME_STATEMENT_DISTRIBUTABLE_ROW + 1}:{cell(INCOME_STATEMENT_DISTRIBUTABLE_ROW, source_col - 1)})"
            ),
            equity_total_row: f"=SUM({cell(capital_row, col)}:{cell(prior_profit_row, col)})",
            total_liabilities_equity_row: f"=SUM({cell(total_liabilities_row, col)},{cell(equity_total_row, col)})",
            balance_gap_row: f"={cell(total_assets_row, col)}-{cell(total_liabilities_equity_row, col)}",
        }
        cached_values = {
            cash_row: model["cash_balance"][year_index + 1],
            cxc_row: model["cxc_end"][year_index],
            inventory_row: 0,
            current_assets_row: model["current_assets"][year_index],
            fixed_net_row: model["fixed_net"][year_index],
            fixed_total_row: model["fixed_net"][year_index],
            intangible_net_row: model["intangible_net"][year_index],
            intangible_total_row: model["intangible_net"][year_index],
            total_assets_row: model["total_assets"][year_index],
            cxp_row: model["cxp_end"][year_index],
            short_debt_row: model["debt"]["principal"][year_index + 1] if year_index < 9 else 0,
            tax_row: model["income_tax"][year_index],
            current_liabilities_row: model["current_liabilities"][year_index],
            long_debt_row: model["long_debt"][year_index],
            long_liabilities_total_row: model["long_debt"][year_index],
            total_liabilities_row: model["total_liabilities"][year_index],
            capital_row: model["equity"],
            current_profit_row: model["retained_profit"][year_index],
            reserve_accum_row: model["reserve_accum"][year_index],
            prior_profit_row: prior_profit,
            equity_total_row: model["equity_values"][year_index],
            total_liabilities_equity_row: model["total_liabilities"][year_index] + model["equity_values"][year_index],
            balance_gap_row: model["balance_gap"][year_index],
        }
        for target_row, formula in row_formulas.items():
            fmt = formats["total"] if target_row in {
                current_assets_row,
                fixed_total_row,
                intangible_total_row,
                total_assets_row,
                current_liabilities_row,
                long_liabilities_total_row,
                total_liabilities_row,
                equity_total_row,
                total_liabilities_equity_row,
                balance_gap_row,
            } else formats["money"]
            ws.write_formula(
                target_row,
                col,
                formula,
                fmt,
                money(cached_values[target_row]),
            )

    ws.write(39, 0, "Firma representante legal", formats["note"])
    ws.write(39, 8, "Firma responsable contable", formats["note"])


def write_ratios(workbook, model, formats):
    ws = workbook.add_worksheet("RAZONES FINANCIERAS")
    setup(ws, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats, 11)
    ws.merge_range(1, 0, 1, 11, "RAZONES FINANCIERAS", formats["caption"])
    ws.set_column(0, 0, 34)
    ws.set_column(1, 10, 12)
    ws.set_column(11, 11, 12)

    def safe_formula(expr: str) -> str:
        return f"=IFERROR({expr},0)"

    def balance_ref(row_index: int, year_col: int) -> str:
        return sheet_cell("BALANCE GENERAL PROYECTADO", row_index, year_col + 1)

    def ratio_header(row: int) -> int:
        ws.write(row, 0, "Concepto", formats["header"])
        for col, label in enumerate(YEAR_LABELS, start=1):
            ws.write(row, col, label, formats["header"])
        ws.write(row, 11, "Unidad", formats["header"])
        return row + 1

    row = 3
    categories = [
        (
            "INDICES DE LIQUIDEZ",
            [
                (
                    "Indice de solvencia",
                    [safe_div(model["current_assets"][i], model["current_liabilities"][i]) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"{balance_ref(BALANCE_CURRENT_ASSETS_ROW, col)}/{balance_ref(BALANCE_CURRENT_LIABILITIES_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "veces",
                ),
                (
                    "Capital de trabajo neto",
                    [model["current_assets"][i] - model["current_liabilities"][i] for i in range(10)],
                    "money",
                    [
                        f"={balance_ref(BALANCE_CURRENT_ASSETS_ROW, col)}-{balance_ref(BALANCE_CURRENT_LIABILITIES_ROW, col)}"
                        for col in range(1, 11)
                    ],
                    "S/",
                ),
                (
                    "Prueba del acido",
                    [safe_div(model["current_assets"][i], model["current_liabilities"][i]) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"{balance_ref(BALANCE_CURRENT_ASSETS_ROW, col)}/{balance_ref(BALANCE_CURRENT_LIABILITIES_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "veces",
                ),
            ],
        ),
        (
            "INDICES DE ACTIVIDAD",
            [
                (
                    "Rotacion del inventario",
                    [0] * 10,
                    "number",
                    ["=0" for _ in range(10)],
                    "No aplica",
                ),
                (
                    "Plazo promedio del inventario",
                    [0] * 10,
                    "number",
                    ["=0" for _ in range(10)],
                    "No aplica",
                ),
                (
                    "Rotacion de cuentas por cobrar",
                    [safe_div(model["sales_total"][i], model["cxc_end"][i]) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_SALES_ROW, col)}/{balance_ref(BALANCE_CXC_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "veces",
                ),
                (
                    "Plazo promedio de cobro",
                    [safe_div(360, safe_div(model["sales_total"][i], model["cxc_end"][i])) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"360/({sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_SALES_ROW, col)}/{balance_ref(BALANCE_CXC_ROW, col)})"
                        )
                        for col in range(1, 11)
                    ],
                    "dias",
                ),
                (
                    "Rotacion de cuentas por pagar",
                    [safe_div(model["direct_total"][i], model["cxp_end"][i]) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"{sheet_cell('COSTOS', COSTOS_DIRECT_TOTAL_ROW, col)}/{balance_ref(BALANCE_CXP_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "veces",
                ),
                (
                    "Plazo promedio de pago",
                    [safe_div(360, safe_div(model["direct_total"][i], model["cxp_end"][i])) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"360/({sheet_cell('COSTOS', COSTOS_DIRECT_TOTAL_ROW, col)}/{balance_ref(BALANCE_CXP_ROW, col)})"
                        )
                        for col in range(1, 11)
                    ],
                    "dias",
                ),
            ],
        ),
        (
            "RAZONES DE ENDEUDAMIENTO",
            [
                (
                    "Razon de endeudamiento",
                    [safe_div(model["total_liabilities"][i], model["total_assets"][i]) for i in range(10)],
                    "percent",
                    [
                        safe_formula(
                            f"{balance_ref(BALANCE_TOTAL_LIABILITIES_ROW, col)}/{balance_ref(BALANCE_TOTAL_ASSETS_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "%",
                ),
                (
                    "Pasivo / Capital",
                    [safe_div(model["total_liabilities"][i], model["equity_values"][i]) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"{balance_ref(BALANCE_TOTAL_LIABILITIES_ROW, col)}/{balance_ref(BALANCE_TOTAL_EQUITY_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "veces",
                ),
            ],
        ),
        (
            "RAZONES DE RENTABILIDAD",
            [
                (
                    "Margen bruto",
                    [safe_div(model["gross_profit"][i], model["sales_total"][i]) for i in range(10)],
                    "percent",
                    [
                        safe_formula(
                            f"{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_GROSS_PROFIT_ROW, col)}/{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_SALES_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "%",
                ),
                (
                    "Margen operativo",
                    [safe_div(model["ebit"][i], model["sales_total"][i]) for i in range(10)],
                    "percent",
                    [
                        safe_formula(
                            f"{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_EBIT_ROW, col)}/{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_SALES_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "%",
                ),
                (
                    "Margen neto",
                    [safe_div(model["net_income"][i], model["sales_total"][i]) for i in range(10)],
                    "percent",
                    [
                        safe_formula(
                            f"{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_NET_INCOME_ROW, col)}/{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_SALES_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "%",
                ),
                (
                    "Rotacion total del activo",
                    [safe_div(model["sales_total"][i], model["total_assets"][i]) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_SALES_ROW, col)}/{balance_ref(BALANCE_TOTAL_ASSETS_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "veces",
                ),
                (
                    "Rendimiento sobre activos (ROA)",
                    [safe_div(model["net_income"][i], model["total_assets"][i]) for i in range(10)],
                    "percent",
                    [
                        safe_formula(
                            f"{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_NET_INCOME_ROW, col)}/{balance_ref(BALANCE_TOTAL_ASSETS_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "%",
                ),
            ],
        ),
        (
            "RAZONES DE COBERTURA",
            [
                (
                    "Cobertura de intereses",
                    [safe_div(model["ebit"][i], model["debt"]["interest"][i]) for i in range(10)],
                    "number",
                    [
                        safe_formula(
                            f"{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_EBIT_ROW, col)}/{sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_INTEREST_ROW, col)}"
                        )
                        for col in range(1, 11)
                    ],
                    "veces",
                ),
            ],
        ),
    ]

    for title, ratios in categories:
        row = write_section(ws, row, title, formats, 11)
        row = ratio_header(row)
        for label, values, kind, formulas, unit in ratios:
            current_row = row
            row = write_series(
                ws,
                row,
                label,
                values,
                formats,
                kind=kind,
                formulas=formulas,
            )
            ws.write(current_row, 11, unit, formats["text"])
        row += 1


def write_ratio_comments(workbook, model, formats):
    ws = workbook.add_worksheet("COMENTARIOS DE LAS RAZONES")
    setup(ws, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats, 12)
    ws.merge_range(1, 0, 1, 12, "COMENTARIOS DE LAS RAZONES FINANCIERAS", formats["caption"])
    ws.set_column(0, 0, 31)
    ws.set_column(1, 12, 12)

    def format_value(value: float, kind: str, unit: str) -> str:
        if kind == "percent":
            return f"{value:.2%}"
        if kind == "money":
            return f"S/ {value:,.2f}"
        return f"{value:.2f} {unit}".strip()

    def comment_formula(
        label: str,
        ratio_row: int,
        values: list[float],
        kind: str,
        unit: str,
        interpretation: str,
    ) -> tuple[str, str]:
        excel_format = {
            "percent": "0.00%",
            "money": "#,##0.00",
            "number": "0.00",
        }[kind]
        value_prefix = "S/ " if kind == "money" else ""
        unit_text = "" if kind in {"percent", "money"} else f" {unit}"
        first_ref = sheet_cell("RAZONES FINANCIERAS", ratio_row, 1)
        last_ref = sheet_cell("RAZONES FINANCIERAS", ratio_row, 10)
        formula = (
            f'="{label}: Ano 1 = {value_prefix}"&TEXT({first_ref},"{excel_format}")'
            f'&"{unit_text}; Ano 10 = {value_prefix}"&TEXT({last_ref},"{excel_format}")'
            f'&"{unit_text}. {interpretation}"'
        )
        cached = (
            f"{label}: Ano 1 = {format_value(values[0], kind, unit)}; "
            f"Ano 10 = {format_value(values[-1], kind, unit)}. {interpretation}"
        )
        return formula, cached

    solvency = [safe_div(model["current_assets"][i], model["current_liabilities"][i]) for i in range(10)]
    working_capital = [model["current_assets"][i] - model["current_liabilities"][i] for i in range(10)]
    cxc_rotation = [safe_div(model["sales_total"][i], model["cxc_end"][i]) for i in range(10)]
    collection_days = [safe_div(360, cxc_rotation[i]) for i in range(10)]
    cxp_rotation = [safe_div(model["direct_total"][i], model["cxp_end"][i]) for i in range(10)]
    payment_days = [safe_div(360, cxp_rotation[i]) for i in range(10)]
    debt_ratio = [safe_div(model["total_liabilities"][i], model["total_assets"][i]) for i in range(10)]
    debt_to_equity = [safe_div(model["total_liabilities"][i], model["equity_values"][i]) for i in range(10)]
    gross_margin = [safe_div(model["gross_profit"][i], model["sales_total"][i]) for i in range(10)]
    operating_margin = [safe_div(model["ebit"][i], model["sales_total"][i]) for i in range(10)]
    net_margin = [safe_div(model["net_income"][i], model["sales_total"][i]) for i in range(10)]
    asset_turnover = [safe_div(model["sales_total"][i], model["total_assets"][i]) for i in range(10)]
    roa = [safe_div(model["net_income"][i], model["total_assets"][i]) for i in range(10)]
    interest_coverage = [safe_div(model["ebit"][i], model["debt"]["interest"][i]) for i in range(10)]

    sections = [
        (
            "INDICES DE LIQUIDEZ",
            [
                (
                    "Indice de solvencia",
                    5,
                    solvency,
                    "number",
                    "veces",
                    "Un valor menor a 1 advierte presion para cubrir obligaciones de corto plazo.",
                ),
                (
                    "Capital de trabajo neto",
                    6,
                    working_capital,
                    "money",
                    "",
                    "La mejora refleja la acumulacion de caja y la expansion de los contratos recurrentes.",
                ),
                (
                    "Prueba del acido",
                    7,
                    solvency,
                    "number",
                    "veces",
                    "Coincide con la solvencia porque MuseIQ no mantiene inventarios de productos terminados.",
                ),
            ],
        ),
        (
            "INDICES DE ACTIVIDAD",
            [
                (
                    "Rotacion de cuentas por cobrar",
                    13,
                    cxc_rotation,
                    "number",
                    "veces",
                    "La cobranza institucional debe controlarse mediante hitos, conformidades y anticipos.",
                ),
                (
                    "Plazo promedio de cobro",
                    14,
                    collection_days,
                    "number",
                    "dias",
                    "El plazo resulta de la politica de credito y del saldo pendiente al cierre.",
                ),
                (
                    "Rotacion de cuentas por pagar",
                    15,
                    cxp_rotation,
                    "number",
                    "veces",
                    "Mide la velocidad con la que se cancelan compras y servicios directos financiados.",
                ),
                (
                    "Plazo promedio de pago",
                    16,
                    payment_days,
                    "number",
                    "dias",
                    "Debe coordinarse con el calendario de cobro para evitar brechas de capital de trabajo.",
                ),
            ],
        ),
        (
            "RAZONES DE ENDEUDAMIENTO",
            [
                (
                    "Razon de endeudamiento",
                    20,
                    debt_ratio,
                    "percent",
                    "",
                    "La proporcion disminuye conforme se amortiza el prestamo y crece el patrimonio.",
                ),
                (
                    "Pasivo / capital",
                    21,
                    debt_to_equity,
                    "number",
                    "veces",
                    "Un nivel alto al inicio responde al financiamiento del 60% de la inversion.",
                ),
            ],
        ),
        (
            "RAZONES DE RENTABILIDAD",
            [
                (
                    "Margen bruto",
                    25,
                    gross_margin,
                    "percent",
                    "",
                    "Mide lo que queda de las ventas despues del costo de prestar e implementar el servicio.",
                ),
                (
                    "Margen operativo",
                    26,
                    operating_margin,
                    "percent",
                    "",
                    "Mejora cuando la cartera recurrente absorbe la estructura administrativa y comercial.",
                ),
                (
                    "Margen neto",
                    27,
                    net_margin,
                    "percent",
                    "",
                    "Incluye costos, gastos operativos, intereses e impuesto a la renta.",
                ),
                (
                    "Rotacion total del activo",
                    28,
                    asset_turnover,
                    "number",
                    "veces",
                    "Relaciona las ventas generadas con los activos totales empleados por MuseIQ.",
                ),
                (
                    "Rendimiento sobre activos",
                    29,
                    roa,
                    "percent",
                    "",
                    "Expresa la utilidad neta obtenida por cada sol mantenido en activos.",
                ),
            ],
        ),
        (
            "RAZONES DE COBERTURA",
            [
                (
                    "Cobertura de intereses",
                    33,
                    interest_coverage,
                    "number",
                    "veces",
                    "Los primeros anos concentran el mayor riesgo financiero; la cobertura mejora con la escala.",
                ),
            ],
        ),
    ]

    row = 3
    for section, comments in sections:
        row = write_section(ws, row, section, formats, 12)
        for label, ratio_row, values, kind, unit, interpretation in comments:
            ws.write(row, 0, label, formats["group_label"])
            ws.merge_range(row, 1, row + 2, 12, "", formats["text"])
            formula, cached = comment_formula(
                label,
                ratio_row,
                values,
                kind,
                unit,
                interpretation,
            )
            ws.write_formula(row, 1, formula, formats["text"], cached)
            row += 4


def write_break_even(workbook, model, formats):
    ws = workbook.add_worksheet("PUNTO DE EQUILIBRIO")
    setup(ws, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats, 11)
    ws.merge_range(1, 0, 1, 11, "PUNTO DE EQUILIBRIO", formats["caption"])
    ws.set_column(0, 0, 39)
    ws.set_column(1, 10, 13)
    row = 2
    row = write_year_header(ws, row, formats)
    sales_row = row
    row = write_series(ws, row, "Ventas", model["sales_total"], formats)
    variable_costs_row = row
    row = write_series(ws, row, "Costos variables", model["variable_costs"], formats)
    direct_costs_row = row
    row = write_series(ws, row, "Costos directos", model["direct_total"], formats)
    variable_labor_row = row
    row = write_series(ws, row, "Mano de obra directa", model["variable_labor"], formats)
    variable_indirect_row = row
    row = write_series(ws, row, "Costos indirectos variables", model["variable_indirect"], formats)
    fixed_costs_row = row
    row = write_series(ws, row, "Costos fijos", model["fixed_costs"], formats)
    fixed_production_row = row
    row = write_series(ws, row, "Estructura fija de produccion", model["fixed_production"], formats)
    fixed_admin_row = row
    row = write_series(ws, row, "Administracion", model["fixed_admin"], formats)
    fixed_sales_row = row
    row = write_series(ws, row, "Ventas", model["fixed_sales"], formats)
    fixed_financial_row = row
    row = write_series(ws, row, "Financieros", model["fixed_financial"], formats)
    break_even_sales_row = row
    row = write_series(
        ws,
        row,
        "Punto de equilibrio en soles",
        model["break_even_sales"],
        formats,
        total=True,
        formulas=[f"=IFERROR({cell(fixed_costs_row, col)}/(1-{cell(variable_costs_row, col)}/{cell(sales_row, col)}),0)" for col in range(1, 11)],
    )
    break_even_units_row = row
    row = write_series(
        ws,
        row,
        "Punto de equilibrio en museos equivalentes",
        model["break_even_units"],
        formats,
        kind="number",
        formulas=[
            f"=IFERROR({cell(break_even_sales_row, col)}/({cell(sales_row, col)}/MAX(1,{sheet_cell('COSTOS', COSTOS_IMPLEMENTED_UNITS_ROW, col)}+{sheet_cell('COSTOS', COSTOS_RECURRENT_UNITS_ROW, col)})),0)"
            for col in range(1, 11)
        ],
    )
    row = write_series(
        ws,
        row,
        "Porcentaje de ventas",
        model["break_even_pct"],
        formats,
        kind="percent",
        formulas=[f"=IFERROR({cell(break_even_sales_row, col)}/{cell(sales_row, col)},0)" for col in range(1, 11)],
    )
    for col in range(1, 11):
        source_col = col
        ws.write_formula(
            sales_row,
            col,
            f"={sheet_cell('ESTADO DE RESULTADOS', INCOME_STATEMENT_SALES_ROW, source_col)}",
            formats["money"],
            money(model["sales_total"][col - 1]),
        )
        ws.write_formula(
            direct_costs_row,
            col,
            f"={sheet_cell('COSTOS', COSTOS_DIRECT_TOTAL_ROW, source_col)}",
            formats["money"],
            money(model["direct_total"][col - 1]),
        )
        ws.write_formula(
            variable_labor_row,
            col,
            f"={sheet_cell('PLANILLA', PLANILLA_DIRECT_PRODUCTION_ROW, source_col)}",
            formats["money"],
            money(model["variable_labor"][col - 1]),
        )
        ws.write_formula(
            variable_indirect_row,
            col,
            f"=SUM({sheet_cell('COSTOS', 66, source_col)}:{cell(69, source_col)})",
            formats["money"],
            money(model["variable_indirect"][col - 1]),
        )
        ws.write_formula(
            variable_costs_row,
            col,
            same_row_formula("SUM", [direct_costs_row, variable_labor_row, variable_indirect_row], col),
            formats["money"],
            money(model["variable_costs"][col - 1]),
        )
        ws.write_formula(
            fixed_production_row,
            col,
            f"={sheet_cell('PLANILLA', PLANILLA_INDIRECT_PRODUCTION_ROW, source_col)}"
            f"+{sheet_cell('COSTOS', 70, source_col)}"
            f"+{sheet_cell('DEPRECIACIONES AMORTIZACION', 103, source_col)}",
            formats["money"],
            money(model["fixed_production"][col - 1]),
        )
        ws.write_formula(
            fixed_admin_row,
            col,
            f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 14, source_col)}",
            formats["money"],
            money(model["fixed_admin"][col - 1]),
        )
        ws.write_formula(
            fixed_sales_row,
            col,
            f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 7, source_col)}",
            formats["money"],
            money(model["fixed_sales"][col - 1]),
        )
        ws.write_formula(
            fixed_financial_row,
            col,
            f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 19, source_col)}",
            formats["money"],
            money(model["fixed_financial"][col - 1]),
        )
        ws.write_formula(
            fixed_costs_row,
            col,
            f"=SUM({cell(fixed_production_row, col)}:{cell(fixed_financial_row, col)})",
            formats["money"],
            money(model["fixed_costs"][col - 1]),
        )
    ws.merge_range(
        row + 1,
        0,
        row + 2,
        10,
        "Criterio: los costos variables cambian con los museos atendidos y el consumo del servicio. Los costos fijos incluyen estructura productiva, administracion, ventas, depreciacion, amortizacion e intereses.",
        formats["note"],
    )
    row += 3
    row = write_section(ws, row, "APALANCAMIENTO OPERATIVO", formats)
    row = write_year_header(ws, row, formats)
    contribution = [model["sales_total"][i] - model["variable_costs"][i] for i in range(10)]
    operating_income = [model["sales_total"][i] - model["variable_costs"][i] - model["fixed_costs"][i] for i in range(10)]
    gao = [safe_div(contribution[i], operating_income[i]) for i in range(10)]
    contribution_row = row
    row = write_series(
        ws,
        row,
        "Ingresos menos costo variable",
        contribution,
        formats,
        formulas=[f"={cell(sales_row, col)}-{cell(variable_costs_row, col)}" for col in range(1, 11)],
    )
    operating_income_row = row
    row = write_series(
        ws,
        row,
        "Resultado antes de impuestos",
        operating_income,
        formats,
        formulas=[f"={cell(contribution_row, col)}-{cell(fixed_costs_row, col)}" for col in range(1, 11)],
    )
    row = write_series(
        ws,
        row,
        "GAO",
        gao,
        formats,
        kind="number",
        formulas=[f"=IFERROR({cell(contribution_row, col)}/{cell(operating_income_row, col)},0)" for col in range(1, 11)],
    )


def write_flow_sheet(
    workbook,
    model,
    formats,
    sheet_name: str,
    title: str,
    *,
    scenario_label: str | None = None,
    scenario_value: float | None = None,
    scenario_kind: str | None = None,
    base_model: dict[str, object] | None = None,
    secondary_scenario_label: str | None = None,
    secondary_scenario_value: float | None = None,
):
    ws = workbook.add_worksheet(sheet_name)
    setup(ws, title, formats, 11)
    row = 3
    headers = ["DESCRIPCION", "AÑO 0"] + YEAR_LABELS
    for col, header in enumerate(headers):
        ws.write(row, col, header, formats["header"])
    row += 1
    lines = [
        ("Inversion", [INITIAL_INVESTMENT] + [0] * 10),
        ("Flujo de ingresos", [0] + model["cash_income"]),
        ("Flujo de egresos", [0] + model["cash_egress"]),
        ("Flujo de efectivo", [0] + [model["cash_income"][i] - model["cash_egress"][i] for i in range(10)]),
        ("Mas liquidacion cuentas por cobrar", [0] * 10 + [model["cxc_end"][-1]]),
        ("Mas liquidacion inventarios finales", [0] * 11),
        ("Mas valor de salvamento", [0] * 10 + [SALVAGE_VALUE]),
        ("Menos pasivo circulante", [0] * 10 + [-(model["cxp_end"][-1] + model["income_tax"][-1])]),
        ("Flujo Neto de efectivo", [INITIAL_INVESTMENT] + model["flow_net"]),
    ]
    line_rows = {}
    for label, values in lines:
        line_rows[label] = row
        total = label == "Flujo Neto de efectivo"
        ws.write(row, 0, label, formats["total"] if total else formats["label"])
        for col, value in enumerate(values, start=1):
            ws.write_formula(
                row,
                col,
                f"={money(value)}",
                formats["total"] if total else formats["money"],
                money(value),
            )
        row += 1

    linked_to_base = sheet_name == "FLUJO DE EFECTIVO"
    is_evaluated_scenario = scenario_label is not None and scenario_value is not None
    ws.write_formula(
        line_rows["Inversion"],
        1,
        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 2)}",
        formats["money"],
        INITIAL_INVESTMENT,
    )
    if linked_to_base:
        for col in range(2, 12):
            source_col = col - 1
            ws.write_formula(
                line_rows["Flujo de ingresos"],
                col,
                f"={sheet_cell('PRESUPUESTO CAJA', 10, col)}",
                formats["money"],
                money(model["cash_income"][source_col - 1]),
            )
            ws.write_formula(
                line_rows["Flujo de egresos"],
                col,
                f"={sheet_cell('PRESUPUESTO CAJA', 25, col)}",
                formats["money"],
                money(model["cash_egress"][source_col - 1]),
            )
        ws.write_formula(
            line_rows["Mas liquidacion cuentas por cobrar"],
            11,
            f"={sheet_cell('BALANCE GENERAL PROYECTADO', BALANCE_CXC_ROW, 11)}",
            formats["money"],
            money(model["cxc_end"][-1]),
        )
        ws.write_formula(
            line_rows["Mas liquidacion inventarios finales"],
            11,
            f"={sheet_cell('BALANCE GENERAL PROYECTADO', 7, 11)}",
            formats["money"],
            0,
        )
        ws.write_formula(
            line_rows["Mas valor de salvamento"],
            11,
            f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 107, 1)}",
            formats["money"],
            SALVAGE_VALUE,
        )
        ws.write_formula(
            line_rows["Menos pasivo circulante"],
            11,
            f"=-{sheet_cell('BALANCE GENERAL PROYECTADO', BALANCE_CURRENT_LIABILITIES_ROW, 11)}",
            formats["money"],
            money(-model["current_liabilities"][-1]),
        )
    elif is_evaluated_scenario and base_model is not None:
        assumption_cell = "$C$17"
        secondary_assumption_cell = "$F$17"

        def revenue_factor_expression() -> str:
            if scenario_kind in {"low_income", "optimistic", "reduced_price"}:
                return f"(1+{assumption_cell})"
            return "1"

        def cost_factor_expression() -> str:
            if scenario_kind == "high_cost":
                return f"(1+{assumption_cell})"
            if scenario_kind == "optimistic":
                return f"(1+{secondary_assumption_cell})"
            return "1"

        def direct_cost_expression(year_col: int) -> str:
            revenue_factor = revenue_factor_expression()
            cost_factor = cost_factor_expression()
            return (
                f"(({sheet_cell('COSTOS', COSTOS_IMPL_SUBTOTAL_ROW, year_col)}"
                f"+{sheet_cell('COSTOS', COSTOS_RECURRENT_SUBTOTAL_ROW, year_col)})"
                f"*{cost_factor}"
                f"+{sheet_cell('COSTOS', COSTOS_EXTRA_SUBTOTAL_ROW, year_col)}"
                f"*{revenue_factor}*{cost_factor})"
            )

        def indirect_cash_expression(year_col: int) -> str:
            return (
                f"SUM({sheet_cell('COSTOS', 66, year_col)}:{cell(70, year_col)})"
                f"*{cost_factor_expression()}"
            )

        def supplier_payment_expression(year_col: int) -> str:
            current_direct = direct_cost_expression(year_col)
            previous_direct = (
                "0"
                if year_col == 1
                else direct_cost_expression(year_col - 1)
            )
            cash_share = sheet_cell("COSTOS", 125, 1, abs_row=True, abs_col=True)
            credit_share = sheet_cell("COSTOS", 126, 1, abs_row=True, abs_col=True)
            paid_current = sheet_cell("COSTOS", 127, 1, abs_row=True, abs_col=True)
            return (
                f"{current_direct}*{cash_share}"
                f"+{current_direct}*{credit_share}*{paid_current}"
                f"+{previous_direct}*{credit_share}*(1-{paid_current})"
            )

        def admin_cash_expression(year_col: int) -> str:
            return (
                f"{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 11, year_col)}"
                f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 12, year_col)}"
                f"*{cost_factor_expression()}"
            )

        def sales_cash_expression(year_col: int) -> str:
            return (
                f"{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 4, year_col)}"
                f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 6, year_col)}"
                f"*{cost_factor_expression()}"
                f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 5, year_col)}"
                f"*{revenue_factor_expression()}*{cost_factor_expression()}"
            )

        high_interest_initial_row = 39
        high_interest_interest_row = 40
        high_interest_principal_row = 41
        high_interest_final_row = 42

        def interest_expression(year_col: int) -> str:
            if scenario_kind == "high_interest":
                return cell(high_interest_interest_row, year_col)
            return sheet_cell("GASTOS_ADM_VTAS_FINANZAS", 19, year_col)

        def principal_expression(year_col: int) -> str:
            if scenario_kind == "high_interest":
                return cell(high_interest_principal_row, year_col)
            return sheet_cell("GASTOS_ADM_VTAS_FINANZAS", 20, year_col)

        def ebt_expression(year_col: int) -> str:
            sales = (
                f"{sheet_cell('INGRESOS', 18, year_col)}"
                f"*{revenue_factor_expression()}"
            )
            production_cost = (
                f"{direct_cost_expression(year_col)}"
                f"+{sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, year_col)}"
                f"+{indirect_cash_expression(year_col)}"
                f"+{sheet_cell('COSTOS', 71, year_col)}"
            )
            operating_expenses = (
                f"{admin_cash_expression(year_col)}"
                f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 13, year_col)}"
                f"+{sales_cash_expression(year_col)}"
            )
            return (
                f"({sales}-({production_cost})-({operating_expenses})"
                f"-{interest_expression(year_col)})"
            )

        def cash_tax_expression(year_col: int) -> str:
            if year_col == 1:
                return "0"
            return (
                f"MAX(0,{ebt_expression(year_col - 1)})"
                f"*{sheet_cell('ESTADO DE RESULTADOS', 22, 1, abs_row=True, abs_col=True)}"
            )

        for col in range(2, 12):
            year_index = col - 2
            source_col = col - 1
            base_income = sheet_cell("PRESUPUESTO CAJA", 10, col)
            if scenario_kind == "low_income":
                income_formula = f"={base_income}*(1+{assumption_cell})"
            elif scenario_kind == "optimistic":
                income_formula = f"={base_income}*(1+{assumption_cell})"
            elif scenario_kind == "bad_debt":
                income_formula = (
                    f"={sheet_cell('INGRESOS', 22, source_col)}"
                    f"+({sheet_cell('INGRESOS', 24, source_col)}"
                    f"+{sheet_cell('INGRESOS', 25, source_col)})"
                    f"*(1-{assumption_cell})"
                )
            elif scenario_kind == "reduced_price":
                income_formula = f"={sheet_cell('Presupuesto Caja Bajando Precio', 10, col)}"
            else:
                income_formula = f"={base_income}"
            ws.write_formula(
                line_rows["Flujo de ingresos"],
                col,
                income_formula,
                formats["money"],
                money(model["cash_income"][year_index]),
            )

            if scenario_kind == "reduced_price":
                egress_formula = (
                    f"={sheet_cell('Presupuesto Caja Bajando Precio', 25, col)}"
                )
            elif scenario_kind == "bad_debt":
                egress_formula = (
                    f"={sheet_cell('PRESUPUESTO CAJA', 25, col)}"
                )
            else:
                egress_formula = (
                    f"={sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, source_col)}"
                    f"+{indirect_cash_expression(source_col)}"
                    f"+{supplier_payment_expression(source_col)}"
                    f"+{admin_cash_expression(source_col)}"
                    f"+{sales_cash_expression(source_col)}"
                    f"+{cash_tax_expression(source_col)}"
                    f"+{interest_expression(source_col)}"
                    f"+{principal_expression(source_col)}"
                )
            ws.write_formula(
                line_rows["Flujo de egresos"],
                col,
                egress_formula,
                formats["money"],
                money(model["cash_egress"][year_index]),
            )

        terminal_cxc_formula = (
            f"={sheet_cell('BALANCE GENERAL PROYECTADO', BALANCE_CXC_ROW, 11)}*(1+{assumption_cell})"
            if scenario_kind in {"low_income", "optimistic", "reduced_price"}
            else f"={sheet_cell('BALANCE GENERAL PROYECTADO', BALANCE_CXC_ROW, 11)}"
        )
        ws.write_formula(
            line_rows["Mas liquidacion cuentas por cobrar"],
            11,
            terminal_cxc_formula,
            formats["money"],
            money(model["cxc_end"][-1]),
        )
        ws.write_formula(
            line_rows["Mas liquidacion inventarios finales"],
            11,
            f"={sheet_cell('BALANCE GENERAL PROYECTADO', 7, 11)}",
            formats["money"],
            0,
        )
        ws.write_formula(
            line_rows["Mas valor de salvamento"],
            11,
            f"={sheet_cell('DEPRECIACIONES AMORTIZACION', 107, 1)}",
            formats["money"],
            SALVAGE_VALUE,
        )
        if scenario_kind in {"reduced_price", "low_income", "high_cost", "optimistic"}:
            final_cxp = (
                f"{direct_cost_expression(10)}"
                f"*{sheet_cell('COSTOS', 126, 1, abs_row=True, abs_col=True)}"
                f"*(1-{sheet_cell('COSTOS', 127, 1, abs_row=True, abs_col=True)})"
            )
            final_tax = (
                f"MAX(0,{ebt_expression(10)})"
                f"*{sheet_cell('ESTADO DE RESULTADOS', 22, 1, abs_row=True, abs_col=True)}"
            )
            terminal_liability_formula = f"=-({final_cxp}+{final_tax})"
        else:
            terminal_liability_formula = (
                f"=-{sheet_cell('BALANCE GENERAL PROYECTADO', BALANCE_CURRENT_LIABILITIES_ROW, 11)}"
            )
        ws.write_formula(
            line_rows["Menos pasivo circulante"],
            11,
            terminal_liability_formula,
            formats["money"],
            money(-model["current_liabilities"][-1]),
        )
    for col in range(1, 12):
        ws.write_formula(
            line_rows["Flujo de efectivo"],
            col,
            f"={cell(line_rows['Flujo de ingresos'], col)}-{cell(line_rows['Flujo de egresos'], col)}",
            formats["money"],
            money(([0] + [model["cash_income"][i] - model["cash_egress"][i] for i in range(10)])[col - 1]),
        )
        ws.write_formula(
            line_rows["Flujo Neto de efectivo"],
            col,
            f"={cell(line_rows['Inversion'], col)}"
            f"+{cell(line_rows['Flujo de efectivo'], col)}"
            f"+SUM({cell(line_rows['Mas liquidacion cuentas por cobrar'], col)}:"
            f"{cell(line_rows['Menos pasivo circulante'], col)})",
            formats["total"],
            money(([INITIAL_INVESTMENT] + model["flow_net"])[col - 1]),
        )
    if is_evaluated_scenario:
        evaluation_row = 15
        ws.merge_range(
            evaluation_row,
            0,
            evaluation_row,
            11,
            "EVALUACION DEL ESCENARIO",
            formats["subtitle"],
        )
        ws.write(evaluation_row + 1, 0, "Supuesto evaluado", formats["label"])
        ws.write(evaluation_row + 1, 1, scenario_label, formats["text"])
        if scenario_kind == "reduced_price":
            ws.write_formula(
                evaluation_row + 1,
                2,
                f"={sheet_cell('Presupuesto Caja Bajando Precio', 35, 1)}",
                formats["percent"],
                scenario_value,
            )
        else:
            ws.write_number(
                evaluation_row + 1,
                2,
                scenario_value,
                formats["percent"],
            )
        if (
            secondary_scenario_label is not None
            and secondary_scenario_value is not None
        ):
            ws.write(evaluation_row + 1, 4, secondary_scenario_label, formats["label"])
            ws.write_number(
                evaluation_row + 1,
                5,
                secondary_scenario_value,
                formats["percent"],
            )
        ws.write(evaluation_row + 2, 0, "Costo de capital del escenario", formats["label"])
        scenario_interest_formula = (
            f"{cell(evaluation_row + 1, 2)}"
            if scenario_kind == "high_interest"
            else sheet_cell("GASTOS_ADM_VTAS_FINANZAS", 48, 1)
        )
        ws.write_formula(
            evaluation_row + 2,
            1,
            f"=({sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}"
            f"/{sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 2)})"
            f"*{sheet_cell('COSTO CAPITAL VAN TIR IR', 4, 3)}"
            f"+({sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}"
            f"/{sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 2)})"
            f"*{scenario_interest_formula}",
            formats["percent"],
            model["cost_capital"],
        )
        ws.write(evaluation_row + 3, 0, "Inflacion promedio", formats["label"])
        ws.write_formula(
            evaluation_row + 3,
            1,
            f"={sheet_cell('COSTO CAPITAL VAN TIR IR', 8, 4)}",
            formats["percent"],
            model["avg_inflation"],
        )
        ws.write(evaluation_row + 4, 0, "Riesgo de la inversion", formats["label"])
        ws.write_formula(
            evaluation_row + 4,
            1,
            f"={cell(evaluation_row + 2, 1)}*{cell(evaluation_row + 3, 1)}",
            formats["percent"],
            model["risk"],
        )
        cut_rate_row = evaluation_row + 5
        ws.write(cut_rate_row, 0, "Tasa de corte", formats["total"])
        ws.write_formula(
            cut_rate_row,
            1,
            f"=SUM({cell(evaluation_row + 2, 1)}:{cell(evaluation_row + 4, 1)})",
            formats["total_pct"],
            model["cut_rate"],
        )

        flow_header_row = evaluation_row + 7
        ws.write(flow_header_row, 0, "Concepto", formats["header"])
        for col, header in enumerate(["Año 0"] + YEAR_LABELS, start=1):
            ws.write(flow_header_row, col, header, formats["header"])
        valuation_flow_row = flow_header_row + 1
        discounted_flow_row = flow_header_row + 2
        ws.write(valuation_flow_row, 0, "Flujos para evaluacion", formats["label"])
        ws.write(discounted_flow_row, 0, "Flujos descontados", formats["label"])
        valuation_flows = [-INITIAL_INVESTMENT] + model["flow_net"]
        for col, value in enumerate(valuation_flows, start=1):
            source_formula = (
                f"=-{cell(line_rows['Inversion'], col)}"
                if col == 1
                else f"={cell(line_rows['Flujo Neto de efectivo'], col)}"
            )
            ws.write_formula(
                valuation_flow_row,
                col,
                source_formula,
                formats["money"],
                money(value),
            )
            ws.write_formula(
                discounted_flow_row,
                col,
                f"={cell(valuation_flow_row, col)}/((1+{cell(cut_rate_row, 1)})^{col - 1})",
                formats["money"],
                money(value / ((1 + model["cut_rate"]) ** (col - 1))),
            )

        van_row = flow_header_row + 4
        tir_row = flow_header_row + 5
        profitability_row = flow_header_row + 6
        decision_row = flow_header_row + 7
        ws.write(van_row, 0, "Valor actual neto", formats["total"])
        ws.write_formula(
            van_row,
            1,
            f"=SUM({cell(discounted_flow_row, 1)}:{cell(discounted_flow_row, 11)})",
            formats["total"],
            model["npv"],
        )
        ws.write(tir_row, 0, "Tasa interna de retorno", formats["total"])
        ws.write_formula(
            tir_row,
            1,
            f"=IRR({cell(valuation_flow_row, 1)}:{cell(valuation_flow_row, 11)})",
            formats["total_pct"],
            model["irr"],
        )
        ws.write(profitability_row, 0, "Indice de rentabilidad", formats["total"])
        ws.write_formula(
            profitability_row,
            1,
            f"=({cell(van_row, 1)}+ABS({cell(valuation_flow_row, 1)}))/ABS({cell(valuation_flow_row, 1)})",
            formats["total_num"],
            model["profitability_index"],
        )
        decision = (
            "PROYECTO SE ACEPTA"
            if model["npv"] > 0
            and model["irr"] > model["cut_rate"]
            and model["profitability_index"] > 1
            else "PROYECTO NO VIABLE EN ESTE ESCENARIO"
        )
        ws.write(decision_row, 0, "Resultado", formats["subtitle"])
        ws.write_formula(
            decision_row,
            1,
            f'=IF(AND({cell(van_row, 1)}>0,{cell(tir_row, 1)}>{cell(cut_rate_row, 1)},'
            f'{cell(profitability_row, 1)}>1),"PROYECTO SE ACEPTA","PROYECTO NO VIABLE EN ESTE ESCENARIO")',
            formats["text"],
            decision,
        )
        ws.merge_range(
            decision_row + 2,
            0,
            decision_row + 3,
            11,
            "Criterio: el escenario se acepta cuando el VAN es positivo, la TIR supera la tasa de corte y el indice de rentabilidad es mayor que uno.",
            formats["note"],
        )
        ws.write(decision_row + 6, 0, "Firma representante legal", formats["note"])
        ws.write(decision_row + 6, 8, "Firma responsable contable", formats["note"])
        if scenario_kind == "high_interest":
            ws.merge_range(
                37,
                0,
                37,
                10,
                "CEDULA DE DEUDA DEL ESCENARIO",
                formats["subtitle"],
            )
            ws.write(38, 0, "Concepto", formats["header"])
            for col, label in enumerate(YEAR_LABELS, start=1):
                ws.write(38, col, label, formats["header"])
            ws.write(high_interest_initial_row, 0, "Saldo inicial", formats["label"])
            ws.write(high_interest_interest_row, 0, "Intereses", formats["label"])
            ws.write(high_interest_principal_row, 0, "Abono a capital", formats["label"])
            ws.write(high_interest_final_row, 0, "Saldo final", formats["total"])
            for col in range(1, 11):
                ws.write_formula(
                    high_interest_initial_row,
                    col,
                    (
                        f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}"
                        if col == 1
                        else f"={cell(high_interest_final_row, col - 1)}"
                    ),
                    formats["money"],
                    money(model["debt"]["initial_balance"][col - 1]),
                )
                ws.write_formula(
                    high_interest_interest_row,
                    col,
                    f"={cell(high_interest_initial_row, col)}*$C$17",
                    formats["money"],
                    money(model["debt"]["interest"][col - 1]),
                )
                ws.write_formula(
                    high_interest_principal_row,
                    col,
                    f"=IF({col}<={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 49, 1)},0,"
                    f"IF({col}<={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 49, 1)}"
                    f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 50, 1)},"
                    f"MIN(PMT($C$17,{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 50, 1)},"
                    f"-{sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)})"
                    f"-{cell(high_interest_interest_row, col)},"
                    f"{cell(high_interest_initial_row, col)}),0))",
                    formats["money"],
                    money(model["debt"]["principal"][col - 1]),
                )
                ws.write_formula(
                    high_interest_final_row,
                    col,
                    f"=MAX(0,{cell(high_interest_initial_row, col)}"
                    f"-{cell(high_interest_principal_row, col)})",
                    formats["total"],
                    money(model["debt"]["final_balance"][col - 1]),
                )
    else:
        ws.merge_range(
            15,
            0,
            16,
            11,
            "El flujo neto incorpora en el Año 10 la recuperacion de cuentas por cobrar, el valor de salvamento y la cancelacion del pasivo circulante, siguiendo la estructura del documento de referencia.",
            formats["note"],
        )
        ws.write(19, 0, "Firma representante legal", formats["note"])
        ws.write(19, 8, "Firma responsable contable", formats["note"])


def write_cost_capital(workbook, model, formats):
    ws = workbook.add_worksheet("COSTO CAPITAL VAN TIR IR")
    setup(ws, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats, 15)
    ws.merge_range(1, 0, 1, 15, "COSTO DE CAPITAL, VAN, TIR E INDICE DE RENTABILIDAD", formats["caption"])
    ws.set_column(0, 0, 30)
    ws.set_column(1, 4, 16)
    ws.set_column(6, 15, 12)
    row = 3
    ws.write(row, 0, "Concepto", formats["header"])
    ws.write(row, 1, "Monto de la inversion", formats["header"])
    ws.write(row, 2, "%", formats["header"])
    ws.write(row, 3, "Costo", formats["header"])
    ws.write(row, 4, "Costo promedio", formats["header"])
    for col, label in enumerate(YEAR_LABELS, start=6):
        ws.write(row, col, label, formats["header"])
    row += 1

    base_row = row
    capital_rows = [
        ("Fondos propios", model["equity"], EQUITY_COST),
        ("Financiamiento", model["loan"], model["interest_rate"]),
    ]
    for offset, (label, amount, cost) in enumerate(capital_rows):
        ws.write(row, 0, label, formats["label"])
        amount_formula = (
            f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 3)}"
            if offset == 0
            else f"={sheet_cell('PLAN DE INVERSION', PLAN_INV_FINANCING_TOTAL_ROW, 5)}"
        )
        ws.write_formula(row, 1, amount_formula, formats["money"], amount)
        excel_row = row + 1
        ws.write_formula(row, 2, f"=B{excel_row}/SUM($B${base_row + 1}:$B${base_row + 2})", formats["percent"], [EQUITY_SHARE, LOAN_SHARE][offset])
        cost_formula = (
            None
            if offset == 0
            else f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 48, 1)}"
        )
        if cost_formula is None:
            ws.write_number(row, 3, cost, formats["percent"])
        else:
            ws.write_formula(row, 3, cost_formula, formats["percent"], cost)
        ws.write_formula(row, 4, f"=C{excel_row}*D{excel_row}", formats["percent"], [EQUITY_SHARE * EQUITY_COST, LOAN_SHARE * model["interest_rate"]][offset])
        row += 1

    ws.write(row, 0, "Costo de capital", formats["label"])
    ws.write_formula(row, 4, f"=SUM(E{base_row + 1}:E{base_row + 2})", formats["percent"], model["cost_capital"])
    cost_capital_row = row + 1
    row += 1

    ws.write(row, 0, "Inflacion anual", formats["label"])
    inflation_values_row = row + 1
    for year_index, (col, value) in enumerate(zip(range(6, 16), INFLATION), start=1):
        ws.write_formula(
            row,
            col,
            f"={sheet_cell('INGRESOS', 31, year_index)}",
            formats["percent"],
            value,
        )
    row += 1

    ws.write(row, 0, "Inflacion promedio", formats["label"])
    ws.write_formula(row, 4, f"=AVERAGE(G{inflation_values_row}:P{inflation_values_row})", formats["percent"], model["avg_inflation"])
    inflation_row = row + 1
    row += 1

    ws.write(row, 0, "Riesgo de la inversion", formats["label"])
    ws.write_formula(row, 4, f"=E{cost_capital_row}*E{inflation_row}", formats["percent"], model["risk"])
    risk_row = row + 1
    row += 1

    ws.write(row, 0, "Tasa de corte", formats["total"])
    ws.write_formula(row, 4, f"=E{cost_capital_row}+E{inflation_row}+E{risk_row}", formats["total_pct"], model["cut_rate"])
    cut_rate_row = row + 1
    row += 1

    row += 2
    ws.write(row, 0, "Valor Actual Neto", formats["subtitle"])
    row += 1
    ws.write(row, 0, "Flujos netos", formats["header"])
    ws.write(row, 1, "Flujos descontados", formats["header"])
    ws.write(row, 2, "Flujo descontado acumulado", formats["header"])
    row += 1
    first_flow_row = row + 1
    cumulative_discounted = 0.0
    cumulative_values = []
    for i, flow in enumerate(model["flows_for_valuation"]):
        label = "INVERSION INICIAL" if i == 0 else f"FLUJO AÑO {i}"
        ws.write(row, 0, label, formats["label"])
        source_col = i + 1
        source_formula = f"=-{sheet_cell('FLUJO DE EFECTIVO', 12, source_col)}" if i == 0 else f"={sheet_cell('FLUJO DE EFECTIVO', 12, source_col)}"
        ws.write_formula(row, 1, source_formula, formats["money"], money(flow))
        excel_row = row + 1
        ws.write_formula(row, 2, f"=B{excel_row}/((1+$E${cut_rate_row})^{i})", formats["money"], flow / ((1 + model["cut_rate"]) ** i))
        discounted_value = flow / ((1 + model["cut_rate"]) ** i)
        cumulative_discounted += discounted_value
        cumulative_values.append(cumulative_discounted)
        cumulative_formula = f"=C{excel_row}" if i == 0 else f"=D{excel_row - 1}+C{excel_row}"
        ws.write_formula(
            row,
            3,
            cumulative_formula,
            formats["money"],
            money(cumulative_discounted),
        )
        row += 1
    last_flow_row = row
    row += 1
    ws.write(row, 0, "VAN", formats["total"])
    ws.write_formula(row, 1, f"=SUM(C{first_flow_row}:C{last_flow_row})", formats["total"], model["npv"])
    van_row = row + 1
    row += 1
    ws.write(row, 0, "TIR", formats["total"])
    ws.write_formula(row, 1, f"=IRR(B{first_flow_row}:B{last_flow_row})", formats["total_pct"], model["irr"])
    row += 1
    ws.write(row, 0, "Indice de rentabilidad", formats["total"])
    ws.write_formula(row, 1, f"=(B{van_row}+ABS(B{first_flow_row}))/ABS(B{first_flow_row})", formats["total_num"], model["profitability_index"])
    row += 1
    ws.write(row, 0, "Periodo recuperacion descontado", formats["total"])
    if math.isfinite(model["payback"]):
        recovery_year = next(
            index
            for index, cumulative in enumerate(cumulative_values)
            if index > 0 and cumulative >= 0
        )
        previous_excel_row = first_flow_row + recovery_year - 1
        recovery_excel_row = first_flow_row + recovery_year
        ws.write_formula(
            row,
            1,
            f"={recovery_year - 1}+ABS(D{previous_excel_row})/C{recovery_excel_row}",
            formats["total_num"],
            model["payback"],
        )
    else:
        ws.write_string(row, 1, "No recupera en 10 años", formats["total_num"])
    row += 1
    decision = "PROYECTO SE ACEPTA" if model["npv"] > 0 and model["irr"] > model["cut_rate"] and model["profitability_index"] > 1 else "PROYECTO NO VIABLE EN ESCENARIO BASE"
    ws.write(row, 0, "Resultado", formats["subtitle"])
    ws.write_formula(
        row,
        1,
        f'=IF(AND(B{van_row}>0,B{van_row + 1}>$E${cut_rate_row},B{van_row + 2}>1),'
        f'"PROYECTO SE ACEPTA","PROYECTO NO VIABLE EN ESCENARIO BASE")',
        formats["text"],
        decision,
    )
    ws.merge_range(
        row + 2,
        0,
        row + 3,
        15,
        "La tasa de corte integra el costo ponderado de fondos propios y financiamiento, la inflacion promedio y una prima de riesgo. La decision exige VAN positivo, TIR superior a la tasa de corte e indice de rentabilidad mayor que uno.",
        formats["note"],
    )


def write_risk_admin(workbook, model, admin_model, formats):
    write_planilla(workbook, admin_model, formats, admin_factor=1.15, sheet_name="Riesgo Administrativo")
    ws = workbook.get_worksheet_by_name("Riesgo Administrativo")
    ws.write(0, 0, "Riesgo Administrativo - Planilla y Evaluacion Financiera", formats["title"])
    row = 273
    row = write_section(ws, row, "EVALUACION DEL RIESGO ADMINISTRATIVO", formats)
    ws.write(row, 0, "Aumento adicional de gastos administrativos", formats["label"])
    ws.write_number(row, 1, 0.15, formats["percent"])
    row += 2
    row = write_year_header(ws, row, formats)
    base_admin_row = row
    row = write_series(
        ws,
        row,
        "Gastos administrativos desembolsables - base",
        [model["admin_labor"][i] + model["admin_other"][i] for i in range(10)],
        formats,
        formulas=[
            f"={sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 11, col)}"
            f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 12, col)}"
            for col in range(1, 11)
        ],
    )
    scenario_admin_row = row
    scenario_admin_cash = [
        admin_model["admin_labor"][i] + admin_model["admin_other"][i]
        for i in range(10)
    ]
    row = write_series(
        ws,
        row,
        "Gastos administrativos desembolsables - escenario",
        scenario_admin_cash,
        formats,
        formulas=[
            f"={cell(base_admin_row, col)}*(1+$B$275)"
            for col in range(1, 11)
        ],
    )
    admin_delta_row = row
    row = write_series(
        ws,
        row,
        "Ajuste administrativo",
        [
            scenario_admin_cash[i]
            - model["admin_labor"][i]
            - model["admin_other"][i]
            for i in range(10)
        ],
        formats,
        formulas=[
            f"={cell(scenario_admin_row, col)}-{cell(base_admin_row, col)}"
            for col in range(1, 11)
        ],
    )
    tax_delta_row = row
    row = write_series(
        ws,
        row,
        "Variacion del impuesto pagado",
        [
            admin_model["cash_tax_paid"][i] - model["cash_tax_paid"][i]
            for i in range(10)
        ],
        formats,
        formulas=[
            "=0"
            if col == 1
            else (
                f"=MAX(0,{sheet_cell('ESTADO DE RESULTADOS', 15, col - 1)}"
                f"-{cell(admin_delta_row, col - 1)})"
                f"*{sheet_cell('ESTADO DE RESULTADOS', 22, 1, abs_row=True, abs_col=True)}"
                f"-{sheet_cell('PRESUPUESTO CAJA', 22, col + 1)}"
            )
            for col in range(1, 11)
        ],
    )
    terminal_delta_values = [0.0] * 9 + [
        admin_model["current_liabilities"][-1] - model["current_liabilities"][-1]
    ]
    terminal_delta_row = row
    row = write_series(
        ws,
        row,
        "Variacion del pasivo circulante final",
        terminal_delta_values,
        formats,
        formulas=[
            "=0"
            if col < 10
            else (
                f"=MAX(0,{sheet_cell('ESTADO DE RESULTADOS', 15, 10)}"
                f"-{cell(admin_delta_row, 10)})"
                f"*{sheet_cell('ESTADO DE RESULTADOS', 22, 1, abs_row=True, abs_col=True)}"
                f"-{sheet_cell('ESTADO DE RESULTADOS', 16, 10)}"
            )
            for col in range(1, 11)
        ],
    )
    base_flow_row = row
    row = write_series(
        ws,
        row,
        "Flujo neto base",
        model["flow_net"],
        formats,
        formulas=[
            f"={sheet_cell('FLUJO DE EFECTIVO', 12, col + 1)}"
            for col in range(1, 11)
        ],
    )
    scenario_flow_row = row
    row = write_series(
        ws,
        row,
        "Flujo neto con riesgo administrativo",
        admin_model["flow_net"],
        formats,
        total=True,
        formulas=[
            f"={cell(base_flow_row, col)}-{cell(admin_delta_row, col)}"
            f"-{cell(tax_delta_row, col)}-{cell(terminal_delta_row, col)}"
            for col in range(1, 11)
        ],
    )
    difference_row = row
    row = write_series(
        ws,
        row,
        "Diferencia frente al escenario base",
        [
            admin_model["flow_net"][i] - model["flow_net"][i]
            for i in range(10)
        ],
        formats,
        formulas=[
            f"={cell(scenario_flow_row, col)}-{cell(base_flow_row, col)}"
            for col in range(1, 11)
        ],
    )

    row += 2
    row = write_section(ws, row, "INDICADORES DEL ESCENARIO", formats)
    ws.write(row, 0, "Concepto", formats["header"])
    for col, header in enumerate(["Año 0"] + YEAR_LABELS, start=1):
        ws.write(row, col, header, formats["header"])
    row += 1
    valuation_row = row
    discounted_row = row + 1
    ws.write(valuation_row, 0, "Flujos para evaluacion", formats["label"])
    ws.write(discounted_row, 0, "Flujos descontados", formats["label"])
    valuation_flows = [-INITIAL_INVESTMENT] + admin_model["flow_net"]
    for col, value in enumerate(valuation_flows, start=1):
        source_formula = (
            f"=-{sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 2)}"
            if col == 1
            else f"={cell(scenario_flow_row, col - 1)}"
        )
        ws.write_formula(
            valuation_row,
            col,
            source_formula,
            formats["money"],
            money(value),
        )
        ws.write_formula(
            discounted_row,
            col,
            f"={cell(valuation_row, col)}/((1+{sheet_cell('COSTO CAPITAL VAN TIR IR', 10, 4)})^{col - 1})",
            formats["money"],
            money(value / ((1 + admin_model["cut_rate"]) ** (col - 1))),
        )
    row = discounted_row + 2
    van_row = row
    tir_row = row + 1
    profitability_row = row + 2
    decision_row = row + 3
    ws.write(van_row, 0, "Valor actual neto", formats["total"])
    ws.write_formula(
        van_row,
        1,
        f"=SUM({cell(discounted_row, 1)}:{cell(discounted_row, 11)})",
        formats["total"],
        admin_model["npv"],
    )
    ws.write(tir_row, 0, "Tasa interna de retorno", formats["total"])
    ws.write_formula(
        tir_row,
        1,
        f"=IRR({cell(valuation_row, 1)}:{cell(valuation_row, 11)})",
        formats["total_pct"],
        admin_model["irr"],
    )
    ws.write(profitability_row, 0, "Indice de rentabilidad", formats["total"])
    ws.write_formula(
        profitability_row,
        1,
        f"=({cell(van_row, 1)}+ABS({cell(valuation_row, 1)}))/ABS({cell(valuation_row, 1)})",
        formats["total_num"],
        admin_model["profitability_index"],
    )
    decision = (
        "PROYECTO SE ACEPTA"
        if admin_model["npv"] > 0
        and admin_model["irr"] > admin_model["cut_rate"]
        and admin_model["profitability_index"] > 1
        else "PROYECTO NO VIABLE EN ESTE ESCENARIO"
    )
    ws.write(decision_row, 0, "Resultado", formats["subtitle"])
    ws.write_formula(
        decision_row,
        1,
        f'=IF(AND({cell(van_row, 1)}>0,{cell(tir_row, 1)}>{sheet_cell("COSTO CAPITAL VAN TIR IR", 10, 4)},'
        f'{cell(profitability_row, 1)}>1),"PROYECTO SE ACEPTA","PROYECTO NO VIABLE EN ESTE ESCENARIO")',
        formats["text"],
        decision,
    )
    ws.merge_range(
        decision_row + 2,
        0,
        decision_row + 3,
        10,
        "El escenario considera un incremento adicional de 15% en los gastos administrativos y muestra por separado su impacto directo, tributario y sobre el pasivo circulante final.",
        formats["note"],
    )


def write_price_analysis(workbook, base_model, formats):
    ws = workbook.add_worksheet("Analisis Precio")
    setup(ws, "MuseIQ - Plataforma Inteligente para Experiencias Museales", formats, 11)
    ws.merge_range(1, 0, 1, 11, "ANALISIS UNIDIMENSIONAL DEL PRECIO", formats["caption"])
    ws.set_column(0, 0, 29)
    ws.set_column(1, 11, 13)
    scenarios = [
        ("Precio -15%", 0.85),
        ("Precio -10%", 0.90),
        ("Precio base", 1.00),
        ("Precio +10%", 1.10),
        ("Precio +15%", 1.15),
    ]
    scenario_models = [
        build_model(name=label, revenue_factor=factor)
        for label, factor in scenarios
    ]

    row = 3
    ws.write(row, 0, "Escenario", formats["header"])
    ws.write(row, 1, "Factor precio", formats["header"])
    ws.write(row, 2, "Ventas acumuladas", formats["header"])
    ws.write(row, 3, "VAN", formats["header"])
    ws.write(row, 4, "TIR", formats["header"])
    ws.write(row, 5, "IR", formats["header"])
    row += 1
    summary_rows = []
    for label, factor in scenarios:
        summary_rows.append(row)
        ws.write(row, 0, label, formats["label"])
        if label == "Precio -10%":
            ws.write_formula(
                row,
                1,
                f"=1+{sheet_cell('Presupuesto Caja Bajando Precio', 35, 1)}",
                formats["percent"],
                factor,
            )
        else:
            ws.write_number(row, 1, factor, formats["percent"])
        ws.write_formula(
            row,
            2,
            f"=SUM({sheet_cell('INGRESOS', 18, 1)}:{cell(18, 10)})*{cell(row, 1)}",
            formats["money"],
            money(sum(base_model["sales_total"]) * factor),
        )
        row += 1

    row += 2
    row = write_year_header(ws, row, formats)
    base_sales_row = row
    row = write_series(
        ws,
        row,
        "Ventas escenario base",
        base_model["sales_total"],
        formats,
        formulas=[
            f"={sheet_cell('INGRESOS', 18, col)}"
            for col in range(1, 11)
        ],
    )
    reduced = scenario_models[1]
    reduced_sales_row = row
    row = write_series(
        ws,
        row,
        "Ventas precio -10%",
        reduced["sales_total"],
        formats,
        formulas=[
            f"={cell(base_sales_row, col)}*{cell(summary_rows[1], 1, abs_row=True, abs_col=True)}"
            for col in range(1, 11)
        ],
    )
    row = write_series(
        ws,
        row,
        "Diferencia",
        [
            reduced["sales_total"][i] - base_model["sales_total"][i]
            for i in range(10)
        ],
        formats,
        formulas=[
            f"={cell(reduced_sales_row, col)}-{cell(base_sales_row, col)}"
            for col in range(1, 11)
        ],
    )

    def price_direct_cost_expression(year_col: int, factor_ref: str) -> str:
        return (
            f"({sheet_cell('COSTOS', COSTOS_IMPL_SUBTOTAL_ROW, year_col)}"
            f"+{sheet_cell('COSTOS', COSTOS_RECURRENT_SUBTOTAL_ROW, year_col)}"
            f"+{sheet_cell('COSTOS', COSTOS_EXTRA_SUBTOTAL_ROW, year_col)}"
            f"*{factor_ref})"
        )

    def price_supplier_payment_expression(year_col: int, factor_ref: str) -> str:
        current_direct = price_direct_cost_expression(year_col, factor_ref)
        previous_direct = (
            "0"
            if year_col == 1
            else price_direct_cost_expression(year_col - 1, factor_ref)
        )
        cash_share = sheet_cell("COSTOS", 125, 1, abs_row=True, abs_col=True)
        credit_share = sheet_cell("COSTOS", 126, 1, abs_row=True, abs_col=True)
        paid_current = sheet_cell("COSTOS", 127, 1, abs_row=True, abs_col=True)
        return (
            f"{current_direct}*{cash_share}"
            f"+{current_direct}*{credit_share}*{paid_current}"
            f"+{previous_direct}*{credit_share}*(1-{paid_current})"
        )

    def price_sales_cash_expression(year_col: int, factor_ref: str) -> str:
        return (
            f"{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 4, year_col)}"
            f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 6, year_col)}"
            f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 5, year_col)}"
            f"*{factor_ref}"
        )

    def price_ebt_expression(year_col: int, factor_ref: str) -> str:
        return (
            f"({sheet_cell('INGRESOS', 18, year_col)}*{factor_ref}"
            f"-{price_direct_cost_expression(year_col, factor_ref)}"
            f"-{sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, year_col)}"
            f"-{sheet_cell('COSTOS', COSTOS_INDIRECT_TOTAL_ROW, year_col)}"
            f"-{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 14, year_col)}"
            f"-{price_sales_cash_expression(year_col, factor_ref)}"
            f"-{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 19, year_col)})"
        )

    def price_cash_tax_expression(year_col: int, factor_ref: str) -> str:
        if year_col == 1:
            return "0"
        return (
            f"MAX(0,{price_ebt_expression(year_col - 1, factor_ref)})"
            f"*{sheet_cell('ESTADO DE RESULTADOS', 22, 1, abs_row=True, abs_col=True)}"
        )

    def price_cash_egress_expression(year_col: int, factor_ref: str) -> str:
        indirect_cash = (
            f"SUM({sheet_cell('COSTOS', 66, year_col)}:{cell(70, year_col)})"
        )
        admin_cash = (
            f"{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 11, year_col)}"
            f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 12, year_col)}"
        )
        return (
            f"{price_supplier_payment_expression(year_col, factor_ref)}"
            f"+{sheet_cell('PLANILLA', PLANILLA_PRODUCTION_SUMMARY_ROW, year_col)}"
            f"+{indirect_cash}"
            f"+{admin_cash}"
            f"+{price_sales_cash_expression(year_col, factor_ref)}"
            f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 19, year_col)}"
            f"+{sheet_cell('GASTOS_ADM_VTAS_FINANZAS', 20, year_col)}"
            f"+{price_cash_tax_expression(year_col, factor_ref)}"
        )

    row += 2
    calculation_rows = []
    for scenario_index, ((label, _), scenario_model) in enumerate(
        zip(scenarios, scenario_models)
    ):
        row = write_section(ws, row, f"FLUJOS DE EVALUACION - {label.upper()}", formats, 11)
        ws.write(row, 0, "Concepto", formats["header"])
        for col, header in enumerate(["Año 0"] + YEAR_LABELS, start=1):
            ws.write(row, col, header, formats["header"])
        row += 1
        valuation_row = row
        discounted_row = row + 1
        ws.write(valuation_row, 0, "Flujos para evaluacion", formats["label"])
        ws.write(discounted_row, 0, "Flujos descontados", formats["label"])
        factor_ref = cell(summary_rows[scenario_index], 1, abs_row=True, abs_col=True)
        valuation_values = [-INITIAL_INVESTMENT] + scenario_model["flow_net"]
        for col, value in enumerate(valuation_values, start=1):
            if col == 1:
                formula = f"=-{sheet_cell('PLAN DE INVERSION', PLAN_INV_EQUITY_TOTAL_ROW, 2)}"
            else:
                year_index = col - 2
                source_col = col - 1
                formula = (
                    f"={sheet_cell('PRESUPUESTO CAJA', 10, col)}*{factor_ref}"
                    f"-({price_cash_egress_expression(source_col, factor_ref)})"
                )
                if year_index == 9:
                    final_cxp_formula = (
                        f"{price_direct_cost_expression(source_col, factor_ref)}"
                        f"*{sheet_cell('COSTOS', 126, 1, abs_row=True, abs_col=True)}"
                        f"*(1-{sheet_cell('COSTOS', 127, 1, abs_row=True, abs_col=True)})"
                    )
                    final_tax_formula = (
                        f"MAX(0,{price_ebt_expression(source_col, factor_ref)})"
                        f"*{sheet_cell('ESTADO DE RESULTADOS', 22, 1, abs_row=True, abs_col=True)}"
                    )
                    formula += (
                        f"+{sheet_cell('BALANCE GENERAL PROYECTADO', BALANCE_CXC_ROW, 11)}"
                        f"*{factor_ref}"
                        f"+{sheet_cell('DEPRECIACIONES AMORTIZACION', 107, 1)}"
                        f"-({final_cxp_formula}+{final_tax_formula})"
                    )
            ws.write_formula(
                valuation_row,
                col,
                formula,
                formats["money"],
                money(value),
            )
            ws.write_formula(
                discounted_row,
                col,
                f"={cell(valuation_row, col)}/((1+{sheet_cell('COSTO CAPITAL VAN TIR IR', 10, 4)})^{col - 1})",
                formats["money"],
                money(value / ((1 + scenario_model["cut_rate"]) ** (col - 1))),
            )
        row = discounted_row + 2
        van_row = row
        tir_row = row + 1
        profitability_row = row + 2
        decision_row = row + 3
        ws.write(van_row, 0, "Valor actual neto", formats["total"])
        ws.write_formula(
            van_row,
            1,
            f"=SUM({cell(discounted_row, 1)}:{cell(discounted_row, 11)})",
            formats["total"],
            scenario_model["npv"],
        )
        ws.write(tir_row, 0, "Tasa interna de retorno", formats["total"])
        ws.write_formula(
            tir_row,
            1,
            f"=IRR({cell(valuation_row, 1)}:{cell(valuation_row, 11)})",
            formats["total_pct"],
            scenario_model["irr"],
        )
        ws.write(profitability_row, 0, "Indice de rentabilidad", formats["total"])
        ws.write_formula(
            profitability_row,
            1,
            f"=({cell(van_row, 1)}+ABS({cell(valuation_row, 1)}))/ABS({cell(valuation_row, 1)})",
            formats["total_num"],
            scenario_model["profitability_index"],
        )
        decision = (
            "VIABLE"
            if scenario_model["npv"] > 0
            and scenario_model["irr"] > scenario_model["cut_rate"]
            and scenario_model["profitability_index"] > 1
            else "NO VIABLE"
        )
        ws.write(decision_row, 0, "Resultado", formats["subtitle"])
        ws.write_formula(
            decision_row,
            1,
            f'=IF(AND({cell(van_row, 1)}>0,{cell(tir_row, 1)}>{sheet_cell("COSTO CAPITAL VAN TIR IR", 10, 4)},'
            f'{cell(profitability_row, 1)}>1),"VIABLE","NO VIABLE")',
            formats["text"],
            decision,
        )
        calculation_rows.append(
            (van_row, tir_row, profitability_row, decision_row)
        )
        row = decision_row + 2

    for summary_row, calculation in zip(summary_rows, calculation_rows):
        van_row, tir_row, profitability_row, decision_row = calculation
        ws.write_formula(
            summary_row,
            3,
            f"={cell(van_row, 1)}",
            formats["money"],
            scenario_models[summary_rows.index(summary_row)]["npv"],
        )
        ws.write_formula(
            summary_row,
            4,
            f"={cell(tir_row, 1)}",
            formats["percent"],
            scenario_models[summary_rows.index(summary_row)]["irr"],
        )
        ws.write_formula(
            summary_row,
            5,
            f"={cell(profitability_row, 1)}",
            formats["number"],
            scenario_models[summary_rows.index(summary_row)]["profitability_index"],
        )
        ws.write_formula(
            summary_row,
            6,
            f"={cell(decision_row, 1)}",
            formats["text"],
            "VIABLE"
            if scenario_models[summary_rows.index(summary_row)]["npv"] > 0
            and scenario_models[summary_rows.index(summary_row)]["irr"]
            > scenario_models[summary_rows.index(summary_row)]["cut_rate"]
            and scenario_models[summary_rows.index(summary_row)]["profitability_index"] > 1
            else "NO VIABLE",
        )

    ws.write(3, 6, "Decision", formats["header"])
    ws.merge_range(
        row,
        0,
        row + 1,
        11,
        "El escenario de precio -10% mantiene VAN positivo, TIR superior a la tasa de corte e indice de rentabilidad mayor que uno, pero presenta un deficit temporal de caja cercano a S/111,946 que exige capital de trabajo o anticipos contractuales. Una reduccion de 15% vuelve inviable el proyecto.",
        formats["note"],
    )


def replace_iferror(formula: str) -> str:
    """Translate IFERROR to a BIFF8-compatible equivalent for xlwt."""
    marker = re.compile(r"IFERROR\(", re.IGNORECASE)
    while match := marker.search(formula):
        args_start = match.end()
        depth = 0
        comma = None
        in_string = False
        index = args_start
        while index < len(formula):
            char = formula[index]
            if char == '"':
                if in_string and index + 1 < len(formula) and formula[index + 1] == '"':
                    index += 2
                    continue
                in_string = not in_string
            elif not in_string:
                if char == "(":
                    depth += 1
                elif char == ")":
                    if depth == 0:
                        break
                    depth -= 1
                elif char == "," and depth == 0 and comma is None:
                    comma = index
            index += 1
        if comma is None or index >= len(formula):
            raise ValueError(f"Formula IFERROR no valida: {formula}")
        expression = formula[args_start:comma]
        fallback = formula[comma + 1:index]
        replacement = f"IF(ISERROR({expression}),{fallback},{expression})"
        formula = formula[:match.start()] + replacement + formula[index + 1:]
    return formula


def xls_color_index(color, palette: dict[str, int], workbook: xlwt.Workbook) -> int:
    if color is None:
        return 0x7FFF
    rgb = None
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:].upper()
    elif color.type == "indexed" and color.indexed is not None:
        indexed = int(color.indexed)
        if 0 <= indexed < len(COLOR_INDEX):
            rgb = COLOR_INDEX[indexed][-6:].upper()
    if not rgb:
        return 0x7FFF
    standard = {
        "000000": xlwt.Style.colour_map["black"],
        "FFFFFF": xlwt.Style.colour_map["white"],
        "FF0000": xlwt.Style.colour_map["red"],
        "008080": xlwt.Style.colour_map["teal"],
    }
    if rgb in standard:
        return standard[rgb]
    if rgb not in palette:
        colour_index = 40 + len(palette)
        if colour_index > 63:
            return xlwt.Style.colour_map["black"]
        palette[rgb] = colour_index
        workbook.set_colour_RGB(
            colour_index,
            int(rgb[0:2], 16),
            int(rgb[2:4], 16),
            int(rgb[4:6], 16),
        )
    return palette[rgb]


def xls_style_from_cell(cell, palette: dict[str, int], workbook: xlwt.Workbook) -> xlwt.XFStyle:
    style = xlwt.XFStyle()

    font = xlwt.Font()
    font.name = cell.font.name or "Arial"
    font.height = int((cell.font.sz or 10) * 20)
    font.bold = bool(cell.font.bold)
    font.italic = bool(cell.font.italic)
    font.underline = xlwt.Font.UNDERLINE_SINGLE if cell.font.underline else xlwt.Font.UNDERLINE_NONE
    font.colour_index = xls_color_index(cell.font.color, palette, workbook)
    style.font = font

    pattern = xlwt.Pattern()
    if cell.fill.fill_type == "solid":
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xls_color_index(cell.fill.fgColor, palette, workbook)
    style.pattern = pattern

    alignment = xlwt.Alignment()
    alignment.horz = {
        "left": xlwt.Alignment.HORZ_LEFT,
        "center": xlwt.Alignment.HORZ_CENTER,
        "right": xlwt.Alignment.HORZ_RIGHT,
        "justify": xlwt.Alignment.HORZ_JUSTIFIED,
    }.get(cell.alignment.horizontal, xlwt.Alignment.HORZ_GENERAL)
    alignment.vert = {
        "top": xlwt.Alignment.VERT_TOP,
        "center": xlwt.Alignment.VERT_CENTER,
        "bottom": xlwt.Alignment.VERT_BOTTOM,
    }.get(cell.alignment.vertical, xlwt.Alignment.VERT_CENTER)
    alignment.wrap = int(bool(cell.alignment.wrap_text))
    style.alignment = alignment

    borders = xlwt.Borders()
    border_styles = {
        None: xlwt.Borders.NO_LINE,
        "thin": xlwt.Borders.THIN,
        "medium": xlwt.Borders.MEDIUM,
        "thick": xlwt.Borders.THICK,
        "double": xlwt.Borders.DOUBLE,
        "dashed": xlwt.Borders.DASHED,
        "dotted": xlwt.Borders.DOTTED,
    }
    for side_name in ["left", "right", "top", "bottom"]:
        source_side = getattr(cell.border, side_name)
        setattr(borders, side_name, border_styles.get(source_side.style, xlwt.Borders.THIN))
        setattr(
            borders,
            f"{side_name}_colour",
            xls_color_index(source_side.color, palette, workbook),
        )
    style.borders = borders
    style.num_format_str = cell.number_format or "General"
    return style


def convert_xlsx_to_xls(source: Path, target: Path) -> None:
    source_workbook = load_workbook(source, data_only=False)
    target_workbook = xlwt.Workbook(encoding="utf-8")
    palette: dict[str, int] = {}
    styles: dict[int, xlwt.XFStyle] = {}
    target_sheets = {
        source_sheet.title: target_workbook.add_sheet(
            source_sheet.title,
            cell_overwrite_ok=True,
        )
        for source_sheet in source_workbook.worksheets
    }

    for source_sheet in source_workbook.worksheets:
        target_sheet = target_sheets[source_sheet.title]
        target_sheet.set_portrait(False)
        target_sheet.set_paper_size_code(9)
        target_sheet.set_fit_width_to_pages(1)
        target_sheet.set_fit_height_to_pages(0)

        for column_letter, dimension in source_sheet.column_dimensions.items():
            column_index = source_sheet[column_letter + "1"].column - 1
            if dimension.width:
                target_sheet.col(column_index).width = min(65535, int(dimension.width * 256))
            target_sheet.col(column_index).hidden = bool(dimension.hidden)
        for row_index, dimension in source_sheet.row_dimensions.items():
            if dimension.height:
                target_row = target_sheet.row(row_index - 1)
                target_row.height = int(dimension.height * 20)
                target_row.height_mismatch = True
            target_sheet.row(row_index - 1).hidden = bool(dimension.hidden)

        for row in source_sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                style = styles.get(cell.style_id)
                if style is None:
                    style = xls_style_from_cell(cell, palette, target_workbook)
                    styles[cell.style_id] = style
                value = cell.value
                if cell.data_type == "f":
                    formula = replace_iferror(value[1:] if value.startswith("=") else value)
                    value = xlwt.Formula(formula)
                target_sheet.write(cell.row - 1, cell.column - 1, value, style)

        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge(
                merged_range.min_row - 1,
                merged_range.max_row - 1,
                merged_range.min_col - 1,
                merged_range.max_col - 1,
            )

        if source_sheet.freeze_panes:
            frozen = source_sheet.freeze_panes
            frozen_cell = source_sheet[frozen] if isinstance(frozen, str) else frozen
            target_sheet.set_panes_frozen(True)
            target_sheet.set_horz_split_pos(frozen_cell.row - 1)
            target_sheet.set_vert_split_pos(frozen_cell.column - 1)

    target_workbook.save(target)


def write_all() -> None:
    base = build_model()
    low_income = build_model(name="Ingresos -20%", revenue_factor=0.80)
    high_cost = build_model(name="Gastos +10%", cost_factor=1.10)
    high_interest = build_model(name="Interes 24%", interest_rate=LOAN_RATE_STRESS)
    bad_debt = build_model(name="No recuperacion CxC 20%", bad_debt_rate=0.20)
    admin_risk = build_model(name="Admin +15%", admin_factor=1.15)
    optimistic = build_model(name="Optimista", revenue_factor=1.12, cost_factor=1.03)
    reduced_price = build_model(name="Precio -10%", revenue_factor=0.90)

    workbook = xlsxwriter.Workbook(OUT)
    formats = make_formats(workbook)
    write_ingresos(workbook, base, formats)
    write_costos(workbook, base, formats)
    write_unit_cost(workbook, base, formats)
    write_planilla(workbook, base, formats)
    write_expenses_finance(workbook, base, formats)
    write_depreciation(workbook, base, formats)
    write_investment_plan(workbook, base, formats)
    write_opening_balance(workbook, base, formats)
    write_cost_of_sales(workbook, base, formats)
    write_income_statement(workbook, base, formats)
    write_cash_budget(workbook, base, formats)
    write_balance(workbook, base, formats)
    write_ratios(workbook, base, formats)
    write_ratio_comments(workbook, base, formats)
    write_break_even(workbook, base, formats)
    write_flow_sheet(workbook, base, formats, "FLUJO DE EFECTIVO", "Flujo Neto de Efectivo - Original")
    write_cost_capital(workbook, base, formats)
    write_flow_sheet(
        workbook,
        low_income,
        formats,
        "Sensibilidad Ingresos Bajos",
        "Flujo Neto Pesimista - Ingresos -20%",
        scenario_label="Reduccion de ingresos",
        scenario_value=-0.20,
        scenario_kind="low_income",
        base_model=base,
    )
    write_flow_sheet(
        workbook,
        high_cost,
        formats,
        "Sensibilidad Incremento Gastos",
        "Flujo Neto Pesimista - Gastos +10%",
        scenario_label="Incremento de costos y gastos",
        scenario_value=0.10,
        scenario_kind="high_cost",
        base_model=base,
    )
    write_flow_sheet(
        workbook,
        high_interest,
        formats,
        "Sensibilidad Incremento Interes",
        "Flujo Neto Pesimista - Interes 24%",
        scenario_label="Tasa de interes del financiamiento",
        scenario_value=LOAN_RATE_STRESS,
        scenario_kind="high_interest",
        base_model=base,
    )
    write_flow_sheet(
        workbook,
        bad_debt,
        formats,
        "Riesgo no recuperacion CxC",
        "Riesgo de no recuperacion de Cuentas por Cobrar",
        scenario_label="Cuentas por cobrar no recuperadas",
        scenario_value=0.20,
        scenario_kind="bad_debt",
        base_model=base,
    )
    write_risk_admin(workbook, base, admin_risk, formats)
    write_flow_sheet(
        workbook,
        optimistic,
        formats,
        "Flujo Optimista",
        "Flujo Neto de Efectivo - Optimista",
        scenario_label="Incremento de ingresos",
        scenario_value=0.12,
        scenario_kind="optimistic",
        base_model=base,
        secondary_scenario_label="Incremento de costos",
        secondary_scenario_value=0.03,
    )
    write_price_analysis(workbook, base, formats)
    write_cash_budget(
        workbook,
        reduced_price,
        formats,
        sheet_name="Presupuesto Caja Bajando Precio",
        base_model=base,
        scenario_value=-0.10,
    )
    write_flow_sheet(
        workbook,
        reduced_price,
        formats,
        "Flujo Efectivo Bajando Precio",
        "Flujo Neto - Bajando Precio 10%",
        scenario_label="Reduccion de precio",
        scenario_value=-0.10,
        scenario_kind="reduced_price",
        base_model=base,
    )
    workbook.close()
    convert_xlsx_to_xls(OUT, OUT_XLS)

    print(f"Archivo generado: {OUT}")
    print(f"Archivo generado: {OUT_XLS}")
    print(f"VAN base: S/ {base['npv']:,.2f}")
    print(f"TIR base: {base['irr']:.2%}")
    print(f"Tasa de corte: {base['cut_rate']:.2%}")
    print(f"Saldo minimo de caja: S/ {min(base['cash_balance']):,.2f}")


if __name__ == "__main__":
    write_all()
